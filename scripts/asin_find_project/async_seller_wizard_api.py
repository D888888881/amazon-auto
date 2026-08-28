import asyncio
import json
import os
import re
import shutil
from pathlib import Path

import aiohttp
import pandas as pd
import numpy as np
from async_read_config import read_main, read_taobao_config
import async_sif_api
from async_advertisement_api import (
    advertisement_main,
    fetch_multiple_asins,
    fetch_multiple_asins_totalUnits,
    ads_cache_get,
    ensure_ads_cached,
    ensure_ads_cached_robust,
)
from typing import List, Dict, Any, Optional
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment
from async_image_search_api import async_price_info_main
from async_fba_api import async_fba_batch
from async_return_rale_api import fetch_refund_rate_for_path, prefetch_refund_rates_batch
from seller_account_guard import is_seller_account_banned_error
from shared_rate_limit import async_api_slot, env_max_concurrent
from taobao__m_h5_tk import get_taobao_tokens
from wizard_progress import emit_progress
# 本地数据根目录：相对本脚本所在目录的 file/，避免「在别的 cwd 运行脚本」时扫不到 Excel
# FILE_DATA_ROOT = Path(r"E:\py_projiect\auto_amazon_project\media\file")
SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = SCRIPT_DIR.parent.parent
FILE_DATA_ROOT = PROJECT_ROOT / "media" / "file"
# 旧主图目录（仅读取兼容）；新图写入 media/images 或 ASIN_IMAGES_ROOT
_LEGACY_IMAGES_DIR = SCRIPT_DIR / "images"


def resolve_asin_images_dir() -> Path:
    """ASIN 主图目录：优先环境变量 / Django settings，默认 media/images。"""
    env = (os.environ.get("ASIN_IMAGES_ROOT") or "").strip()
    if env:
        return Path(env).expanduser().resolve()
    try:
        from django.conf import settings

        if getattr(settings, "configured", False):
            root = getattr(settings, "ASIN_IMAGES_ROOT", None)
            if root:
                return Path(root).resolve()
            media = getattr(settings, "MEDIA_ROOT", None)
            if media:
                return (Path(media) / "images").resolve()
    except Exception:
        pass
    return (PROJECT_ROOT / "media" / "images").resolve()


# 兼容旧引用名；运行时请用 resolve_asin_images_dir()，避免 import 时 Django 未就绪
IMAGES_DIR = PROJECT_ROOT / "media" / "images"

# SIF 无 ASIN 数据时的广告指标默认值
SIF_DEFAULT_AD_CPC = 1.0
SIF_DEFAULT_AD_CLICKS = 100.0
SIF_DEFAULT_CONVERSION_RATE = 0.10  # 10%

# 其它接口缺数据时的 ROI 默认值（仍生成 ROI-US-pack.xlsx）
ROI_DEFAULT_UNIT_PURCHASE = 10.0
ROI_DEFAULT_FBA_FEE = 5.0
ROI_DEFAULT_REFUND_RATE = 10.0
ROI_DEFAULT_PRODUCT_PRICE = 29.99


def _positive_float_or_none(value) -> float | None:
    try:
        num = float(value)
        if num > 0:
            return num
    except (TypeError, ValueError):
        pass
    return None


def resolve_sif_ad_metrics(
    product_asin: str,
    asin_cpc_list: list,
    daily_orders: float,
) -> tuple[float, float, float]:
    """
    从 SIF 聚合结果解析广告 cpc / 转化率 / 点击数。
    字段为空、None 或 <=0 时使用默认值，避免 ROI-US-pack 无法生成。
    """
    asin_cpc_dict: dict[str, dict] = {}
    cpc_items = asin_cpc_list if isinstance(asin_cpc_list, list) else []
    for item in cpc_items:
        if not isinstance(item, dict):
            continue
        for asin_key, info in item.items():
            if not isinstance(info, dict):
                continue
            key = str(asin_key).strip().upper()
            cpc_info = info.get('cpc') or {}
            cpc_median = cpc_info.get('median') if isinstance(cpc_info, dict) else None
            asin_cpc_dict[key] = {
                'cpc_median': cpc_median,
                'clickPurchaseRatio': info.get('clickPurchaseRatio'),
            }

    product_key = str(product_asin).strip().upper()
    cpc_info = asin_cpc_dict.get(product_key, {})
    used_defaults = False

    ad_cpc = _positive_float_or_none(cpc_info.get('cpc_median'))
    if ad_cpc is None:
        ad_cpc = SIF_DEFAULT_AD_CPC
        used_defaults = True

    # SIF 偶发返回百分数（如 10 表示 10%），统一成 0~1 小数
    conversion_rate = _positive_float_or_none(cpc_info.get('clickPurchaseRatio'))
    if conversion_rate is not None and conversion_rate > 1:
        conversion_rate = conversion_rate / 100.0
    if conversion_rate is None:
        conversion_rate = SIF_DEFAULT_CONVERSION_RATE
        used_defaults = True

    if daily_orders > 0 and conversion_rate > 0:
        # 总流量中广告占比 7、自然流量占比 3
        ad_clicks = daily_orders / conversion_rate * 0.7
    else:
        ad_clicks = None
    if ad_clicks is None or ad_clicks <= 0:
        ad_clicks = SIF_DEFAULT_AD_CLICKS
        used_defaults = True

    if used_defaults:
        print(
            f'ASIN {product_key} SIF 广告数据缺失，使用默认值：'
            f'cpc={ad_cpc}, 点击={ad_clicks}, 转化率={conversion_rate * 100:.0f}%'
        )

    return ad_cpc, conversion_rate, ad_clicks

# 卖家精灵导出 Excel 列名 -> 内部 API 风格字段（与 colum_mapping.json 中译名对齐）
COLUMN_MAPPING_EXCEL_TO_API = {
    "ASIN": "asin",
    "父ASIN": "parent",
    "SKU": "sku",
    "品牌": "brand",
    "品牌链接": "brandUrl",
    "搜索排名": "lqs",
    "商品标题": "title",
    "商品详情页链接": "asinUrl",
    "商品主图": "imageUrl",
    "类目路径": "nodeLabelPath",
    "节点标签路径": "nodeLabelPath",
    "节点路径": "nodeLabelPath",
    "大类目": "categoryName",
    "大类BSR": "bsrRank",
    "小类目": "bsrLabel",
    "小类BSR": "subcategories",
    "月销量": "totalUnits",
    "月销量增长率": "totalUnitsGrowth",
    "月销售额($)": "totalAmount",
    "子体销量": "fbaUnits",
    "子体销售额($)": "fbaAmount",
    "变体数": "variations",
    "价格($)": "price",
    "Prime价格($)": "primeExclusivePrice",
    "Coupon": "coupon",
    "Q&A数": "questions",
    "评分数": "reviews",
    "月新增评分数": "reviewsIncreasement",
    "评分": "rating",
    "留评率": "reviewsRate",
    "FBA($)": "fba",
    "毛利率": "profit",
    "上架时间": "availableDate",
    "上架天数": "availableDays",
    "LQS": "lqs",
    "卖家数": "sellers",
    "Best Seller标识": "bestSeller",
    "Amazon's Choice": "amazonChoice",
    "New Release标识": "newRelease",
    "商品重量": "weight",
    "商品尺寸": "dimensions",
    "包装重量": "pkgWeight",
    "包装尺寸": "pkgDimensions",
}


with open('config_file/colum_mapping.json', 'r', encoding='utf-8') as f:
    column_mapping = json.load(f)


# 验证/生成保存目录：优先使用显式 ASIN（与 file/{ASIN} 结构一致）
async def verify_path(
        keyword_dict: Optional[Dict[str, Any]] = None,
        keyword: Optional[str] = None,
        asin: Optional[str] = None,
) -> str:
    """
    返回文件保存的目录路径。

    规则：
    - 同时提供 asin 和 keyword：返回 FILE_DATA_ROOT / asin / keyword
    - 只提供 asin，不提供 keyword：返回 FILE_DATA_ROOT / asin
    - 只提供 keyword，且提供 keyword_dict：尝试匹配 ASIN，返回 FILE_DATA_ROOT / asin / keyword
    - 其他情况：返回默认目录 FILE_DATA_ROOT
    """
    base = Path(FILE_DATA_ROOT)

    # 1. 同时有 asin 和 keyword（最优先）
    if asin and keyword:
        path = base / asin / keyword

    # 2. 只有 asin，没有 keyword
    elif asin and not keyword:
        path = base / asin

    # 3. 只有 keyword，但提供了 keyword_dict 用于匹配 asin
    elif keyword and keyword_dict:
        matched_asin = None
        for key, values in keyword_dict.items():
            norm = values if isinstance(values, (list, tuple)) else [values]
            if keyword in norm:
                matched_asin = key
                break
        if matched_asin:
            path = base / matched_asin / keyword
        else:
            print(f"警告：关键词 '{keyword}' 未在 keyword_dict 中匹配到 ASIN，保存到默认目录。")
            path = base

    # 4. 其他情况
    else:
        path = base

    # 自动创建目录
    path.mkdir(parents=True, exist_ok=True)
    return str(path)


def _extract_asin_reference_price(
    asin: str,
    asin_info_dict: dict | None,
    df: pd.DataFrame | None = None,
) -> float | None:
    """解析当前 ASIN 参考价；优先批量广告数据，其次 Search 表内本 ASIN 价格。"""
    a = str(asin or '').strip().upper()
    if asin_info_dict and isinstance(asin_info_dict, dict):
        raw = asin_info_dict.get('avg_price', asin_info_dict.get('price'))
        if raw is None and a in asin_info_dict and isinstance(asin_info_dict[a], dict):
            raw = asin_info_dict[a].get('avg_price', asin_info_dict[a].get('price'))
        try:
            p = float(raw)
            if p > 0:
                return p
        except (TypeError, ValueError):
            pass
    if df is not None and not df.empty and 'asin' in df.columns and 'price' in df.columns:
        sub = df.copy()
        sub['_asin_norm'] = sub['asin'].astype(str).str.strip().str.upper()
        row = sub[sub['_asin_norm'] == a]
        if not row.empty:
            p = pd.to_numeric(row.iloc[0]['price'], errors='coerce')
            if not pd.isna(p) and float(p) > 0:
                return float(p)
    return None


def _load_existing_data_origin(output_path: str) -> pd.DataFrame | None:
    if not os.path.isfile(output_path):
        return None
    try:
        existing = pd.read_excel(output_path)
        if existing is not None and not existing.empty:
            return existing
    except Exception as e:
        print(f'读取已有 data_origin 失败 {output_path}: {e}')
    return None


def _keyword_artifact_paths(output_dir: str, keyword: str) -> dict[str, str]:
    return {
        'data_origin': os.path.join(output_dir, f'{keyword}_data_origin.xlsx'),
        'review': os.path.join(output_dir, f'{keyword}_review_interval_analysis_US.xlsx'),
        'review_legacy': os.path.join(output_dir, f'{keyword}_review_interval_analysis.xlsx'),
        'ad_efficiency': os.path.join(output_dir, f'{keyword}_ad_efficiency_table.xlsx'),
    }


def _review_interval_file_exists(paths: dict[str, str]) -> bool:
    return os.path.isfile(paths.get('review', '')) or os.path.isfile(paths.get('review_legacy', ''))


def _read_ranking_from_ad_efficiency_xlsx(
    path: str,
    target_asin: str,
) -> float | None:
    """从已保存的广告效率表读取 ranking_percent；文件不存在或无效时返回 None。"""
    if not os.path.isfile(path):
        return None
    try:
        df = pd.read_excel(path)
        if df.empty or '链接' not in df.columns:
            return None
        work = df.copy()
        work['_asin_norm'] = work['链接'].astype(str).str.strip().str.upper()
        work = work[work['_asin_norm'].str.match(r'^B[A-Z0-9]{9}$', na=False)]
        if work.empty:
            return None
        if '广告词数量' in work.columns:
            work['广告词数量'] = pd.to_numeric(work['广告词数量'], errors='coerce')
            valid_df = work[work['广告词数量'] > 0].reset_index(drop=True)
        else:
            return None
        if valid_df.empty:
            return None
        target = str(target_asin or '').strip().upper()
        if target not in valid_df['_asin_norm'].values:
            return None
        rank = valid_df[valid_df['_asin_norm'] == target].index[0] + 1
        total_valid = len(valid_df)
        if total_valid <= 0:
            return None
        return round((rank / total_valid) * 100, 3)
    except Exception as e:
        print(f'读取广告效率表 ranking 失败 {path}: {e}')
        return None


async def _rotate_ad_account_for_keyword_resume(
    asin_key: str,
    keyword: str,
    *,
    rotation_state: dict,
) -> None:
    from seller_account_guard import (
        SellerAccountBannedError,
        bulk_rotate_if_available,
        clear_seller_login_cache,
        ensure_seller_login,
    )

    emit_progress(
        f'广告难度：关键词「{keyword}」({asin_key}) 遇禁号，切换账号并断点续算…'
    )
    clear_seller_login_cache()
    if not await bulk_rotate_if_available(
        [asin_key],
        rotation_state=rotation_state,
        pending_task='ad',
    ):
        raise SellerAccountBannedError(
            f'批量账号被禁且无法轮换（关键词 {keyword} / {asin_key}）'
        )
    await ensure_seller_login()


# 清洗脏数据
async def save_cleaned_data_orign_to_excel(
    df: pd.DataFrame,
    keyword: str,
    asin: str,
    asin_info_dict: dict = None,
    *,
    ads_cache: dict | None = None,
):
    """
    依赖字段（卖家精灵 Excel 映射后）：
    - brand / parent：去重；缺失则填空，避免 KeyError
    - price / totalUnits / reviews：清洗与分箱；缺失则转为 NaN
    - reviews：评论区间、sales_level
    - availableDate 等：导出多为日期字符串，非毫秒时间戳
    """

    df_deduplicated = df.copy()
    if "brand" not in df_deduplicated.columns:
        df_deduplicated["brand"] = ""
    if "parent" not in df_deduplicated.columns:
        df_deduplicated["parent"] = (
            df_deduplicated["asin"] if "asin" in df_deduplicated.columns else ""
        )
    for col in ("price", "totalUnits", "reviews"):
        if col not in df_deduplicated.columns:
            df_deduplicated[col] = np.nan
    df_deduplicated["brand"] = df_deduplicated["brand"].fillna("")
    df_deduplicated["parent"] = df_deduplicated["parent"].fillna("")
    df_deduplicated["totalUnits"] = pd.to_numeric(df_deduplicated["totalUnits"], errors="coerce")
    # df_deduplicated["reviews"] = pd.to_numeric(df_deduplicated["reviews"], errors="coerce")

    # 去重
    df_deduplicated = df_deduplicated.drop_duplicates(subset=["brand", "parent","asin"]).copy()

    # 去除广告位
    before_count = len(df_deduplicated)
    mask = ~df_deduplicated['lqs'].astype(str).str.contains("广告位", na=False)
    df_deduplicated = df_deduplicated[mask].copy()
    after_count = len(df_deduplicated)
    print(f"已删除搜索排名包含'广告位'的行，删除前 {before_count} 行，剩余 {after_count} 行")

    # 价格清洗：基于当前 ASIN 价格过滤
    df_deduplicated['price'] = pd.to_numeric(df_deduplicated['price'], errors='coerce')
    df_deduplicated['totalUnits'] = pd.to_numeric(df_deduplicated['totalUnits'], errors='coerce')

    price_cleaning = df_deduplicated.dropna(subset=['price', 'totalUnits']).copy()
    asin_price = _extract_asin_reference_price(asin, asin_info_dict, df_deduplicated)
    if asin_price is None:
        cached_row = ads_cache_get(ads_cache, asin) if ads_cache else None
        if cached_row:
            asin_price = _extract_asin_reference_price(
                asin, cached_row, df_deduplicated
            )
            if asin_price is not None:
                print(f'ASIN {asin} 价格（广告缓存）：{asin_price}')
    if asin_price is None and not asin_info_dict and not ads_cache:
        try:
            info_dict = await advertisement_main([asin])
            a_key = str(asin).strip().upper()
            row = (info_dict or {}).get(asin) or (info_dict or {}).get(a_key) or {}
            asin_price = _extract_asin_reference_price(asin, row if isinstance(row, dict) else None, df_deduplicated)
            if asin_price is not None:
                print(f'ASIN {asin} 价格（广告接口）：{asin_price}')
        except Exception as e:
            if is_seller_account_banned_error(e):
                raise
            print(f'价格获取失败（将跳过价格过滤，避免清空 data_origin）：{e}')

    if asin_price is not None and asin_price > 0:
        upper_bound = asin_price * 1.3
        lower_bound = asin_price * 0.6
        filtered = price_cleaning[
            (price_cleaning['price'] >= lower_bound) & (price_cleaning['price'] <= upper_bound)
        ].copy()
        if not filtered.empty:
            price_cleaning_data = filtered
        else:
            print(
                f'警告: ASIN {asin} 关键词「{keyword}」价格过滤后无数据，'
                f'保留过滤前 {len(price_cleaning)} 行'
            )
            price_cleaning_data = price_cleaning.copy()
    else:
        print(
            f'警告: ASIN {asin} 关键词「{keyword}」未获取有效参考价，'
            f'跳过价格过滤（保留 {len(price_cleaning)} 行）'
        )
        price_cleaning_data = price_cleaning.copy()
    print(f'价格过滤后数据行数：{len(price_cleaning_data)}')
    if not price_cleaning_data.empty:
        print(price_cleaning_data['price'].describe())

    # # 删除无用列
    # columns_to_drop = [
    #     'New Release标识', 'A+页面', '视频介绍', 'SP广告', '品牌故事', '品牌广告', '7天促销', 'Best Seller标识',
    #     "Amazon's Choice", 'CPF绿标', '评级'
    # ]
    # price_cleaning_data.drop(columns=columns_to_drop, inplace=True, errors='ignore')
    print(f"清洗后的数据行数：{len(price_cleaning_data)}")

    output_dir = await verify_path(asin=asin, keyword=keyword)
    os.makedirs(output_dir, exist_ok=True)
    output_path = os.path.join(output_dir, f'{keyword}_data_origin.xlsx')

    if price_cleaning_data.empty:
        existing = _load_existing_data_origin(output_path)
        if existing is not None:
            print(f'警告: 清洗结果为空，保留已有 data_origin: {output_path}')
            return existing
        fallback = price_cleaning.copy()
        if fallback.empty:
            fallback = df_deduplicated.copy()
        if fallback.empty:
            raise RuntimeError(
                f'ASIN {asin} 关键词「{keyword}」清洗后无有效数据，已跳过写入 data_origin'
            )
        print(
            f'警告: ASIN {asin} 关键词「{keyword}」清洗结果为空，'
            f'使用去广告位后的 {len(fallback)} 行作为 data_origin'
        )
        price_cleaning_data = fallback

    # 时间字段：API 为毫秒时间戳；Excel 多为日期或字符串
    time_fields = ["syncTime", "amzUnitDate", "updatedTime", "availableDate", "firstReviewDate"]
    for field in time_fields:
        if field not in price_cleaning_data.columns:
            continue
        col = price_cleaning_data[field]
        try:
            num = pd.to_numeric(col, errors="coerce")
            if num.notna().any() and (num.dropna() > 1e11).all():
                price_cleaning_data[field] = pd.to_datetime(num, unit="ms", errors="coerce").dt.strftime(
                    "%Y-%m-%d"
                )
            else:
                price_cleaning_data[field] = pd.to_datetime(col, errors="coerce").dt.strftime("%Y-%m-%d")
        except Exception as e:
            print(f"<时间列格式化>{field}: {e}")

    # 定义 bins：最后一个值用很大的数或无穷大
    bins = [0, 30, 60, 100, 150, 200, np.inf]  # 或 1e10

    labels = ['0-30', '31-50', '51-100', '101-150', '151-200', '200以上']
    #
    # 使用 cut 进行分级
    price_cleaning_data['sales_level'] = pd.cut(price_cleaning_data['reviews'], bins=bins, labels=labels,
                                                right=True, include_lowest=True)
    # 列名中文化
    # price_cleaning_data_chinese = price_cleaning_data.rename(columns=column_mapping)

    # 保存 Excel（禁止用空表覆盖已有 data_origin，避免 Search 全表被标黄）
    price_cleaning_data.to_excel(output_path, index=False)
    print(f"文件已保存至: {output_path}")
    return price_cleaning_data


# 整理出市场容量前五
async def save_top5_market_capacity_to_excel(price_cleaning_data: pd.DataFrame, keyword: str, asin: str):
    # 1. 确保销量列为数值，并去掉缺失值
    price_cleaning_data['totalUnits'] = pd.to_numeric(price_cleaning_data['totalUnits'], errors='coerce')
    price_cleaning_data = price_cleaning_data.dropna(subset=['totalUnits'])

    # 2. 按销量降序排序，取前5名
    top5 = price_cleaning_data.sort_values('totalUnits', ascending=False).head(5).copy()
    top5.reset_index(drop=True, inplace=True)

    if top5.empty:
        print(
            f"警告: ASIN {asin} 关键词「{keyword}」清洗后无有效月销量(totalUnits)，"
            f"跳过 top5 市场容量分析，目标月销记为 0。"
        )
        return 0.0

    # 3. 初始化新列
    top5['实际增长倍数'] = np.nan
    top5['是否1.3-1.5递增减'] = ''
    top5['模拟增长数'] = top5['totalUnits'].astype(float).copy()

    # 4. 计算实际增长倍数和判断（基于真实销量，不变）
    for i in range(len(top5) - 1):
        current = top5.loc[i, 'totalUnits']
        next_val = top5.loc[i + 1, 'totalUnits']
        if next_val > 0:
            ratio = current / next_val
            top5.loc[i, '实际增长倍数'] = round(ratio, 2)
            if 1.3 <= ratio <= 1.5:
                top5.loc[i, '是否1.3-1.5递增减'] = '是'
            else:
                top5.loc[i, '是否1.3-1.5递增减'] = '否'
        else:
            top5.loc[i, '实际增长倍数'] = np.nan
            top5.loc[i, '是否1.3-1.5递增减'] = '否'
    if len(top5) > 0:
        top5.loc[len(top5) - 1, '实际增长倍数'] = np.nan
        top5.loc[len(top5) - 1, '是否1.3-1.5递增减'] = '-'

    # 5. 按新规则计算模拟增长数（以第五名 / 最后一名为基准，向上乘以1.3）
    n = len(top5)
    base_sales = top5.loc[n - 1, 'totalUnits']   # 最后一名的真实销量
    for i in range(n):
        # 第 i 名（0 为第一名）的等级 = n - 1 - i
        top5.loc[i, '模拟增长数'] = round(base_sales * (1.3 ** (n - 1 - i)), 2)

    # 6. 全局统计值（站内月销上限 = 第一名模拟增长数，目标月销 = 上限/3）
    top_actual = top5.loc[0, '模拟增长数']
    total_upper_limit = round(top_actual, 2)
    target_monthly_sales = round(total_upper_limit / 3, 2)

    # 7. 构建明细表并重命名
    detail_df = top5[['totalUnits', '模拟增长数', '实际增长倍数', '是否1.3-1.5递增减']].copy()
    detail_df.rename(columns={'totalUnits': '月销'}, inplace=True)

    # 8. 添加全局统计列（只在第一行显示）
    final_df = detail_df.copy()
    final_df['站内月销上限'] = ''
    final_df['目标月销'] = ''
    final_df.loc[0, '站内月销上限'] = total_upper_limit
    final_df.loc[0, '目标月销'] = target_monthly_sales

    output_dir = await verify_path(asin=asin, keyword=keyword)

    # 9. 确保保存目录存在
    os.makedirs(output_dir, exist_ok=True)

    # 10. 保存 Excel
    output_path = os.path.join(output_dir, f'{keyword}_top5_market_capacity.xlsx')
    final_df.to_excel(output_path, index=False)
    print(f"分析结果已保存为：{output_path}")

    # 11. 控制台输出
    print("前五名分析结果：")
    print(final_df.to_string(index=False))
    return target_monthly_sales


# 分析出区间前五数据
async def save_review_interval_analysis_to_excel(
    df: pd.DataFrame,
    keyword: str,
    asin: str,
    *,
    marketplace: str = 'US',
):
    """
    对评论进行区间分级，计算每个区间月销前五的平均值，运营难度，并保存结果。
    marketplace=US|UK 决定评论分段与输出文件名 *_review_interval_analysis_{US|UK}.xlsx
    返回: dict {'区间': list, '平均销量': list, '运营难度': list}
    """
    mp = str(marketplace or 'US').strip().upper()
    if mp not in ('US', 'UK'):
        mp = 'US'

    # 1. 确保评论列为数值，并去除缺失值
    df['reviews'] = pd.to_numeric(df['reviews'], errors='coerce')
    df = df.dropna(subset=['reviews', 'totalUnits']).copy()

    # 2. 定义评论区间（英美不同）
    if mp == 'UK':
        bins = [0, 10, 30, 50, 100, 150, np.inf]
        labels = ['0-10', '11-30', '31-50', '51-100', '101-150', '150以上']
    else:
        bins = [0, 30, 50, 100, 200, np.inf]
        labels = ['0-30', '31-50', '51-100', '101-200', '200以上']
    df['评论区间'] = pd.cut(df['reviews'], bins=bins, labels=labels, right=True, include_lowest=True)

    # 3. 计算全量数据月销量前五的平均值（作为分母）
    top5_overall_sales = df.nlargest(5, 'totalUnits')['totalUnits']
    overall_avg_top5 = top5_overall_sales.mean()

    # 4. 计算每个区间月销前五的平均值
    interval_stats = []
    for interval in labels:
        group = df[df['评论区间'] == interval]
        if len(group) == 0:
            avg_sales = 0
        else:
            top5_sales = group.nlargest(5, 'totalUnits')['totalUnits']
            avg_sales = top5_sales.mean()
        interval_stats.append({'区间': interval, '平均销量': avg_sales})

    result_df = pd.DataFrame(interval_stats)

    # 5. 计算运营难度（百分比，分母为 overall_avg_top5）
    if overall_avg_top5 != 0:
        result_df['运营难度'] = (result_df['平均销量'] / overall_avg_top5 * 100).round(3).astype(str) + '%'
    else:
        result_df['运营难度'] = '0.000%'

    # 6. 定义难度标签函数，作为注释1列
    def difficulty_label(percent_str):
        try:
            percent = float(percent_str.replace('%', ''))
        except Exception:
            return '-'
        if percent < 10:
            return '困难'
        elif percent <= 15:
            return '适中'
        else:
            return '简单'

    result_df['注释1'] = result_df['运营难度'].apply(difficulty_label)

    # 7. 调整主体部分列顺序
    result_df = result_df[['区间', '平均销量', '运营难度', '注释1']]

    # ========== 构建注释行（放在数据行之后） ==========
    comment_row = pd.DataFrame({
        '区间': [''],
        '平均销量': [''],
        '运营难度': [''],
        '注释1': [''],
        '注释2': ['和评论关系不大，非标产品居多，主要看产品'],
        '参照': ['1、低于10%=困难  2、10%-15%=适中  3、大于15%=简单']
    })

    # ========== 构建汇总行 ==========
    summary_row = pd.DataFrame({
        '区间': ['自然排名最高的五个产品的平均值 销量：'],
        '平均销量': [overall_avg_top5],
        '运营难度': [''],
        '注释1': [''],
        '注释2': [''],
        '参照': ['']
    })

    # 由于主体部分缺少注释2和参照列，concat时会自动添加，对应行填充NaN
    final_df = pd.concat([result_df, comment_row, summary_row], ignore_index=True)

    # 确保最终列顺序
    final_df = final_df[['区间', '平均销量', '运营难度', '注释1', '注释2', '参照']]

    # 保存 Excel（新写一律带站点后缀；US 可读旧无后缀文件）
    output_dir = await verify_path(asin=asin, keyword=keyword)
    os.makedirs(output_dir, exist_ok=True)
    output_path = os.path.join(output_dir, f'{keyword}_review_interval_analysis_{mp}.xlsx')
    final_df.to_excel(output_path, index=False)
    print(f"评论区间前五分析已保存至: {output_path}")

    print("\n评论区间分析结果预览：")
    print(final_df.to_string(index=False))

    # ========== 返回字典格式 ==========
    return {
        '区间': result_df['区间'].tolist(),
        '平均销量': result_df['平均销量'].tolist(),
        '运营难度': result_df['运营难度'].tolist()
    }


def _ad_asin_delay_sec() -> float:
    try:
        return max(0.0, float(os.environ.get('AD_DIFFICULTY_ASIN_DELAY_SEC', '3')))
    except (TypeError, ValueError):
        return 3.0


AD_EFFICIENCY_FAILED = -1
AD_EFFICIENCY_ZERO_AD_WORDS = -2
AD_KEYWORD_ERROR_ZERO_AD_WORDS = '广告词数量为0'


def _ad_word_count_from_ads_data(data: dict | None) -> int:
    if not isinstance(data, dict):
        return 0
    total = 0
    for key in ('ads', 'highly_rated', 'sponsor_video', 'sponsor_brand'):
        try:
            total += int(data.get(key, 0) or 0)
        except (TypeError, ValueError):
            pass
    return total


def _ads_row_has_counter_fields(data: dict | None) -> bool:
    if not isinstance(data, dict):
        return False
    return any(key in data for key in ('ads', 'highly_rated', 'sponsor_video', 'sponsor_brand'))


def _ad_keyword_failure_message(reason: str) -> str:
    mapping = {
        'zero_ad_words': AD_KEYWORD_ERROR_ZERO_AD_WORDS,
        'empty_products': '关键词无有效产品数据',
        'invalid_ranking': '广告效率排名无效',
        'no_ad_efficiency_table': '未能生成广告效率表',
    }
    return mapping.get(reason, reason or '广告难度未成功计算')


async def _fetch_target_ad_word_count(
    target_asin: str,
    *,
    ads_cache: dict | None = None,
    price_info: dict | None = None,
    max_concurrent: int | None = None,
    ads_fetch_rounds: int = 3,
) -> int | None:
    """
    拉取目标 ASIN 广告词数量。
    返回 None 表示未能确认；返回 0 表示已确认广告词数量为 0。
    """
    asin_key = str(target_asin or '').strip().upper()
    if not asin_key:
        return None

    for src in (
        price_info,
        ads_cache_get(ads_cache, asin_key) if ads_cache else None,
    ):
        if _ads_row_has_counter_fields(src):
            return _ad_word_count_from_ads_data(src)

    max_ss = max_concurrent if max_concurrent is not None else env_max_concurrent('sellersprite', 6)
    row: dict | None = None
    if ads_cache is not None:
        await ensure_ads_cached_robust(
            ads_cache,
            [asin_key],
            max_concurrent=max_ss,
            max_rounds=max(1, ads_fetch_rounds),
            pause_sec=2.0,
        )
        row = ads_cache_get(ads_cache, asin_key)
    else:
        result = await advertisement_main([asin_key], max_concurrent=max_ss)
        row = (result or {}).get(asin_key) or (result or {}).get(target_asin)
        if not isinstance(row, dict):
            row = None

    if _ads_row_has_counter_fields(row):
        return _ad_word_count_from_ads_data(row)
    return None


def _zero_ad_words_skip_payload(asin_key: str) -> dict:
    return {
        'ranking_percent': None,
        'computed_ad': False,
        'keywords': {},
        'error': AD_KEYWORD_ERROR_ZERO_AD_WORDS,
        'skip_reason': 'zero_ad_words',
    }


async def _prefilter_zero_ad_word_asins(
    nested_result: dict,
    *,
    ads_cache: dict,
    max_concurrent: int,
    ads_fetch_rounds: int = 3,
) -> tuple[dict, dict]:
    """登录后、批量预取竞品广告数据前，先剔除目标 ASIN 广告词数量为 0 的项。"""
    filtered: dict = {}
    early_out: dict = {}
    for asin, kw_map in (nested_result or {}).items():
        asin_key = str(asin).strip().upper()
        ad_words = await _fetch_target_ad_word_count(
            asin_key,
            ads_cache=ads_cache,
            max_concurrent=max_concurrent,
            ads_fetch_rounds=ads_fetch_rounds,
        )
        if ad_words == 0:
            emit_progress(f'{asin_key} 目标 ASIN 广告词数量为 0，跳过广告难度计算')
            early_out[asin_key] = _zero_ad_words_skip_payload(asin_key)
            continue
        filtered[asin] = kw_map
    return filtered, early_out


async def _enrich_total_units_from_api(
    total_units_by_asin: dict,
    asin_list: list,
    *,
    max_concurrent: int = 1,
) -> None:
    """为 clean_data 中缺失或为 0 的 ASIN 批量补拉父体销量。"""
    need: list[str] = []
    seen: set[str] = set()
    for raw in asin_list:
        key = str(raw or '').strip().upper()
        if not key or key in seen:
            continue
        seen.add(key)
        val = total_units_by_asin.get(key)
        try:
            missing = val is None or pd.isna(val) or float(val) <= 0
        except (TypeError, ValueError):
            missing = True
        if missing:
            need.append(key)
    if not need:
        return
    try:
        fetched = await fetch_multiple_asins_totalUnits(need, max_concurrent=max_concurrent)
    except Exception as e:
        if is_seller_account_banned_error(e):
            raise
        print(f'警告：批量补拉父体销量失败: {e}')
        return
    for asin, data in (fetched or {}).items():
        key = str(asin).strip().upper()
        tu = pd.to_numeric((data or {}).get('totalUnits'), errors='coerce')
        if not pd.isna(tu):
            total_units_by_asin[key] = float(tu)


async def save_ad_efficiency_table(
        clean_data: pd.DataFrame,
        keyword: str,
        target_asin: str,
        *,
        ads_cache: dict | None = None,
        ads_max_concurrent: int | None = None,
        ads_fetch_rounds: int = 3,
):
    """
    保存广告效率表：
    - 广告词数量/图片/均价/均评：来自 advertisement_main
    - 父体销量 totalUnits：优先来自 clean_data（按 asin 匹配/聚合）；
      若 target_asin 在 clean_data 中缺失，则回退到 fetch_multiple_asins_totalUnits 获取。
    将排名注释放在右侧区域顶部。
    target_asin：当前分析对应的 ASIN（与 file/{ASIN} 一致）。
    """
    # ==================== 1. 确定当前 ASIN ====================
    current_asin = target_asin.strip().upper() if target_asin else None

    # ==================== 2. 获取 ASIN 列表 ====================
    asin_list = clean_data['asin'].dropna().unique().tolist()
    if current_asin and current_asin not in asin_list:
        print(f"警告：当前 ASIN {current_asin} 不在清洗数据中，将临时添加以便获取广告词数据。")
        asin_list.append(current_asin)

    # ==================== 3. 获取广告词数据（不含 totalUnits） ====================
    max_ss = ads_max_concurrent if ads_max_concurrent is not None else env_max_concurrent('sellersprite', 6)
    try:
        if ads_cache is not None:
            await ensure_ads_cached_robust(
                ads_cache,
                [str(a).strip().upper() for a in asin_list],
                max_concurrent=max_ss,
                max_rounds=max(1, ads_fetch_rounds),
                pause_sec=2.0,
            )
            ads_result = {}
            for a in asin_list:
                row = ads_cache_get(ads_cache, a)
                if row:
                    ads_result[str(a).strip().upper()] = row
                    ads_result[a] = row
        else:
            ads_result = await advertisement_main(asin_list, max_concurrent=max_ss)
        print("广告词数据示例：", list(ads_result.items())[:3])
    except Exception as e:
        if is_seller_account_banned_error(e):
            raise
        print(f"获取广告词数据失败：{e}")
        ads_result = {}

    if current_asin:
        target_row = (
            ads_result.get(current_asin)
            or ads_result.get(str(current_asin).upper())
            or (ads_cache_get(ads_cache, current_asin) if ads_cache else None)
        )
        if _ads_row_has_counter_fields(target_row):
            if _ad_word_count_from_ads_data(target_row) == 0:
                print(f'目标 ASIN {current_asin} 广告词数量为 0，跳过广告效率表')
                return AD_EFFICIENCY_ZERO_AD_WORDS

    # ==================== 4. 从 clean_data 构建 asin -> totalUnits 映射 ====================
    total_units_by_asin = {}
    if "asin" in clean_data.columns and "totalUnits" in clean_data.columns:
        tmp = clean_data[["asin", "totalUnits"]].copy()
        tmp["asin"] = tmp["asin"].astype(str).str.strip().str.upper()
        tmp["totalUnits"] = pd.to_numeric(tmp["totalUnits"], errors="coerce")
        # 同一个 asin 可能多行：取最大值更稳妥（避免某行缺失导致取到 0/NaN）
        total_units_by_asin = (
            tmp.dropna(subset=["asin"])
            .groupby("asin")["totalUnits"]
            .max()
            .to_dict()
        )

    await _enrich_total_units_from_api(
        total_units_by_asin,
        asin_list,
        max_concurrent=1 if max_ss <= 1 else min(max_ss, 3),
    )

    # 兼容旧逻辑：目标 ASIN 仍单独尝试一次（上面批量已覆盖）
    need_fetch_target_units = False
    if current_asin:
        v = total_units_by_asin.get(current_asin)
        try:
            need_fetch_target_units = v is None or pd.isna(v) or float(v) <= 0
        except (TypeError, ValueError):
            need_fetch_target_units = True

    if need_fetch_target_units:
        try:
            target_asin_total_units = await fetch_multiple_asins_totalUnits([current_asin], 1)
            fetched = (target_asin_total_units or {}).get(current_asin, {}).get("totalUnits")
            fetched = pd.to_numeric(fetched, errors="coerce")
            if not pd.isna(fetched) and float(fetched) > 0:
                total_units_by_asin[current_asin] = float(fetched)
                print(f"target_asin={current_asin} 的 totalUnits 回退获取成功: {fetched}")
            else:
                print(f"警告：target_asin={current_asin} 回退接口未返回有效 totalUnits")
        except Exception as e:
            if is_seller_account_banned_error(e):
                raise
            print(f"警告：回退获取 target_asin totalUnits 失败: {e}")

    # ==================== 5. 构建 DataFrame（广告字段来自 ads_result） ====================
    rows = []
    total_units1 = -1
    ad_words1 = -1
    for asin in asin_list:
        data = ads_result.get(asin)
        if not data:
            continue
        asin_norm = str(asin).strip().upper()
        image_url = data.get("imageUrl", "")
        price = data.get("avg_price", data.get("price", 0))
        reviews = data.get("avg_reviews", data.get("reviews", 0))
        total_units = total_units_by_asin.get(asin_norm, np.nan)
        ad_words = sum(data.get(k, 0) for k in ['ads', 'highly_rated', 'sponsor_video', 'sponsor_brand'])
        if asin == target_asin:
            total_units1 = total_units
            ad_words1 = ad_words
        ad_efficiency = total_units / ad_words if ad_words > 0 else np.nan

        rows.append({
            'imageUrl': image_url,
            'asin': asin,
            '广告词数量': ad_words,
            'totalUnits': total_units,
            'reviews': reviews,
            'price': price,
            '广告效率': ad_efficiency
        })

    if not rows:
        print("警告：没有有效的广告词数据，无法生成广告效率表")
        return -1

    result_df = pd.DataFrame(rows)
    result_df = result_df.sort_values('广告效率', ascending=False, na_position='last').reset_index(drop=True)

    # ==================== 6. 计算排名率 ====================
    ranking_note = ""
    ranking_percent = -1
    if current_asin:
        valid_df = result_df[result_df["广告词数量"] > 0].reset_index(drop=True)
        valid_df["_asin_norm"] = valid_df["asin"].astype(str).str.strip().str.upper()
        if current_asin in valid_df["_asin_norm"].values:
            rank = valid_df[valid_df["_asin_norm"] == current_asin].index[0] + 1
            total_valid = len(valid_df)
            ranking_percent = (rank / total_valid) * 100 if total_valid > 0 else 0

            if ranking_percent <= 70:
                ranking_note = f"当前产品在排序中处于{ranking_percent:.1f}%的位置，排名小于70%，相对好运营"
            else:
                ranking_note = f"当前产品在排序中处于{ranking_percent:.1f}%的位置，排名大于等于70%，运营难度较大"
        else:
            ranking_note = "当前产品广告词数量为0，无法计算广告效率排名"

    # ==================== 7. 构建结果 DataFrame ====================
    result_df = result_df[['imageUrl', 'asin', '广告词数量', 'totalUnits', 'reviews', 'price', '广告效率']].copy()
    result_df.rename(columns={
        'imageUrl': '图片',
        'asin': '链接',
        'totalUnits': '父体销量',
        'reviews': '评论数量',
        'price': '价格'
    }, inplace=True)

    # ==================== 8. 添加说明行（仅方法说明，不含排名注释） ====================
    notes = [
        "具体方法:",
        "1、找到最精准的关键词(建议用H10的Cerebro ASIN版),进去搜索结果页面",
        "2、把所有的产品打开,一个个去统计他们的父体销量,广告词数量,然后拿父体销量/广告词数量,得到广告效率,最后进行排序",
        "3、如果我们想做的这个产品在排序中处于70%以上,说明相对好运营"
    ]
    note_rows = []
    for note in notes:
        row = {col: '' for col in result_df.columns}
        row['图片'] = note
        note_rows.append(row)
    notes_df = pd.DataFrame(note_rows)
    final_df = pd.concat([result_df, notes_df], ignore_index=True)

    # ==================== 9. 直接保存 URL 文本（不下载图片） ====================
    output_dir = await verify_path(asin=target_asin, keyword=keyword)
    os.makedirs(output_dir, exist_ok=True)
    output_path = os.path.join(output_dir, f'{keyword}_ad_efficiency_table.xlsx')
    final_df.to_excel(output_path, index=False, engine='openpyxl')
    wb = load_workbook(output_path)
    ws = wb.active

    # ==================== 10. 高亮当前 ASIN 的链接 ====================
    if current_asin:
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=2)  # 链接列是第二列
            cv = str(cell.value).strip().upper() if cell.value is not None else ""
            if cv == current_asin:
                cell.fill = yellow_fill

    # ==================== 11. 将排名注释添加到右侧区域顶部 ====================
    if ranking_note:
        right_cols = [8, 9, 10]
        target_row = 5
        # 清空第5行右侧三列原有内容（可选）
        for col in right_cols:
            ws.cell(row=target_row, column=col).value = ''
        # 在右侧第一列写入注释
        note_cell = ws.cell(row=target_row, column=right_cols[0])
        note_cell.value = ranking_note
        note_cell.font = Font(size=26, bold=True)
        note_cell.alignment = Alignment(horizontal='center', vertical='center')
        # 设置右侧三列列宽为150
        for col in right_cols:
            ws.column_dimensions[get_column_letter(col)].width = 150
        # 设置行高
        ws.row_dimensions[target_row].height = 50

    # 保存最终文件
    wb.save(output_path)
    print(f"广告效率表已保存至: {output_path}")
    return ranking_percent


def _collect_ad_fetch_asins_from_nested(nested_result: dict) -> list[str]:
    """汇总目标 ASIN 及本地竞品列表中所有需请求广告数据的 ASIN。"""
    seen: set[str] = set()
    for asin, kw_map in (nested_result or {}).items():
        a = str(asin).strip().upper()
        if a:
            seen.add(a)
        for products in (kw_map or {}).values():
            for rec in products or []:
                if isinstance(rec, dict):
                    ca = str(rec.get('asin') or '').strip().upper()
                    if ca:
                        seen.add(ca)
    return sorted(seen)


def _ad_progress_emit_interval(total: int) -> int:
    if total <= 20:
        return 1
    if total <= 100:
        return 5
    return max(1, total // 40)


def _ops_gt10_ignore_200(review_interval: dict) -> bool:
    """运营难度是否命中 >10%，仅判断前4段（忽略 200以上）。"""
    if not isinstance(review_interval, dict):
        return False
    ops = review_interval.get('运营难度') or []
    if not isinstance(ops, list):
        return False
    for x in ops[:4]:
        t = str(x).replace('%', '').replace('％', '').strip()
        if not t:
            continue
        try:
            if float(t) > 10:
                return True
        except Exception:
            m = re.search(r'(\d+(?:\.\d+)?)', t)
            if m and float(m.group(1)) > 10:
                return True
    return False


async def _ad_difficulty_for_one_asin(
    asin: str,
    kw_map: dict,
    source_map: dict,
    asin_price_dict: dict,
    *,
    ads_cache: dict | None = None,
    progress_state: dict | None = None,
    keyword_sequential: bool = False,
    ads_max_concurrent: int | None = None,
    ads_fetch_rounds: int = 3,
    ad_rotation_state: dict | None = None,
    keyword_ban_retry: bool = False,
    marketplace: str = 'US',
) -> tuple[str, dict]:
    """单 ASIN 广告难度；定时顺序模式下支持关键词级断点续算。"""
    asin_key = str(asin).strip().upper()
    mp = str(marketplace or 'US').strip().upper()
    if mp not in ('US', 'UK'):
        mp = 'US'
    price_info = (
        asin_price_dict.get(asin_key)
        or asin_price_dict.get(asin)
        or (ads_cache_get(ads_cache, asin_key) if ads_cache else None)
        or {}
    )
    rp_candidates: list[float] = []
    details: dict = {}
    max_ss = ads_max_concurrent if ads_max_concurrent is not None else (
        1 if keyword_sequential else env_max_concurrent('sellersprite', 6)
    )

    target_ad_words = await _fetch_target_ad_word_count(
        asin_key,
        ads_cache=ads_cache,
        price_info=price_info if price_info else None,
        max_concurrent=max_ss,
        ads_fetch_rounds=ads_fetch_rounds,
    )
    if target_ad_words == 0:
        emit_progress(f'{asin_key} 目标 ASIN 广告词数量为 0，跳过广告难度计算')
        if progress_state is not None:
            n_kw = len(kw_map or {})
            if n_kw:
                async with progress_state['lock']:
                    progress_state['done'] += n_kw
        return asin_key, _zero_ad_words_skip_payload(asin_key)

    async def _one_keyword(keyword: str, products: list) -> tuple[str, dict]:
        output_dir = await verify_path(asin=asin_key, keyword=keyword)
        paths = _keyword_artifact_paths(output_dir, keyword)

        existing_rp = _read_ranking_from_ad_efficiency_xlsx(
            paths['ad_efficiency'], asin_key
        )
        if existing_rp is not None and existing_rp >= 0:
            emit_progress(
                f'跳过已完成关键词「{keyword}」({asin_key})，广告难度 {existing_rp}%'
            )
            return keyword, {
                'computed': True,
                'matched': True,
                'ranking_percent': float(existing_rp),
                'resumed': True,
            }

        df = pd.DataFrame(products)
        if df.empty:
            return keyword, {
                'computed': False,
                'matched': False,
                'ranking_percent': None,
                'reason': 'empty_products',
            }
        source_kind = (source_map.get(asin_key, {}) or source_map.get(asin, {}) or {}).get(
            keyword, 'search'
        )
        existing_clean = _load_existing_data_origin(paths['data_origin'])
        if existing_clean is not None and not existing_clean.empty:
            clean_data = existing_clean
            emit_progress(f'续算关键词「{keyword}」({asin_key})：复用已有 data_origin')
        elif source_kind == 'data_origin':
            clean_data = df
        else:
            clean_data = await save_cleaned_data_orign_to_excel(
                df,
                keyword,
                asin_key,
                price_info,
                ads_cache=ads_cache,
            )

        if _review_interval_file_exists(paths):
            emit_progress(
                f'续算关键词「{keyword}」({asin_key})：复用已有 review_interval_analysis'
            )
        else:
            await save_review_interval_analysis_to_excel(
                clean_data, keyword, asin_key, marketplace=mp
            )

        rp = await save_ad_efficiency_table(
            clean_data,
            keyword,
            asin_key,
            ads_cache=ads_cache,
            ads_max_concurrent=max_ss,
            ads_fetch_rounds=ads_fetch_rounds,
        )
        rp_num = pd.to_numeric(rp, errors='coerce')
        if not pd.isna(rp_num) and float(rp_num) >= 0:
            return keyword, {
                'computed': True,
                'matched': True,
                'ranking_percent': float(rp_num),
            }
        if not pd.isna(rp_num) and float(rp_num) == AD_EFFICIENCY_ZERO_AD_WORDS:
            return keyword, {
                'computed': False,
                'matched': False,
                'ranking_percent': None,
                'reason': 'zero_ad_words',
            }
        reason = 'invalid_ranking'
        if pd.isna(rp_num) or float(rp_num) == AD_EFFICIENCY_FAILED:
            reason = 'no_ad_efficiency_table'
        return keyword, {
            'computed': False,
            'matched': False,
            'ranking_percent': None,
            'reason': reason,
        }

    async def _one_keyword_resilient(keyword: str, products: list) -> tuple[str, dict]:
        if not (keyword_ban_retry and ad_rotation_state is not None):
            return await _one_keyword(keyword, products)
        max_rot = int(ad_rotation_state.get('max') or 16)
        attempts = 0
        while True:
            try:
                return await _one_keyword(keyword, products)
            except Exception as e:
                if not is_seller_account_banned_error(e):
                    raise
                attempts += 1
                if attempts > max_rot:
                    raise
                await _rotate_ad_account_for_keyword_resume(
                    asin_key,
                    keyword,
                    rotation_state=ad_rotation_state,
                )

    kw_items = list((kw_map or {}).items())
    if not kw_items:
        return asin_key, {'ranking_percent': None, 'computed_ad': False, 'keywords': {}}

    if keyword_sequential:
        kw_results = []
        for kw, prods in kw_items:
            kw_results.append(await _one_keyword_resilient(kw, prods))
    else:
        kw_results = await asyncio.gather(
            *[_one_keyword(kw, prods) for kw, prods in kw_items],
            return_exceptions=True,
        )

    for item in kw_results:
        if isinstance(item, Exception):
            raise item
        keyword, detail = item
        details[keyword] = detail
        if detail.get('computed') and detail.get('ranking_percent') is not None:
            rp_candidates.append(float(detail['ranking_percent']))
        if progress_state is not None:
            async with progress_state['lock']:
                progress_state['done'] += 1
                done_n = progress_state['done']
                total_n = progress_state['total']
            step = _ad_progress_emit_interval(total_n)
            if done_n >= total_n or done_n == 1 or done_n % step == 0:
                emit_progress(f'广告难度计算进度 {done_n}/{total_n}：{asin_key}')

    payload = {
        'ranking_percent': round(min(rp_candidates), 3) if rp_candidates else None,
        'computed_ad': bool(rp_candidates),
        'keywords': details,
    }
    if not rp_candidates:
        reasons = [
            str(d.get('reason') or 'failed')
            for d in details.values()
            if isinstance(d, dict)
        ]
        if reasons and all(r == 'zero_ad_words' for r in reasons):
            payload['error'] = AD_KEYWORD_ERROR_ZERO_AD_WORDS
            payload['skip_reason'] = 'zero_ad_words'
        else:
            messages = [_ad_keyword_failure_message(r) for r in reasons[:3]]
            payload['error'] = '; '.join(messages) if messages else 'no_valid_keyword_ranking'
    return asin_key, payload


def _ad_difficulty_partial_or_raise(
    out: dict,
    ban_pending: list[str],
    message: str,
) -> dict:
    """
    禁号且无法轮换时：未完成 ASIN 写入续算队列。
    若 out 中已有成功结果则返回带 __ban__ 的 dict（不抛异常），否则抛出 SellerAccountBannedError。
    """
    from bulk_account_pool import record_ban_pending_asins
    from seller_account_guard import SellerAccountBannedError

    pending = sorted(
        {str(a).strip().upper() for a in ban_pending if str(a).strip()}
    )
    if pending:
        record_ban_pending_asins(pending, task='ad')
    clean_out = {
        k: v for k, v in out.items() if not str(k).startswith('__')
    }
    has_success = any(
        isinstance(v, dict)
        and v.get('computed_ad')
        and v.get('ranking_percent') is not None
        for v in clean_out.values()
    )
    if has_success:
        result = dict(clean_out)
        result['__ban__'] = {'pending': pending, 'message': message}
        return result
    raise SellerAccountBannedError(
        message,
        partial_results=clean_out,
        ban_pending=pending,
    )


async def calculate_ad_difficulty_for_asins(
    target_asins: list[str] | None = None,
    *,
    sequential: bool = False,
    marketplace: str = 'US',
) -> dict:
    """
    从本地 file/{ASIN}/{关键词} 数据重算广告难度（使用批量账号池）：
    - 每个关键词先判断运营难度前4段是否存在 >10%（忽略 200以上）
    - 命中才生成广告效率表并产出该关键词 ranking_percent
    - ASIN 级 ranking_percent 取有效关键词最小值；若都不命中则为 0
    - sequential=True：定时任务顺序模式（逐个 ASIN、逐个请求卖家精灵，无并发）
    - 默认多 ASIN 并发；遇禁号自动切换最久未用批量账号并续算未完成 ASIN
    - marketplace=US|UK：卖家精灵 market / marketId 与本地 Search 文件站点
    """
    from sellersprite_market import (
        normalize_sellersprite_marketplace,
        reset_sellersprite_marketplace,
        set_sellersprite_marketplace,
    )

    mp = normalize_sellersprite_marketplace(marketplace)
    _ss_token = set_sellersprite_marketplace(mp)

    try:
        return await _calculate_ad_difficulty_for_asins_body(
            target_asins,
            sequential=sequential,
            marketplace=mp,
        )
    finally:
        reset_sellersprite_marketplace(_ss_token)


async def _calculate_ad_difficulty_for_asins_body(
    target_asins: list[str] | None = None,
    *,
    sequential: bool = False,
    marketplace: str = 'US',
) -> dict:
    from seller_account_guard import (
        bulk_rotate_if_available,
        clear_seller_login_cache,
        ensure_seller_login,
    )

    mp = marketplace
    try:
        from bulk_account_pool import pop_ban_pending_asins

        pending_resume = pop_ban_pending_asins(task='ad')
        if pending_resume:
            if target_asins:
                target_asins = sorted(
                    {str(x).strip().upper() for x in target_asins if str(x).strip()}
                    | set(pending_resume)
                )
            else:
                target_asins = pending_resume
            emit_progress(f'续算上次因禁号未完成的广告难度：{len(pending_resume)} 个 ASIN…')
    except ImportError:
        pass

    emit_progress(f'广告难度站点：{mp}')
    nested_result, _keyword_dict, source_map = await asyncio.to_thread(
        load_products_from_local_files, None, target_asins, marketplace=mp
    )
    if not nested_result:
        return {}

    rotation_state: dict = {'count': 0, 'max': 16}
    max_ss = 1 if sequential else env_max_concurrent('sellersprite', 6)
    ads_cache: dict = {}
    fetch_progress_state = {'last': 0}

    def _on_ad_fetch_progress(done: int, total: int, _asin: str = '') -> None:
        step = _ad_progress_emit_interval(total)
        if done >= total or done == 1 or done % step == 0:
            if fetch_progress_state.get('last') != done:
                fetch_progress_state['last'] = done
                emit_progress(f'广告数据请求进度 {done}/{total}')

    emit_progress(f'正在登录卖家精灵（批量账号）…')
    await ensure_seller_login()

    nested_result, zero_ad_out = await _prefilter_zero_ad_word_asins(
        nested_result,
        ads_cache=ads_cache,
        max_concurrent=max_ss,
    )
    out: dict = dict(zero_ad_out)
    if not nested_result:
        if zero_ad_out:
            emit_progress(
                f'广告难度：{len(zero_ad_out)} 个 ASIN 因广告词数量为 0 已全部跳过'
            )
        return out

    n_asin = len(nested_result)
    n_kw = sum(len(km or {}) for km in nested_result.values())
    all_fetch_asins = _collect_ad_fetch_asins_from_nested(nested_result)
    skipped_zero = len(zero_ad_out)
    if skipped_zero:
        emit_progress(
            f'广告难度：已跳过 {skipped_zero} 个广告词为 0 的 ASIN，'
            f'继续计算 {n_asin} 个 ASIN、{n_kw} 个关键词'
        )
    else:
        emit_progress(f'广告难度：共 {n_asin} 个 ASIN、{n_kw} 个关键词待分析')

    async def _prefetch_ads(
        batch_fetch: list[str],
        *,
        rotated: bool = False,
        ban_pending_asins: list[str] | None = None,
    ) -> dict | None:
        try:
            async with async_api_slot('sellersprite'):
                await ensure_ads_cached_robust(
                    ads_cache,
                    batch_fetch,
                    max_concurrent=max_ss,
                    max_rounds=3,
                    pause_sec=2.0,
                    on_progress=_on_ad_fetch_progress,
                    marketplace=mp,
                )
        except Exception as e:
            if is_seller_account_banned_error(e) and not rotated:
                emit_progress('广告难度：卖家精灵被禁，正在切换批量账号…')
                clear_seller_login_cache()
                if await bulk_rotate_if_available(
                    ban_pending_asins or batch_fetch,
                    rotation_state=rotation_state,
                    pending_task='ad',
                ):
                    await ensure_seller_login()
                    return await _prefetch_ads(
                        batch_fetch,
                        rotated=True,
                        ban_pending_asins=ban_pending_asins,
                    )
            if is_seller_account_banned_error(e):
                return _ad_difficulty_partial_or_raise(
                    out,
                    ban_pending_asins or batch_fetch,
                    str(e),
                )
            raise RuntimeError(f'广告数据批量获取失败: {e}') from e
        return None

    remaining = list(nested_result.keys())

    if sequential:
        emit_progress(
            f'广告难度：定时任务顺序模式，共 {n_asin} 个 ASIN、{n_kw} 个关键词'
            f'（逐个 ASIN 顺序请求卖家精灵）'
        )
        asin_price_dict: dict = {}
    else:
        emit_progress(
            f'正在向卖家精灵请求 {len(all_fetch_asins)} 个 ASIN 的广告数据（并发 {max_ss}）…'
        )
        partial = await _prefetch_ads(
            all_fetch_asins,
            ban_pending_asins=list(nested_result.keys()),
        )
        if partial is not None:
            return partial
        emit_progress(f'广告数据请求完成（{len(all_fetch_asins)} 个 ASIN）')
        asin_price_dict = {
            a: ads_cache_get(ads_cache, a)
            for a in nested_result.keys()
            if ads_cache_get(ads_cache, a)
        }
    calc_total = n_kw
    progress_state = {
        'done': 0,
        'total': calc_total,
        'lock': asyncio.Lock(),
    }
    if calc_total:
        emit_progress(f'开始计算广告难度（0/{calc_total} 个关键词任务）…')

    while remaining:
        if sequential:
            work_batch = [remaining[0]]
            per_fetch = _collect_ad_fetch_asins_from_nested(
                {work_batch[0]: nested_result[work_batch[0]]}
            )
            emit_progress(
                f'顺序模式：{work_batch[0]} 请求广告数据（{len(per_fetch)} 个 ASIN）…'
            )
            partial = await _prefetch_ads(
                per_fetch,
                ban_pending_asins=list(remaining),
            )
            if partial is not None:
                return partial
            batch_price = {
                work_batch[0]: ads_cache_get(ads_cache, work_batch[0]) or {}
            }
        else:
            work_batch = remaining
            batch_price = asin_price_dict

        max_asin = 1 if sequential else env_max_concurrent('scheduler_asin', 4)
        sem = asyncio.Semaphore(max_asin)

        async def _run_one(asin: str, kw_map: dict):
            async with sem:
                return await _ad_difficulty_for_one_asin(
                    asin,
                    kw_map,
                    source_map,
                    batch_price,
                    ads_cache=ads_cache,
                    progress_state=progress_state,
                    keyword_sequential=sequential,
                    ads_max_concurrent=max_ss,
                    ads_fetch_rounds=3,
                    ad_rotation_state=rotation_state,
                    keyword_ban_retry=sequential,
                    marketplace=mp,
                )

        raw_results = await asyncio.gather(
            *[_run_one(a, nested_result[a]) for a in work_batch],
            return_exceptions=True,
        )

        ban_pending: list[str] | None = None
        for asin, item in zip(work_batch, raw_results):
            if isinstance(item, Exception) and is_seller_account_banned_error(item):
                done_keys = {
                    a for a, it in zip(work_batch, raw_results)
                    if not isinstance(it, Exception)
                }
                ban_pending = [a for a in remaining if a not in done_keys]
                break

        for asin, item in zip(work_batch, raw_results):
            if ban_pending and asin in ban_pending:
                continue
            if isinstance(item, Exception):
                if is_seller_account_banned_error(item):
                    continue
                print(f"警告: ASIN {asin} 广告难度计算失败: {item}")
                out[asin] = {
                    'ranking_percent': None,
                    'computed_ad': False,
                    'keywords': {},
                    'error': str(item),
                }
                continue
            asin_key, payload = item
            out[asin_key] = payload

        if ban_pending:
            emit_progress('广告难度：检测到子账号被禁，切换账号并续算未完成 ASIN…')
            clear_seller_login_cache()
            if not await bulk_rotate_if_available(
                ban_pending,
                rotation_state=rotation_state,
                pending_task='ad',
            ):
                return _ad_difficulty_partial_or_raise(
                    out,
                    ban_pending,
                    '批量账号被禁且无法轮换（广告难度）',
                )
            await ensure_seller_login()
            refetch = _collect_ad_fetch_asins_from_nested(
                {a: nested_result[a] for a in ban_pending if a in nested_result}
            )
            partial = await _prefetch_ads(refetch, ban_pending_asins=ban_pending)
            if partial is not None:
                return partial
            remaining = ban_pending
            continue

        if sequential:
            remaining = remaining[len(work_batch):]
            delay = _ad_asin_delay_sec()
            if remaining and delay > 0:
                await asyncio.sleep(delay)
        else:
            break

    if calc_total:
        emit_progress(
            f'广告难度全部完成：{progress_state["done"]}/{calc_total} 个关键词任务，'
            f'{len(out)} 个 ASIN 有结果'
        )
    return out


def _node_path_from_record(record: dict) -> str:
    """从单行产品字典中取类目路径（兼容映射后字段名与 Excel 原文列名）。"""
    if not record:
        return ""
    for key in (
            "nodeLabelPath",
            "类目路径",
            "节点标签路径",
            "节点路径",
            "nodeLabelPathLocale",
    ):
        val = record.get(key)
        if val is None:
            continue
        if isinstance(val, float) and pd.isna(val):
            continue
        s = str(val).strip()
        if s and s.lower() != "nan":
            return s
    return ""


async def collect_node_label_paths(
        keyword_dict: dict, result: dict, target_asins: list = None
) -> dict:
    """
    从爬取/本地加载结果中为每个目标 ASIN 收集一个 nodeLabelPath。

    result 支持两种结构：
    - 嵌套：{asin: {keyword: products_list}}（本地 file/{ASIN}/{关键词}/ 扫描结果）
    - 扁平：{keyword: products_list}（旧版 fetch_multiple_keywords）
    """
    if target_asins is None:
        target_asins = list(keyword_dict.keys())
    else:
        target_asins = [a for a in target_asins if a in keyword_dict]

    asin_to_path = {}
    if not result:
        return asin_to_path

    first_val = next(iter(result.values()))
    nested = isinstance(first_val, dict) and not isinstance(first_val, list)

    if nested:
        for asin in target_asins:
            kw_map = result.get(asin) or {}
            node_path = ""
            src_kw = ""
            for kw, products in kw_map.items():
                for rec in products or []:
                    node_path = _node_path_from_record(rec)
                    if node_path:
                        src_kw = kw
                        break
                if node_path:
                    break
            if node_path:
                asin_to_path[asin] = node_path
                print(f"为 ASIN {asin} 获取 nodeLabelPath: {node_path} (关键词: {src_kw})")
            else:
                asin_to_path[asin] = ""
                print(
                    f"警告：ASIN {asin} 在 Excel 中未找到有效「类目路径」/nodeLabelPath，"
                    f"ROI 中相关接口将使用空路径。"
                )
        return asin_to_path

    for asin in target_asins:
        node_path = ""
        src_kw = ""
        for ky, products in result.items():
            if ky not in keyword_dict.get(asin, []):
                continue
            for rec in products or []:
                node_path = _node_path_from_record(rec)
                if node_path:
                    src_kw = ky
                    break
            if node_path:
                break
        if node_path:
            asin_to_path[asin] = node_path
            print(f"为 ASIN {asin} 获取 nodeLabelPath: {node_path} (关键词: {src_kw})")
        else:
            asin_to_path[asin] = ""
            print(
                f"警告：ASIN {asin} 未找到有效类目路径，ROI 中相关接口将使用空路径。"
            )
    return asin_to_path


def asin_image_file(asin: str, *, for_write: bool = False) -> Path:
    """ASIN 主图路径：写入 media/images（或 ASIN_IMAGES_ROOT）；读取时可回退旧目录。"""
    name = f"{str(asin).strip().upper()}.jpg"
    primary = resolve_asin_images_dir() / name
    if for_write:
        return primary
    if primary.is_file():
        return primary
    legacy = _LEGACY_IMAGES_DIR / name
    if legacy.is_file():
        return legacy
    return primary


def _parse_unit_purchase_cell(raw) -> float | None:
    if raw is None or (isinstance(raw, float) and pd.isna(raw)):
        return None
    t = str(raw).replace('￥', '').replace(',', '').strip()
    if not t or t.upper() in ('N/A', 'NA', '-', '—'):
        return None
    try:
        return float(t)
    except (TypeError, ValueError):
        return None


def _positive_float(val) -> float | None:
    """解析为正数；0/空/无效视为缺失。"""
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return None
    try:
        n = float(val)
    except (TypeError, ValueError):
        return None
    return n if n > 0 else None


def _parse_money(val) -> float | None:
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return None
    try:
        s = str(val).replace('$', '').replace('￥', '').replace(',', '').strip()
        if not s or s.upper() in ('N/A', 'NA', '-', '—'):
            return None
        n = float(s)
        return n if n > 0 else None
    except (TypeError, ValueError):
        return None


def _parse_percent(val) -> float | None:
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return None
    try:
        s = str(val).replace('%', '').replace(',', '').strip()
        if not s or s.upper() in ('N/A', 'NA', '-', '—'):
            return None
        n = float(s)
        return n if n >= 0 else None
    except (TypeError, ValueError):
        return None


def read_labeled_field_from_roi_pack(asin: str, label: str) -> float | None:
    """从 ROI-US-pack 任意列组读取带「字段/值」结构的数值。"""
    a = str(asin).strip().upper()
    pack = FILE_DATA_ROOT / a / f"{a}_ROI-US-pack.xlsx"
    if not pack.is_file():
        return None
    try:
        df = pd.read_excel(pack)
        for _, row in df.iterrows():
            for col_idx in range(len(row)):
                field = str(row.iloc[col_idx]).strip() if pd.notna(row.iloc[col_idx]) else ''
                if field != label:
                    continue
                val_col = col_idx + 1
                if val_col < len(row):
                    cell = row.iloc[val_col]
                    if label == '退款率':
                        parsed = _parse_percent(cell)
                    else:
                        parsed = _parse_money(cell)
                    if parsed is None:
                        parsed = _parse_unit_purchase_cell(cell)
                    if parsed is not None:
                        return parsed
    except Exception as e:
        print(f"读取 ROI 字段「{label}」失败 ({pack}): {e}")
    return None


def read_text_field_from_roi_pack(asin: str, label: str) -> str | None:
    a = str(asin).strip().upper()
    pack = FILE_DATA_ROOT / a / f"{a}_ROI-US-pack.xlsx"
    if not pack.is_file():
        return None
    try:
        df = pd.read_excel(pack)
        for _, row in df.iterrows():
            for col_idx in range(len(row)):
                field = str(row.iloc[col_idx]).strip() if pd.notna(row.iloc[col_idx]) else ''
                if field != label:
                    continue
                val_col = col_idx + 1
                if val_col < len(row) and pd.notna(row.iloc[val_col]):
                    url = str(row.iloc[val_col]).strip()
                    if url and url.upper() not in ('N/A', 'NA', 'NAN'):
                        return url
    except Exception as e:
        print(f"读取 ROI 文本字段「{label}」失败 ({pack}): {e}")
    return None


def build_local_product_hints(nested_result: dict, asin: str) -> dict:
    """从本地 Excel 提取类目路径（供退款率接口入参，不作数值回落）。"""
    hints: dict = {}
    a = str(asin).strip().upper()
    kw_map = nested_result.get(a) or nested_result.get(asin) or {}
    for products in kw_map.values():
        for rec in products or []:
            if not isinstance(rec, dict):
                continue
            if not hints.get('nodeLabelPath'):
                path = _node_path_from_record(rec)
                if path:
                    hints['nodeLabelPath'] = path
                    break
        if hints.get('nodeLabelPath'):
            break
    return hints


def _resolve_image_url_for_roi(
    asin: str,
    info: dict | None,
    *,
    marketplace: str | None = None,
) -> str:
    from sellersprite_market import sellersprite_amazon_dp_url

    url = str((info or {}).get('imageUrl') or '').strip()
    if url and url.upper() not in ('N/A', 'NA', 'NAN'):
        return url
    cached = read_image_url_from_roi_pack(asin)
    if cached and str(cached).strip().upper() not in ('N/A', 'NA', 'NAN', ''):
        return str(cached).strip()
    return sellersprite_amazon_dp_url(asin, marketplace)


async def _resolve_refund_rate(
    asin: str,
    node_label_path: str,
    hints: dict | None,
    refund_cache: dict[str, float] | None = None,
) -> float:
    path = (node_label_path or '').strip() or str((hints or {}).get('nodeLabelPath') or '').strip()
    if not path:
        print(
            f'警告: ASIN {asin} 缺少类目路径，退款率使用默认值 {ROI_DEFAULT_REFUND_RATE}%'
        )
        return ROI_DEFAULT_REFUND_RATE
    if refund_cache is not None and path in refund_cache:
        return refund_cache[path]
    try:
        refund_raw = await fetch_refund_rate_for_path(path)
        text = str(refund_raw or '').replace('%', '').strip()
        if not text:
            raise ValueError('退款率接口无数据')
        val = float(text)
        if val <= 0:
            raise ValueError(f'退款率无效: {val}')
    except Exception as exc:
        print(
            f'警告: ASIN {asin} 退款率获取失败（{exc}），'
            f'使用默认值 {ROI_DEFAULT_REFUND_RATE}%'
        )
        return ROI_DEFAULT_REFUND_RATE
    if refund_cache is not None:
        refund_cache[path] = val
    return val


async def prefetch_refund_rates(
    paths: list[str],
    *,
    max_concurrent: int | None = None,
    marketplace: str | None = None,
) -> dict[str, float]:
    """按类目路径批量预取退款率（同一路径只请求一次；失败类目不中断整批）。"""
    limit = max_concurrent or env_max_concurrent('sellersprite', 6)
    out, failures = await prefetch_refund_rates_batch(
        paths, max_concurrent=limit, marketplace=marketplace
    )
    if failures:
        emit_progress(
            f'退款率：{len(out)} 个类目成功，{len(failures)} 个类目无数据（相关 ASIN 将单独失败）'
        )
        for row in failures[:8]:
            emit_progress(f"  退款率跳过: {str(row.get('path', ''))[:55]}…")
        if len(failures) > 8:
            emit_progress(f'  … 另有 {len(failures) - 8} 个类目无退款率')
    return out


async def _resolve_fba_and_head(
    asin: str,
    fba_info_dict: dict,
    head_distance_override: float | None,
    default_fba_fee: float | None = None,
) -> tuple[float, float]:
    info = fba_info_dict.get(asin) or {}
    if not isinstance(info, dict):
        info = {}

    fba_fee = _parse_money(info.get('FBA'))
    head_distance = info.get('head_distance')
    if head_distance is not None:
        try:
            head_distance = float(head_distance)
        except (TypeError, ValueError):
            head_distance = None

    fba_fallback = float(default_fba_fee) if default_fba_fee is not None else ROI_DEFAULT_FBA_FEE
    if fba_fee is None or fba_fee <= 0:
        print(
            f'警告: ASIN {asin} 无法获取 FBA 配送费，'
            f'使用默认值 ${fba_fallback}'
        )
        fba_fee = fba_fallback

    if head_distance_override is not None:
        head_distance = head_distance_override
    elif head_distance is None:
        head_distance = 0.0

    return float(fba_fee), float(head_distance)


def read_unit_purchase_from_roi_pack(asin: str) -> float | None:
    """从已有 ROI-US-pack 表读取「单件采购」，供图搜失败时回落。"""
    return read_labeled_field_from_roi_pack(asin, '单件采购')


def read_image_url_from_roi_pack(asin: str) -> str | None:
    """从已有 ROI-US-pack 读取「图片链接」。"""
    return read_text_field_from_roi_pack(asin, '图片链接')


async def ensure_asin_image_path(
    asin: str,
    info: dict | None,
    current_path: str = "",
) -> str | None:
    """确保 ASIN 主图存在；优先卖家精灵 imageUrl，其次旧 ROI 表图片链接。"""
    p = Path(current_path) if current_path else asin_image_file(asin)
    if p.is_file():
        return str(p.resolve())
    local = asin_image_file(asin)
    if local.is_file():
        return str(local.resolve())

    url = str((info or {}).get('imageUrl') or '').strip()
    if not url or url.upper() in ('N/A', 'NA', 'NAN'):
        cached = read_image_url_from_roi_pack(asin)
        url = str(cached or '').strip()
    if not url or url.upper() in ('N/A', 'NA', 'NAN'):
        return None
    # 商品页 URL 无法当主图下载
    if 'amazon.' in url.lower() and '/dp/' in url.lower() and 'images-' not in url.lower():
        return None

    got = await download_one(asin, {'imageUrl': url})
    if got:
        return got
    return None


async def _fetch_unit_purchase_via_taobao(image_file: str, tokens: list[str]) -> float | None:
    """1688 以图搜价；失败时用 prefer_network 刷新 token 再试一次。"""
    if not tokens or len(tokens) < 2:
        print('淘宝 token 不可用，跳过图搜')
        return None
    val = await async_price_info_main(image_file, tokens[0], tokens[1])
    if val is not None:
        return val
    print('图搜无结果或 token 失效，尝试刷新淘宝 token 后重试…')
    fresh = await asyncio.to_thread(lambda: get_taobao_tokens(prefer_network=True))
    return await async_price_info_main(image_file, fresh['_m_h5_tk'], fresh['_m_h5_tk_enc'])


async def save_roi_us_pack(nodeLabelPath: str,
                           fba_info_dict: dict,
                           asin: str,
                           asin_cpc_list: list,
                           monthly_sales_dict: dict,
                           tokens: list[str],
                           image_path: str,
                           exchange_rate: float = 7.2,
                           asin_info_dict: dict = None,
                           unit_purchase_override: float | None = None,
                           head_distance_override: float | None = None,
                           local_hints: dict | None = None,
                           refund_cache: dict[str, float] | None = None,
                           marketplace: str = 'US',
                           roi_defaults: dict | None = None):
    """
    生成 ROI-US-pack 表，输出三组并排数据：左侧基础成本、中间广告相关、右侧流量与利润指标，
    每组包含字段、值、单位三列。
    marketplace=US|UK；roi_defaults 可覆盖平台佣金与默认采购/FBA/退款率。
    """
    product_asin = asin
    hints = local_hints or {}
    node_label_path = (nodeLabelPath or '').strip() or str(hints.get('nodeLabelPath') or '').strip()
    mp = str(marketplace or 'US').strip().upper()
    if mp not in ('US', 'UK'):
        mp = 'US'
    rd = roi_defaults if isinstance(roi_defaults, dict) else {}
    default_unit_purchase = float(rd.get('default_unit_purchase') or ROI_DEFAULT_UNIT_PURCHASE)
    default_fba_fee = float(rd.get('default_fba_fee') or ROI_DEFAULT_FBA_FEE)
    default_refund_rate = float(rd.get('default_refund_rate') or ROI_DEFAULT_REFUND_RATE)
    try:
        platform_commission = float(rd['platform_commission']) if 'platform_commission' in rd else (
            25 if mp == 'UK' else 15
        )
    except (TypeError, ValueError):
        platform_commission = 25 if mp == 'UK' else 15

    # ==================== 1. 获取基础数据 ====================
    fba_fee, head_distance = await _resolve_fba_and_head(
        product_asin,
        fba_info_dict,
        head_distance_override,
        default_fba_fee=default_fba_fee,
    )

    if unit_purchase_override is not None:
        unit_purchase = unit_purchase_override
    else:
        resolved_image = await ensure_asin_image_path(
            product_asin, asin_info_dict or {}, image_path or ''
        )
        if resolved_image:
            unit_purchase = await _fetch_unit_purchase_via_taobao(resolved_image, tokens)
            print(unit_purchase, asin, '3333')
        else:
            unit_purchase = None
            print(f'警告: ASIN {asin} 主图不可用，跳过图搜')
        if unit_purchase is None:
            unit_purchase = read_unit_purchase_from_roi_pack(product_asin)
        if unit_purchase is None:
            unit_purchase = default_unit_purchase
            print(
                f'警告: ASIN {asin} 无法获取单件采购价，'
                f'使用默认值 ￥{default_unit_purchase}'
            )

    refund_rate = await _resolve_refund_rate(
        product_asin, node_label_path, hints, refund_cache=refund_cache
    )

    asin_price = _positive_float((asin_info_dict or {}).get('avg_price'))

    # 计算基础售价
    if unit_purchase is not None and head_distance is not None:
        cost_cny = unit_purchase + head_distance
        cost_usd = cost_cny / exchange_rate
        total_cost = cost_usd + fba_fee
        denominator = 1 - (platform_commission + refund_rate) / 100
        if denominator <= 0:
            print(f'警告: ASIN {asin} 佣金+退款率过高，退款率改用默认值')
            refund_rate = default_refund_rate
            denominator = 1 - (platform_commission + refund_rate) / 100
        lowest_price = total_cost / denominator if denominator > 0 else None
    else:
        lowest_price = None

    if asin_price is not None:
        product_price = asin_price * 1.02
    elif lowest_price is not None:
        product_price = lowest_price * 2
    else:
        product_price = ROI_DEFAULT_PRODUCT_PRICE
        print(
            f'警告: ASIN {asin} 无法推算售价，'
            f'使用默认值 ${ROI_DEFAULT_PRODUCT_PRICE}'
        )

    discount = 0
    if product_price is not None:
        discounted_price = product_price * (1 - discount / 100)
        refund_amount = product_price * (refund_rate / 100)
        commission_amount = discounted_price * (platform_commission / 100)
        shipping_fba = fba_fee

        sales_return = discounted_price - refund_amount - commission_amount - shipping_fba
        if unit_purchase is not None and head_distance is not None:
            cost_cny = unit_purchase + head_distance
            cost_usd = cost_cny / exchange_rate
            actual_profit = sales_return - cost_usd
            actual_cost = discounted_price - actual_profit
        else:
            actual_profit = None
            actual_cost = None
    else:
        discounted_price = None
        sales_return = None
        actual_profit = None
        actual_cost = None

    if actual_profit is None or discounted_price is None or discounted_price <= 0:
        if product_price is None:
            product_price = ROI_DEFAULT_PRODUCT_PRICE
        discounted_price = product_price * (1 - discount / 100)
        if unit_purchase is None:
            unit_purchase = default_unit_purchase
        if head_distance is None:
            head_distance = 0.0
        if fba_fee is None or fba_fee <= 0:
            fba_fee = default_fba_fee
        cost_usd = (float(unit_purchase) + float(head_distance)) / exchange_rate
        commission_amount = discounted_price * (platform_commission / 100)
        refund_amount = product_price * (refund_rate / 100)
        sales_return = discounted_price - refund_amount - commission_amount - fba_fee
        actual_profit = max(sales_return - cost_usd, 0.01)
        actual_cost = discounted_price - actual_profit
        print(
            f'警告: ASIN {asin} 利润数据不完整，已用默认值估算 '
            f'（折后价={discounted_price:.2f}，利润={actual_profit:.2f}）'
        )

    product_key = str(product_asin).strip().upper()
    monthly_sales = (
        monthly_sales_dict.get(product_asin)
        or monthly_sales_dict.get(product_key)
        or 0.0
    )
    if not monthly_sales:
        monthly_sales = 0.0
    daily_orders = monthly_sales / 30.0 if monthly_sales > 0 else 0.0

    # ==================== 2. 广告相关计算 ====================
    ad_cpc, conversion_rate, ad_clicks = resolve_sif_ad_metrics(
        product_asin,
        asin_cpc_list,
        daily_orders,
    )

    ad_budget = ad_cpc * ad_clicks

    # 日利润、月利润（出单量为 0 时按 0 计，仍保留单位经济学指标）
    daily_profit1 = actual_profit * daily_orders - ad_budget
    daily_profit2 = daily_profit1 * exchange_rate
    monthly_profit1 = daily_profit1 * 30
    monthly_profit2 = daily_profit2 * 30

    # 广告费占比（出单量为 0 时广告占比为 0，去广告毛利率 = 利润率）
    if discounted_price > 0:
        if daily_orders > 0:
            ad_cost_ratio = ad_budget / (discounted_price * daily_orders) * 100
        else:
            ad_cost_ratio = 0.0
    else:
        ad_cost_ratio = 0.0

    # ==================== 3. 右侧新增字段计算 ====================
    cost_cny_total = (
            unit_purchase + head_distance) if unit_purchase is not None and head_distance is not None else None

    # 总流量
    if conversion_rate is not None and conversion_rate > 0 and daily_orders is not None:
        total_traffic = daily_orders / conversion_rate
    else:
        total_traffic = None

    # 日自然流量（总流量的 30%；广告点击为 70%）
    if total_traffic is not None:
        daily_natural_traffic = total_traffic * 0.3
    else:
        daily_natural_traffic = None

    # 每单需点击
    if conversion_rate is not None and conversion_rate > 0:
        clicks_per_order = 1 / conversion_rate
    else:
        clicks_per_order = None

    # 去广告投产
    if cost_cny_total is not None and cost_cny_total > 0 and daily_orders > 0:
        profit_cny = daily_profit1 * exchange_rate
        ad_removed_roi = profit_cny / (cost_cny_total * daily_orders)
    elif cost_cny_total is not None and cost_cny_total > 0:
        ad_removed_roi = 0.0
    else:
        ad_removed_roi = None

    # 每单利润
    if daily_orders > 0:
        profit_per_order = daily_profit2 / daily_orders
    else:
        profit_per_order = 0.0

    # 投产比
    if cost_cny_total is not None and cost_cny_total > 0:
        roi_ratio = (actual_profit * exchange_rate) / cost_cny_total * 100
    else:
        roi_ratio = None

    # 利润率 / 去广告毛利率（正常流程必须产出数值）
    profit_margin = (actual_profit / discounted_price) * 100
    ad_removed_gross_margin = profit_margin - (ad_cost_ratio or 0.0)
    # 单件采购+单件头程
    unit_head_price = (
        unit_purchase + head_distance
        if unit_purchase is not None and head_distance is not None
        else None
    )
    from sellersprite_market import sellersprite_amazon_dp_url

    target_asin_url = sellersprite_amazon_dp_url(product_asin, mp)
    link1688 = 'https://aibuy.1688.com/landingpage/home/inventory/products.html?bizType=selectionTool&customerId=sellerspriteLP&lang=zh&currency=CNY'
    imageUrl = _resolve_image_url_for_roi(
        product_asin, asin_info_dict, marketplace=mp
    )

    # ==================== 4. 构建三组带单位的 DataFrame ====================
    # 左侧：基础成本与售价
    left_fields = [
        '单件采购', '单件头程', '平台佣金', '退款率', 'FBA配送',
        '最低售价', '产品售价', '优惠折扣', '折后价格', '平台佣金',
        '实际利润', '销售回款', '实际成本', '投产比'
    ]
    left_values = [
        f"{unit_purchase:.2f}" if unit_purchase is not None else 'N/A',
        f"{head_distance:.2f}" if head_distance is not None else 'N/A',
        f"{platform_commission}",
        f"{refund_rate:.2f}",
        f"{fba_fee:.2f}",
        f"{lowest_price:.2f}" if lowest_price is not None else 'N/A',
        f"{product_price:.2f}" if product_price is not None else 'N/A',
        f"{discount}",
        f"{discounted_price:.2f}" if discounted_price is not None else 'N/A',
        f"{platform_commission}",
        f"{actual_profit:.2f}" if actual_profit is not None else 'N/A',
        f"{sales_return:.2f}" if sales_return is not None else 'N/A',
        f"{actual_cost:.2f}" if actual_cost is not None else 'N/A',
        f"{roi_ratio:.2f}" if roi_ratio is not None else 'N/A'
    ]
    left_units = [
        '￥', '￥', '%', '%', '$',
        '$', '$', '%', '$', '%',
        '$', '$', '$', '%'
    ]

    # 中间：广告相关
    middle_fields = [
        '广告预算', '广告cpc', '广告点击', '转化率', '月出单量', '出单量',
        '日利润1', '日利润2', '月利润1', '月利润2', '广告费占比','图片链接',
        'asin链接','1688链接'
    ]
    middle_values = [
        f"{ad_budget:.2f}",
        f"{ad_cpc:.2f}",
        f"{ad_clicks:.2f}",
        f"{conversion_rate * 100:.2f}",
        f"{daily_orders * 30:.2f}" if daily_orders else 'N/A',
        f"{daily_orders:.2f}" if daily_orders else 'N/A',
        f"{daily_profit1:.2f}" if daily_profit1 is not None else 'N/A',
        f"{daily_profit2:.2f}" if daily_profit2 is not None else 'N/A',
        f"{monthly_profit1:.2f}" if monthly_profit1 is not None else 'N/A',
        f"{monthly_profit2:.2f}" if monthly_profit2 is not None else 'N/A',
        f"{ad_cost_ratio:.2f}" if ad_cost_ratio is not None else 'N/A',
        imageUrl,
        f"{target_asin_url}" if target_asin_url is not None else 'N/A',
        f"{link1688}" if link1688 is not None else 'N/A'
    ]
    middle_units = [
        '$', '$', '次', '%', '单', '单',
        '$', '￥', '$', '￥', '%','a',
        'a','a'
    ]

    # 右侧：流量与利润指标
    extra_fields = [
        '总流量', '日自然流量', '每单需点击', '去广告投产',
        '每单利润', '采购总价格', '利润率', '去广告毛利率'
    ]
    extra_values = [
        f"{total_traffic:.2f}" if total_traffic is not None else 'N/A',
        f"{daily_natural_traffic:.2f}" if daily_natural_traffic is not None else 'N/A',
        f"{clicks_per_order:.2f}" if clicks_per_order is not None else 'N/A',
        f"{ad_removed_roi * 100:.2f}" if ad_removed_roi is not None else 'N/A',
        f"{profit_per_order:.2f}" if profit_per_order is not None else 'N/A',
        f"{unit_head_price:.2f}" if unit_head_price is not None else 'N/A',
        f"{profit_margin:.2f}",
        f"{ad_removed_gross_margin:.2f}"
    ]
    extra_units = [
        '次', '次', '次', '%',
        '￥', '￥', '%', '%'
    ]

    # 构建三个 DataFrame
    left_df = pd.DataFrame({'字段': left_fields, '值': left_values, '单位': left_units})
    middle_df = pd.DataFrame({'字段': middle_fields, '值': middle_values, '单位': middle_units})
    extra_df = pd.DataFrame({'字段': extra_fields, '值': extra_values, '单位': extra_units})

    # 水平合并
    result_df = pd.concat([left_df, middle_df, extra_df], axis=1)

    # 保存 Excel（先写临时文件，成功后再替换，避免失败时覆盖旧 ROI-US-pack）
    output_dir = await verify_path(asin=asin)
    os.makedirs(output_dir, exist_ok=True)
    output_path = Path(output_dir) / f'{product_asin}_ROI-US-pack.xlsx'
    tmp_path = output_path.with_suffix('.part.xlsx')
    if tmp_path.is_file():
        tmp_path.unlink(missing_ok=True)
    try:
        result_df.to_excel(tmp_path, index=False)

        # ==================== 样式设置（含移动右侧特殊字段） ====================
        wb = load_workbook(tmp_path)
        ws = wb.active

        # 特殊字段名称
        special_field_names = ['去广告投产', '采购总价格', '每单利润', '利润率', '去广告毛利率']

        # 右侧三列的列号（假设三组列分别为 1-3, 4-6, 7-9）
        right_cols = [7, 8, 9]  # 字段列、值列、单位列

        # 找到右侧字段列中所有特殊字段的行
        special_rows = []
        for row in range(2, ws.max_row + 1):
            field_cell = ws.cell(row=row, column=right_cols[0])
            if field_cell.value in special_field_names:
                special_rows.append(row)

        if special_rows:
            # 收集这些行的右侧三列数据
            special_data = []
            for row in special_rows:
                row_data = []
                for col in right_cols:
                    row_data.append(ws.cell(row=row, column=col).value)
                special_data.append(row_data)

            # 在原位置清空这些单元格
            for row in special_rows:
                for col in right_cols:
                    ws.cell(row=row, column=col).value = ''

            # 找到表格末尾行（最后一行的下一行）
            last_row = ws.max_row
            # 将特殊数据追加到底部
            for i, row_data in enumerate(special_data):
                new_row = last_row + 1 + i
                for j, col in enumerate(right_cols):
                    ws.cell(row=new_row, column=col).value = row_data[j]

            # 为移动后的特殊行设置样式（行高、字体）
            for i, row_data in enumerate(special_data):
                new_row = last_row + 1 + i
                ws.row_dimensions[new_row].height = 50
                for col in right_cols:
                    cell = ws.cell(row=new_row, column=col)
                    cell.font = Font(size=36, bold=True)
                    cell.alignment = Alignment(horizontal='center', vertical='center')

        # 1. 设置列宽
        for col in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col)
            cell_value = ws.cell(row=1, column=col).value
            if cell_value == '单位':
                ws.column_dimensions[col_letter].width = 10
            else:
                ws.column_dimensions[col_letter].width = 25

        # 右侧值列（第8列）宽45
        right_value_col = 8
        ws.column_dimensions[get_column_letter(right_value_col)].width = 45
        ws.column_dimensions[get_column_letter(7)].width = 45

        # 2. 设置行高（所有行先30）
        for row in range(1, ws.max_row + 1):
            ws.row_dimensions[row].height = 30

        # 3. 设置全局字体（所有单元格18加粗，居中）
        bold_font = Font(size=18, bold=True)
        for row in ws.iter_rows():
            for cell in row:
                cell.font = bold_font
                cell.alignment = Alignment(horizontal='center', vertical='center')

        # 4. 重新为移动后的特殊行设置样式（避免被全局覆盖）
        if special_rows:
            for i, row_data in enumerate(special_data):
                new_row = last_row + 1 + i
                ws.row_dimensions[new_row].height = 50
                for col in right_cols:
                    cell = ws.cell(row=new_row, column=col)
                    cell.font = Font(size=36, bold=True)
                    cell.alignment = Alignment(horizontal='center', vertical='center')

        # 5. 高亮指定字段
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        highlight_fields = [
            '最低售价', '折后价格', '平台佣金', '实际利润', '销售回款', '实际成本',
            '广告预算', '广告点击', '出单量', '日利润1', '日利润2', '月利润1', '月利润2',
            '广告费占比', '总流量', '日自然流量', '每单需点击', '去广告投产', '每单利润',
            '投产比', '利润率', '去广告毛利率', '采购总价格'
        ]

        # 找到所有“值”列
        value_cols = []
        for col in range(1, ws.max_column + 1):
            if ws.cell(row=1, column=col).value == '值':
                value_cols.append(col)

        for row in range(2, ws.max_row + 1):
            for val_col in value_cols:
                field_col = val_col - 1
                field_cell = ws.cell(row=row, column=field_col)
                if field_cell.value in highlight_fields:
                    value_cell = ws.cell(row=row, column=val_col)
                    value_cell.fill = yellow_fill

        # 保存最终样式
        wb.save(tmp_path)
        shutil.move(str(tmp_path), str(output_path))
    except Exception as exc:
        if tmp_path.is_file():
            tmp_path.unlink(missing_ok=True)
        print(f'警告: ASIN {product_asin} 样式写入失败（{exc}），改为保存简化版 ROI-US-pack')
        result_df.to_excel(output_path, index=False)
        emit_progress(f'ASIN {product_asin} ROI-US-pack 已生成（简化版）')

    print(f"ROI-US-pack 表已保存至: {output_path}")
    if not output_path.is_file():
        raise RuntimeError(f'ASIN {product_asin} ROI-US-pack 写入失败')
    emit_progress(f'ASIN {product_asin} ROI-US-pack 已生成')
    return {
        product_asin: {
            # 看板「去广告毛利率」列对应 profit_margin 字段
            'profit_margin': round(ad_removed_gross_margin, 2),
            'gross_profit_rate': round(profit_margin, 2),
            'unit_purchase': unit_purchase,
            'monthly_profit1': monthly_profit1,
            'monthly_results': monthly_sales,
            'profit_per_order': profit_per_order,
            'head_distance': head_distance,
            'actual_cost': actual_cost,
            'ad_removed_roi': ad_removed_roi,
        }
    }


async def force_write_minimal_roi_us_pack(
    asin: str,
    exchange_rate: float = 7.2,
    marketplace: str = 'US',
    roi_defaults: dict | None = None,
) -> dict:
    """极端兜底：用默认值写出 ROI-US-pack（确保计算 ROI 页能识别为已计算）。"""
    product_asin = str(asin).strip().upper()
    mp = str(marketplace or 'US').strip().upper()
    if mp not in ('US', 'UK'):
        mp = 'US'
    rd = roi_defaults if isinstance(roi_defaults, dict) else {}
    try:
        platform_commission = float(rd['platform_commission']) if 'platform_commission' in rd else (
            25 if mp == 'UK' else 15
        )
    except (TypeError, ValueError):
        platform_commission = 25 if mp == 'UK' else 15
    unit_purchase = float(rd.get('default_unit_purchase') or ROI_DEFAULT_UNIT_PURCHASE)
    head_distance = 0.0
    fba_fee = float(rd.get('default_fba_fee') or ROI_DEFAULT_FBA_FEE)
    refund_rate = float(rd.get('default_refund_rate') or ROI_DEFAULT_REFUND_RATE)
    product_price = ROI_DEFAULT_PRODUCT_PRICE
    discounted_price = product_price
    actual_profit = max(discounted_price * 0.15, 0.01)
    ad_cpc = SIF_DEFAULT_AD_CPC
    conversion_rate = SIF_DEFAULT_CONVERSION_RATE
    ad_clicks = SIF_DEFAULT_AD_CLICKS
    ad_budget = ad_cpc * ad_clicks
    profit_margin = (actual_profit / discounted_price) * 100
    ad_removed_gross_margin = profit_margin

    left_fields = ['单件采购', '单件头程', '平台佣金', '退款率', 'FBA配送', '折后价格', '实际利润']
    left_values = [
        f'{unit_purchase:.2f}', f'{head_distance:.2f}', f'{platform_commission}', f'{refund_rate:.2f}',
        f'{fba_fee:.2f}', f'{discounted_price:.2f}', f'{actual_profit:.2f}',
    ]
    left_units = ['￥', '￥', '%', '%', '$', '$', '$']
    from sellersprite_market import sellersprite_amazon_dp_url

    _dp = sellersprite_amazon_dp_url(product_asin, mp)
    middle_fields = ['广告预算', '广告cpc', '广告点击', '转化率', '图片链接', 'asin链接']
    middle_values = [
        f'{ad_budget:.2f}', f'{ad_cpc:.2f}', f'{ad_clicks:.2f}',
        f'{conversion_rate * 100:.2f}',
        _dp,
        _dp,
    ]
    middle_units = ['$', '$', '次', '%', 'a', 'a']
    extra_fields = ['利润率', '去广告毛利率']
    extra_values = [f'{profit_margin:.2f}', f'{ad_removed_gross_margin:.2f}']
    extra_units = ['%', '%']

    left_df = pd.DataFrame({'字段': left_fields, '值': left_values, '单位': left_units})
    middle_df = pd.DataFrame({'字段': middle_fields, '值': middle_values, '单位': middle_units})
    extra_df = pd.DataFrame({'字段': extra_fields, '值': extra_values, '单位': extra_units})
    result_df = pd.concat([left_df, middle_df, extra_df], axis=1)

    output_dir = await verify_path(asin=product_asin)
    os.makedirs(output_dir, exist_ok=True)
    output_path = Path(output_dir) / f'{product_asin}_ROI-US-pack.xlsx'
    result_df.to_excel(output_path, index=False)
    emit_progress(f'ASIN {product_asin} 已用默认值强制生成 ROI-US-pack')
    print(f'强制生成 ROI-US-pack: {output_path}')
    return {
        product_asin: {
            'profit_margin': round(ad_removed_gross_margin, 2),
            'gross_profit_rate': round(profit_margin, 2),
            'unit_purchase': unit_purchase,
            'monthly_profit1': 0.0,
            'monthly_results': 0.0,
            'profit_per_order': 0.0,
            'head_distance': head_distance,
            'actual_cost': discounted_price - actual_profit,
            'ad_removed_roi': 0.0,
        }
    }


async def get_month_number(asin_list: list, data_nested: dict, key_list: dict):
    """
    按 ASIN 汇总目标月销（来自各关键词），再对关键词取平均。

    data_nested: {asin: {keyword: 目标月销数值}}
    key_list: {asin: [keyword, ...]}，须与扫描/业务关键词列表一致
    """
    asin_month = {}
    for a in asin_list:
        kws = key_list.get(a) or []
        if not kws:
            asin_month[a] = 0.0
            continue
        inner = data_nested.get(a) or {}
        total = sum(inner.get(kw, 0) for kw in kws)
        asin_month[a] = total / len(kws)
    return asin_month


# 并发下载图片到本地（绝对路径：media/images/<ASIN>.jpg，可用 ASIN_IMAGES_ROOT 覆盖）
async def download_one(asin, info, *, retries: int = 2):
    images_dir = resolve_asin_images_dir()
    images_dir.mkdir(parents=True, exist_ok=True)
    existing = asin_image_file(asin)
    if existing.is_file():
        print(f"图片已存在: {existing}")
        return str(existing)
    local_path = asin_image_file(asin, for_write=True)

    image_url = (info or {}).get('imageUrl') if isinstance(info, dict) else None
    if not image_url:
        print(f"警告: ASIN {asin} 无图片 URL，无法下载主图到 {local_path}")
        return None

    from sellersprite_market import (
        get_sellersprite_marketplace,
        sellersprite_amazon_host,
    )

    host = sellersprite_amazon_host(get_sellersprite_marketplace())
    last_err = ''
    for attempt in range(max(1, retries)):
        try:
            async with aiohttp.ClientSession(
                headers={
                    'User-Agent': (
                        'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 '
                        '(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'
                    ),
                    'Accept': 'image/avif,image/webp,image/apng,image/*,*/*;q=0.8',
                    'Referer': f'https://{host}/',
                }
            ) as session:
                timeout = aiohttp.ClientTimeout(total=90)
                async with session.get(image_url, timeout=timeout) as resp:
                    if resp.status == 200:
                        content = await resp.read()
                        if len(content) < 256:
                            last_err = f'主图过小({len(content)} bytes)'
                            print(f"下载 {asin} {last_err}，可能不是有效图片")
                        else:
                            local_path.write_bytes(content)
                            print(f"图片已下载: {local_path} ({len(content)} bytes)")
                            return str(local_path)
                    else:
                        last_err = f'HTTP {resp.status}'
                        print(f"下载失败 {asin}: {last_err} url={str(image_url)[:120]}")
        except Exception as e:
            last_err = str(e)
            print(f"下载异常 {asin} (attempt {attempt + 1}): {e}")
        if attempt + 1 < retries:
            await asyncio.sleep(0.6 * (attempt + 1))
    if last_err:
        print(f"警告: ASIN {asin} 主图下载最终失败: {last_err}")
    return None


async def _download_asin_images(
    asin_info_dict: dict,
    asin_to_image_path: dict | None = None,
) -> dict[str, str]:
    """为有广告数据的 ASIN 下载主图；已有本地文件则跳过。"""
    out = dict(asin_to_image_path or {})
    need: list[tuple[str, dict]] = []
    for asin, info in (asin_info_dict or {}).items():
        key = str(asin).strip().upper()
        if not key or not isinstance(info, dict):
            continue
        existing = out.get(key) or ''
        if existing and Path(existing).is_file():
            continue
        local = asin_image_file(key)
        if local.is_file():
            out[key] = str(local.resolve())
            continue
        need.append((key, info))
    if not need:
        return out
    paths = await asyncio.gather(*[download_one(a, info) for a, info in need])
    for (asin, _), path in zip(need, paths):
        if path:
            out[asin] = path
    return out


def _build_image_meta(
    target_asins: list[str],
    asin_info_dict: dict,
    asin_to_image_path: dict | None,
) -> dict[str, dict]:
    """供自动 ROI 校验本地主图与数据来源。"""
    meta: dict[str, dict] = {}
    for raw in target_asins or []:
        key = str(raw).strip().upper()
        if not key:
            continue
        info = asin_info_dict.get(key) if isinstance(asin_info_dict, dict) else None
        path = (asin_to_image_path or {}).get(key, '')
        has_local = bool(path and Path(path).is_file()) or asin_image_file(key).is_file()
        src = 'none'
        image_url = ''
        ad_data = isinstance(info, dict) and bool(info)
        if isinstance(info, dict):
            image_url = str(info.get('imageUrl') or '')[:240]
            src = str(info.get('_image_source') or ('ad_api' if image_url else 'none'))
        if has_local and src == 'none':
            src = 'local_existing'
        meta[key] = {
            'has_local_image': has_local,
            'image_source': src,
            'ad_data': ad_data,
            'image_url': image_url,
        }
    return meta


async def _prepare_asin_images(
    target_asins: list[str],
    asin_info_dict: dict,
    asin_to_image_path: dict | None,
    *,
    max_concurrent: int,
    marketplace: str,
) -> tuple[dict, dict[str, str]]:
    """competing-lookup 补 imageUrl 并下载到 media/images。"""
    from async_advertisement_api import supplement_image_urls_from_lookup

    info = await supplement_image_urls_from_lookup(
        asin_info_dict,
        target_asins,
        max_concurrent=max_concurrent,
        marketplace=marketplace,
    )
    paths = await _download_asin_images(info, asin_to_image_path)
    missing = [
        str(a).strip().upper()
        for a in target_asins
        if str(a).strip().upper() not in paths
        and not asin_image_file(str(a).strip().upper()).is_file()
    ]
    if missing:
        emit_progress(f'主图下载缺口 {len(missing)}，重试一次…')
        paths = await _download_asin_images(info, paths)
    return info, paths


async def _refill_missing_sellersprite(
    target_asins: list[str],
    fba_info_dict: dict,
    asin_info_dict: dict,
    *,
    max_concurrent: int = 6,
    max_rounds: int = 3,
    rotation_state: dict | None = None,
    pending_task: str = 'roi',
    marketplace: str | None = None,
) -> tuple[dict, dict]:
    """
    对首次批量拉取后缺失的广告/FBA 做多轮补拉（复用广告难度路径的 robust 逻辑）。
    遇禁号时尝试 bulk 换号后重试一次；仍失败则上抛，避免吞掉禁号继续写残缺包。
    """
    from async_advertisement_api import ensure_ads_cached_robust
    from async_fba_api import async_fba_batch
    from seller_account_guard import (
        SellerAccountBannedError,
        SellerSpriteTransientError,
        bulk_rotate_if_available,
        clear_seller_login_cache,
        is_seller_account_banned_error,
    )
    from sellersprite_market import normalize_sellersprite_marketplace

    mp = normalize_sellersprite_marketplace(marketplace)
    targets = [str(a).strip().upper() for a in (target_asins or []) if str(a).strip()]
    ads = dict(asin_info_dict or {})
    fba = dict(fba_info_dict or {})
    rot_state = rotation_state if rotation_state is not None else {'count': 0, 'max': 16}

    async def _with_ban_retry(label: str, coro_factory):
        try:
            return await coro_factory()
        except Exception as e:
            if not is_seller_account_banned_error(e):
                raise
            emit_progress(f'{label}遇禁号，尝试解禁换号后重试…')
            clear_seller_login_cache()
            if not await bulk_rotate_if_available(
                targets,
                rotation_state=rot_state,
                pending_task=pending_task,
            ):
                raise SellerAccountBannedError(f'{label}禁号且无法轮换: {e}') from e
            return await coro_factory()

    missing_ads = [a for a in targets if not isinstance(ads.get(a), dict)]
    if missing_ads:
        emit_progress(
            f'广告数据缺失 {len(missing_ads)}/{len(targets)}，开始补拉（最多 {max_rounds} 轮）…'
        )

        async def _pull_ads():
            return await ensure_ads_cached_robust(
                ads,
                missing_ads,
                max_concurrent=max_concurrent,
                max_rounds=max_rounds,
                pause_sec=2.0,
                marketplace=mp,
            )

        ads = await _with_ban_retry('广告补拉', _pull_ads)
        ads = {
            str(k).strip().upper(): v
            for k, v in (ads or {}).items()
            if isinstance(v, dict)
        }
        still_ads = [a for a in targets if a not in ads]
        if still_ads:
            emit_progress(f'警告：补拉后仍缺广告数据 {len(still_ads)} 个（将影响主图/采购价）')
            for a in still_ads[:12]:
                emit_progress(f'  缺广告: {a}')
            if len(still_ads) > 12:
                emit_progress(f'  … 另有 {len(still_ads) - 12} 个')
        else:
            emit_progress(f'广告补拉完成：{len(ads)}/{len(targets)}')

    def _fba_ok(val) -> bool:
        return isinstance(val, dict) and (
            val.get('FBA') is not None or val.get('head_distance') is not None
        )

    missing_fba = [a for a in targets if not _fba_ok(fba.get(a))]
    if missing_fba:
        emit_progress(f'FBA 数据缺失 {len(missing_fba)}/{len(targets)}，开始补拉…')

        async def _pull_fba():
            return await async_fba_batch(
                missing_fba, max_concurrent=max_concurrent, marketplace=mp
            )

        try:
            extra = await _with_ban_retry('FBA 补拉', _pull_fba)
            if isinstance(extra, dict):
                for k, v in extra.items():
                    key = str(k).strip().upper()
                    if _fba_ok(v):
                        fba[key] = v
        except Exception as e:
            if is_seller_account_banned_error(e):
                raise
            emit_progress(f'警告：FBA 补拉失败：{e}')
            print(f'FBA 补拉失败: {e}')

    return fba, ads


async def async_return_info(asin_dict: dict, info_list: list):
    """
    将 monthly_results（按关键词）聚合为按 ASIN 的结构：
    {
      ASIN: {
        "monthly_results": 该 ASIN 下关键词目标月销最大值,
        "review_interval": {
          ASIN: {
            keyword: {"review_interval": <该关键词区间分析字典>}
          }
        },
        "ranking_percent": 该 ASIN 下关键词 ranking_percent 的最小值（越小越优）
      }
    }
    """
    result_dict: Dict[str, Dict[str, Any]] = {}
    try:
        for asin, kw_list in asin_dict.items():
            result_dict[asin] = {
                "monthly_results": 0.0,
                "review_interval": {asin: {}},
                "ranking_percent": 0.0,
            }
            max_monthly = -1.0
            rp_candidates: List[float] = []

            for item in info_list:
                if not item:
                    continue
                try:
                    item_asin = item.get("asin")
                    if item_asin and item_asin != asin:
                        continue

                    monthly_map = item.get("monthly_results", {})
                    if not monthly_map:
                        continue
                    kw = list(monthly_map.keys())[0]
                    if kw not in kw_list:
                        continue

                    month_val = pd.to_numeric(monthly_map.get(kw), errors="coerce")
                    if not pd.isna(month_val) and float(month_val) > max_monthly:
                        max_monthly = float(month_val)

                    review_payload = item.get("review_interval", {})
                    result_dict[asin]["review_interval"][asin][kw] = {
                        "review_interval": review_payload
                    }

                    rp_payload = item.get("ranking_percent", -1)
                    rp_num = pd.to_numeric(rp_payload, errors="coerce")
                    if not pd.isna(rp_num) and rp_num >= 0:  # 允许 0 作为默认值
                        rp_candidates.append(float(rp_num))
                except Exception as e:
                    print(f"<UNK> {asin_dict}: {e}")

            if max_monthly >= 0:
                result_dict[asin]["monthly_results"] = max_monthly
            if rp_candidates:
                result_dict[asin]["ranking_percent"] = round(min(rp_candidates), 3)
    except Exception as e:
        print(f"<不好。数据出问题了1> {asin_dict}: {e}")
    return result_dict


async def async_merging_data(info_list: list, info_dict: dict):
    """
    合并 ROI 结果到 info_dict（按 ASIN 键合并）。
    保留 async_return_info 生成的 ranking_percent 嵌套结构，不做重写。
    """
    try:
        for info in info_list:
            if not info:
                continue
            asin_key = list(info.keys())[0]
            converted_info = {
                k: round(float(v), 2) if isinstance(v, np.float64) else v
                for k, v in info[asin_key].items()
            }
            if asin_key not in info_dict:
                info_dict[asin_key] = {}
            info_dict[asin_key].update(converted_info)

        for key in info_dict:
            info_dict[key] = {
                k: round(float(v), 2) if isinstance(v, np.float64) else v
                for k, v in info_dict[key].items()
            }
    except Exception as e:
        print(f"<不好，数据出问题了3> {info_dict}: {e}")
    return info_dict


ASIN_FOLDER_PATTERN = re.compile(r"^B[A-Z0-9]{9}$")


def normalize_excel_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    """列名去空格、按 COLUMN_MAPPING 重命名；多列映射到同一 API 名时保留首列。"""
    out = df.copy()
    out.columns = pd.Index([str(c).strip() for c in out.columns])
    rename_map = {
        k: v for k, v in COLUMN_MAPPING_EXCEL_TO_API.items() if k in out.columns
    }
    out = out.rename(columns=rename_map)
    if out.columns.duplicated().any():
        out = out.loc[:, ~out.columns.duplicated(keep="first")]
    return out


def load_products_from_local_files(
    base_dir: Optional[Path] = None,
    asins: list[str] | None = None,
    *,
    marketplace: str = 'US',
) -> tuple:
    """
    自动扫描本地目录：file/{ASIN}/{关键词文件夹}/Search(*)-*-{US|UK}-*.xlsx
    不依赖外部传入的关键词列表；ASIN 与关键词均来自文件夹名。

    :param base_dir: 数据根目录，默认 FILE_DATA_ROOT（即 ./file）
    :param asins: 若指定则只扫描这些 ASIN 目录（避免 media 下数百个 ASIN 全量读 Excel）
    :param marketplace: US|UK，优先匹配对应站点的 Search 导出文件
    :return: (nested_result, keyword_dict, source_map)
        nested_result: {asin: {keyword_folder_name: [product_dict, ...]}}
        keyword_dict: {asin: [keyword_folder_name, ...]}，供后续加权月销等逻辑使用
        source_map: {asin: {keyword_folder_name: 'data_origin'|'search'|'xlsx'}}
    """
    mp = str(marketplace or 'US').strip().upper()
    if mp not in ('US', 'UK'):
        mp = 'US'
    search_glob = f'Search(*)-*-{mp}-*.xlsx'
    root = Path(base_dir) if base_dir is not None else FILE_DATA_ROOT
    nested: Dict[str, Dict[str, list]] = {}
    keyword_dict: Dict[str, list] = {}

    source_map: Dict[str, Dict[str, str]] = {}
    if not root.exists():
        print(f"警告：数据根目录不存在: {root.resolve()}")
        return nested, keyword_dict, source_map

    if asins is not None:
        allow = {str(x).strip().upper() for x in asins if str(x).strip()}
        if not allow:
            print('定向扫描：ASIN 列表为空，跳过读盘')
            return nested, keyword_dict, source_map
        asin_dirs: list[Path] = []
        for asin in sorted(allow):
            if not ASIN_FOLDER_PATTERN.match(asin):
                print(f'警告：跳过非法 ASIN 目录名: {asin}')
                continue
            asin_dir = root / asin
            if asin_dir.is_dir():
                asin_dirs.append(asin_dir)
            else:
                print(f'警告：ASIN 目录不存在: {asin_dir}')
        print(f'定向扫描：仅读取 {len(asin_dirs)} 个 ASIN 目录（共指定 {len(allow)} 个）')
    else:
        asin_dirs = [
            d for d in sorted(root.iterdir())
            if d.is_dir() and ASIN_FOLDER_PATTERN.match(d.name.upper())
        ]
        print(f'全量扫描：未指定 ASIN，遍历 {len(asin_dirs)} 个目录')

    for asin_dir in asin_dirs:
        asin = asin_dir.name.upper()

        nested[asin] = {}
        keyword_dict[asin] = []
        source_map[asin] = {}
        has_roi_sheet = any(
            p.is_file()
            and p.suffix.lower() == ".xlsx"
            and ("roi-us" in p.name.lower() or "roi_us" in p.name.lower())
            for p in asin_dir.glob("*.xlsx")
        )

        for kw_dir in sorted(asin_dir.iterdir()):
            if not kw_dir.is_dir():
                continue
            kw_name = kw_dir.name

            source_kind = "search"
            if has_roi_sheet:
                excel_files = sorted(kw_dir.glob("*_data_origin.xlsx"))
                if excel_files:
                    source_kind = "data_origin"
                else:
                    excel_files = sorted(kw_dir.glob(search_glob))
                    source_kind = "search"
            else:
                excel_files = sorted(kw_dir.glob(search_glob))
            if not excel_files:
                # 回落：另一站点 Search，再任意 xlsx
                alt = 'UK' if mp == 'US' else 'US'
                excel_files = sorted(kw_dir.glob(f'Search(*)-*-{alt}-*.xlsx'))
                source_kind = "search"
            if not excel_files:
                excel_files = sorted(kw_dir.glob("*.xlsx"))
                source_kind = "xlsx"
            if not excel_files:
                print(f"警告：{kw_dir} 中未找到 xlsx")
                continue

            excel_path = excel_files[0]
            print(f"正在读取文件：{excel_path}")
            try:
                df = pd.read_excel(excel_path)
                df = normalize_excel_dataframe(df)
                for col in ["price", "totalUnits", "reviews"]:
                    if col in df.columns:
                        df[col] = pd.to_numeric(df[col], errors="coerce")
                products = df.to_dict(orient="records")
                nested[asin][kw_name] = products
                keyword_dict[asin].append(kw_name)
                source_map[asin][kw_name] = source_kind
            except Exception as e:
                print(f"读取 Excel 失败 {excel_path}: {e}")

    return nested, keyword_dict, source_map


async def seller_wizard_main(
        parity: float,
        asins: list[str] | None = None,
        cost_overrides: dict | None = None,
        marketplace: str = 'US',
        roi_defaults: dict | None = None,
):
    emit_progress('seller_wizard_main 已开始…')
    mp = str(marketplace or 'US').strip().upper()
    if mp not in ('US', 'UK'):
        mp = 'US'
    from sellersprite_market import (
        reset_sellersprite_marketplace,
        sellersprite_market_com,
        sellersprite_market_id,
        set_sellersprite_marketplace,
    )

    _ss_token = set_sellersprite_marketplace(mp)
    rd = roi_defaults if isinstance(roi_defaults, dict) else {}
    try:
        _pc = float(rd['platform_commission']) if 'platform_commission' in rd else (
            25 if mp == 'UK' else 15
        )
    except (TypeError, ValueError):
        _pc = 25 if mp == 'UK' else 15
    emit_progress(
        f'当前站点：{mp}（平台佣金 {_pc:g}%；'
        f'SellerSprite marketId={sellersprite_market_id(mp)} / market={sellersprite_market_com(mp)}）'
    )

    try:
        return await _seller_wizard_main_body(
            parity,
            asins=asins,
            cost_overrides=cost_overrides,
            marketplace=mp,
            roi_defaults=rd,
        )
    finally:
        reset_sellersprite_marketplace(_ss_token)


async def _seller_wizard_main_body(
        parity: float,
        asins: list[str] | None = None,
        cost_overrides: dict | None = None,
        marketplace: str = 'US',
        roi_defaults: dict | None = None,
):
    mp = str(marketplace or 'US').strip().upper()
    if mp not in ('US', 'UK'):
        mp = 'US'
    rd = roi_defaults if isinstance(roi_defaults, dict) else {}
    from sellersprite_market import sellersprite_amazon_dp_url

    def _finish(payload):
        return payload

    FILE_DATA_ROOT.mkdir(parents=True, exist_ok=True)
    print(f"本地 Excel 根目录: {FILE_DATA_ROOT.resolve()}")
    try:
        taobao_config = await asyncio.to_thread(get_taobao_tokens)
        tokens = [taobao_config["_m_h5_tk"], taobao_config["_m_h5_tk_enc"]]
    except Exception as e:
        print(f"警告: 获取淘宝 token 失败，图搜将尝试回落已有采购价: {e}")
        tokens = []

    # 1. 从 file/{ASIN}/{关键词}/ 读取 Excel（指定 asins 时只扫对应目录，不全库遍历）
    try:
        from bulk_account_pool import pop_ban_pending_asins

        pending_resume = pop_ban_pending_asins(task='roi')
        if pending_resume:
            if asins:
                asins = sorted({str(x).strip().upper() for x in asins if str(x).strip()} | set(pending_resume))
            else:
                asins = pending_resume
            emit_progress(f'续算上次因禁号未完成的 {len(pending_resume)} 个 ASIN…')
    except ImportError:
        pass
    if asins:
        n = len({str(x).strip().upper() for x in asins if str(x).strip()})
        emit_progress(f'正在扫描本地 Excel（仅 {n} 个目标 ASIN，不全库扫盘）…')
    else:
        emit_progress('正在扫描本地 Excel（未指定 ASIN，全库扫盘）…')
    nested_result, keyword_dict, source_map = await asyncio.to_thread(
        load_products_from_local_files, None, asins, marketplace=mp
    )
    target_asins = [str(a).strip().upper() for a in nested_result.keys()]
    emit_progress(f'已扫描本地数据：共 {len(target_asins)} 个 ASIN（读盘完成）')
    print("从本地扫描得到的 keyword_dict:", keyword_dict)
    if not target_asins:
        emit_progress(f'未在 {FILE_DATA_ROOT} 下发现 ASIN 数据')
        print(f"未在 {FILE_DATA_ROOT} 下发现符合 B0XXXXXXXXX 结构的 ASIN 数据，结束。")
        return _finish({})

    from seller_account_guard import (
        SellerAccountBannedError,
        SellerSpriteTransientError,
        bulk_rotate_if_available,
        clear_seller_login_cache,
        ensure_seller_login,
        is_seller_account_banned_error,
    )

    rotation_state: dict = {'count': 0, 'max': 16}

    emit_progress('正在登录卖家精灵并批量拉取 FBA / 广告数据…')
    await ensure_seller_login()
    max_ss = env_max_concurrent('sellersprite', 6)

    async def _fetch_fba_and_ad(batch_asins: list[str]):
        while True:
            try:
                async with async_api_slot('sellersprite'):
                    return await asyncio.gather(
                        async_fba_batch(
                            batch_asins, max_concurrent=max_ss, marketplace=mp
                        ),
                        advertisement_main(
                            batch_asins, max_concurrent=max_ss, marketplace=mp
                        ),
                    )
            except Exception as e:
                if isinstance(e, SellerSpriteTransientError):
                    raise
                if not is_seller_account_banned_error(e):
                    raise RuntimeError(f'卖家精灵 FBA/广告批量获取失败: {e}') from e
                emit_progress('卖家精灵会话失效或子账号被禁，正在解禁并切换批量账号…')
                clear_seller_login_cache()
                if not await bulk_rotate_if_available(
                    batch_asins,
                    rotation_state=rotation_state,
                    pending_task='roi',
                ):
                    # single 或池不足：至少解禁当前号再试一次
                    try:
                        from unlock_seller_info import activate_children

                        activate_children()
                        clear_seller_login_cache()
                        await ensure_seller_login(force_refresh=True)
                        async with async_api_slot('sellersprite'):
                            return await asyncio.gather(
                                async_fba_batch(
                                    batch_asins, max_concurrent=max_ss, marketplace=mp
                                ),
                                advertisement_main(
                                    batch_asins, max_concurrent=max_ss, marketplace=mp
                                ),
                            )
                    except Exception as e2:
                        raise SellerAccountBannedError(str(e2 or e)) from e
                continue

    fba_info_dict, asin_info_dict = await _fetch_fba_and_ad(target_asins)
    print(fba_info_dict, '666666')
    if not isinstance(asin_info_dict, dict):
        asin_info_dict = {}
    asin_info_dict = {
        str(k).strip().upper(): v
        for k, v in asin_info_dict.items()
        if isinstance(v, dict)
    }
    if isinstance(fba_info_dict, dict):
        fba_info_dict = {
            str(k).strip().upper(): v
            for k, v in fba_info_dict.items()
        }
    else:
        fba_info_dict = {}
    emit_progress(
        f'卖家精灵首轮：FBA {sum(1 for v in fba_info_dict.values() if isinstance(v, dict))}/{len(target_asins)}，'
        f'广告 {len(asin_info_dict)}/{len(target_asins)}'
    )
    fba_info_dict, asin_info_dict = await _refill_missing_sellersprite(
        target_asins,
        fba_info_dict,
        asin_info_dict,
        max_concurrent=max_ss,
        max_rounds=3,
        rotation_state=rotation_state,
        pending_task='roi',
        marketplace=mp,
    )
    emit_progress(
        f'卖家精灵数据就绪：FBA {sum(1 for v in fba_info_dict.values() if isinstance(v, dict))}/{len(target_asins)}，'
        f'广告 {len(asin_info_dict)}/{len(target_asins)}'
    )

    asin_info_dict, asin_to_image_path = await _prepare_asin_images(
        target_asins,
        asin_info_dict,
        {},
        max_concurrent=max_ss,
        marketplace=mp,
    )

    # 2. SIF：CPC 等（关键词以本地扫描为准；英国站 country=UK）
    emit_progress(f'正在请求 SIF 广告 CPC / 转化率（站点 {mp}，{len(target_asins)} 个 ASIN）…')
    try:
        # Django 侧解析 authorization JWT exp；过期则直接失败，避免整批默默用 $1/10%
        try:
            from django.conf import settings as dj_settings

            base = Path(getattr(dj_settings, 'BASE_DIR', Path.cwd()))
            auth_path = base / 'scripts' / 'asin_find_project' / 'config_file' / 'sif_authorization.txt'
            if auth_path.is_file():
                import base64
                import json
                from datetime import datetime, timezone as dt_tz

                jwt = auth_path.read_text(encoding='utf-8').strip()
                if jwt.count('.') >= 2:
                    payload_b64 = jwt.split('.')[1]
                    pad = '=' * (-len(payload_b64) % 4)
                    payload = json.loads(base64.urlsafe_b64decode(payload_b64 + pad).decode('utf-8'))
                    exp_ts = int(payload.get('exp') or 0)
                    if exp_ts > 0:
                        left = exp_ts - int(datetime.now(tz=dt_tz.utc).timestamp())
                        exp_label = datetime.fromtimestamp(exp_ts).strftime('%Y-%m-%d %H:%M:%S')
                        if left <= 0:
                            msg = (
                                f'SIF authorization（JWT）已过期（到期 {exp_label}）。'
                                '请到「凭证配置」更新 JWT 并刷新 Token 后重试；'
                                '本次不会继续用默认 $1/10% 掩盖问题。'
                            )
                            emit_progress(msg)
                            raise RuntimeError(msg)
                        if left <= 48 * 3600:
                            emit_progress(
                                f'警告：SIF JWT 将在约 {max(1, left // 3600)} 小时内过期（{exp_label}），请尽快更新'
                            )
        except RuntimeError:
            raise
        except Exception as e:
            emit_progress(f'SIF JWT 到期检查跳过：{e}')

        async with async_api_slot('sif'):
            asin_cpc, _ = await async_sif_api.sif_main(target_asins, country=mp)
    except Exception as e:
        emit_progress(f'SIF CPC 获取失败：{e}')
        print(f"警告: SIF CPC 获取失败: {e}")
        # JWT 过期属于配置错误，整批中止，避免自动/手动 ROI 静默默认值
        err_s = str(e)
        if 'JWT' in err_s or 'authorization' in err_s.lower() or '已过期' in err_s:
            raise
        emit_progress(f'SIF 失败，将使用默认 CPC=$1 / 转化率=10%：{e}')
        asin_cpc = []
    if not isinstance(asin_cpc, list):
        asin_cpc = []
    sif_ok = 0
    for item in asin_cpc:
        if not isinstance(item, dict):
            continue
        for _k, info in item.items():
            if not isinstance(info, dict):
                continue
            cpc = info.get('cpc') if isinstance(info.get('cpc'), dict) else {}
            med = cpc.get('median')
            ratio = info.get('clickPurchaseRatio')
            try:
                if (med is not None and float(med) > 0) or (
                    ratio is not None and float(ratio) > 0
                ):
                    sif_ok += 1
                    break
            except (TypeError, ValueError):
                pass
    emit_progress(
        f'SIF 结果：{sif_ok}/{len(target_asins)} 个 ASIN 有可用 CPC/转化率'
        + ('（其余将用默认 $1 / 10%）' if sif_ok < len(target_asins) else '')
    )

    asin_path_dict = await collect_node_label_paths(keyword_dict, nested_result, target_asins)

    emit_progress('正在批量获取退款率…')
    refund_cache = await prefetch_refund_rates(
        list(asin_path_dict.values()),
        max_concurrent=max_ss,
        marketplace=mp,
    )
    emit_progress(f'退款率预取完成：{len(refund_cache)} 个类目')

    roi_failures: list[dict[str, str]] = []

    # ========== 3. 并发处理每个 (ASIN, 关键词) ==========
    async def process_one_keyword(asin: str, keyword: str, products: list):
        """清洗 -> 市场容量 -> 评论区间 -> 广告效率表，输出写入 file/{asin}/"""
        try:
            print(f"\n处理 ASIN={asin} 关键词: {keyword}")
            print(f"清洗前的数据行数：{len(products)}")
            df = pd.DataFrame(products)
            ranking_percent = 0
            if df.empty:
                return None
            source_kind = (source_map.get(asin, {}) or {}).get(keyword, "search")
            if source_kind == "data_origin":
                clean_data = df
                print(f"ASIN={asin} 关键词={keyword} 检测到 ROI 表，直接使用 data_origin 作为数据源。")
            else:
                price_info = asin_info_dict.get(str(asin).strip().upper()) or {}
                clean_data = await save_cleaned_data_orign_to_excel(
                    df, keyword, asin, price_info if isinstance(price_info, dict) else None
                )
            target_monthly = await save_top5_market_capacity_to_excel(clean_data, keyword, asin)
            review_interval = await save_review_interval_analysis_to_excel(
                clean_data, keyword, asin, marketplace=mp
            )
            ranking_percent = 0

            print(f"<UNK> review_interval={review_interval}")
            return {
                "asin": asin,
                "keyword": keyword,
                'review_interval': review_interval,
                "monthly_results": {keyword: target_monthly},
                "ranking_percent": ranking_percent,
            }
        except Exception as e:
            print(f"警告: ASIN={asin} 关键词={keyword} 处理失败，已跳过: {e}")
            return None

    keyword_tasks = [
        process_one_keyword(asin, kw, products)
        for asin, kw_map in nested_result.items()
        for kw, products in kw_map.items()
    ]
    if not keyword_tasks:
        print(
            f"警告：未加载到任何关键词 Excel 数据（请确认 xlsx 在 {FILE_DATA_ROOT}/{{ASIN}}/{{关键词}}/ 下，"
            f"且从脚本/项目目录运行）。"
        )
    monthly_results_raw = await asyncio.gather(*keyword_tasks, return_exceptions=True)
    monthly_results = []
    for item in monthly_results_raw:
        if isinstance(item, Exception):
            print(f"警告: 关键词任务异常: {item}")
            continue
        if item:
            monthly_results.append(item)
    print(monthly_results, "wangxian1")

    info_dict = await async_return_info(asin_dict=keyword_dict, info_list=monthly_results)
    print(info_dict, "ppppp")

    target_monthly_sales: Dict[str, Dict[str, float]] = {}
    for res in monthly_results:
        if not res:
            continue
        a = res["asin"]
        target_monthly_sales.setdefault(a, {}).update(res["monthly_results"])
    print(target_monthly_sales)

    monthly_sales_dict = await get_month_number(target_asins, target_monthly_sales, keyword_dict)
    print(monthly_sales_dict)
    emit_progress(f'关键词 Excel 处理完成，开始生成 ROI-US-pack（0/{len(target_asins)}）…')

    async def save_roi_safe(
        asin: str,
        node_label_path: str,
        info: dict,
        path: str,
        up_val,
        hd_val,
    ):
        try:
            hints = build_local_product_hints(nested_result, asin)
            result = await save_roi_us_pack(
                node_label_path,
                fba_info_dict,
                asin,
                asin_cpc,
                monthly_sales_dict,
                tokens,
                path,
                parity,
                info,
                unit_purchase_override=up_val,
                head_distance_override=hd_val,
                local_hints=hints,
                refund_cache=refund_cache,
                marketplace=mp,
                roi_defaults=rd,
            )
            return result
        except Exception as e:
            if is_seller_account_banned_error(e):
                raise
            err = f'{type(e).__name__}: {e}'
            print(f"警告: ASIN {asin} ROI-US-pack 生成失败: {e}，尝试强制写入默认值…")
            try:
                return await force_write_minimal_roi_us_pack(
                    asin, exchange_rate=parity, marketplace=mp, roi_defaults=rd
                )
            except Exception as e2:
                roi_failures.append({'asin': asin, 'error': f'{err}; 兜底失败: {e2}'})
                emit_progress(f'失败 {asin}：{e2}')
                print(f"警告: ASIN {asin} 强制生成 ROI-US-pack 仍失败: {e2}")
                return None

    async def _build_roi_tasks(asin_list: list[str]):
        tasks = []
        keys: list[str] = []
        for asin in asin_list:
            asin_key = str(asin).strip().upper()
            info = dict(asin_info_dict.get(asin_key) or {})
            if not info:
                print(f"警告: ASIN {asin_key} 无卖家精灵 advertisement 数据，将用默认值生成 ROI-US-pack")
                emit_progress(f'提示 {asin_key}：无广告接口数据，主图/采购价可能为默认值')
            elif not str(info.get('imageUrl') or '').strip():
                emit_progress(f'提示 {asin_key}：广告数据无 imageUrl，尝试本地/旧表主图')
            path = asin_to_image_path.get(asin_key, "")
            co = (cost_overrides or {}).get(asin) or (cost_overrides or {}).get(asin_key) or {}
            up = pd.to_numeric(co.get('unit_purchase'), errors='coerce') if isinstance(co, dict) else np.nan
            hd = pd.to_numeric(co.get('head_distance'), errors='coerce') if isinstance(co, dict) else np.nan
            up_val = None if pd.isna(up) else float(up)
            hd_val = None if pd.isna(hd) else float(hd)
            keys.append(asin_key)
            tasks.append(
                save_roi_safe(asin_key, asin_path_dict.get(asin_key, ""), info, path, up_val, hd_val)
            )
        return keys, tasks

    remaining_roi = list(target_asins)
    info_result: list = []
    done_count = 0

    while remaining_roi:
        roi_task_asins, roi_tasks = await _build_roi_tasks(remaining_roi)
        if not roi_tasks:
            break

        info_result_raw = await asyncio.gather(*roi_tasks, return_exceptions=True)
        ban_pending: list[str] | None = None
        for asin_key, item in zip(roi_task_asins, info_result_raw):
            if isinstance(item, Exception) and is_seller_account_banned_error(item):
                done_keys = {
                    a for a, it in zip(roi_task_asins, info_result_raw)
                    if it and not isinstance(it, Exception)
                }
                ban_pending = [a for a in roi_task_asins if a not in done_keys]
                break

        for asin_key, item in zip(roi_task_asins, info_result_raw):
            if ban_pending and asin_key in ban_pending:
                continue
            if isinstance(item, Exception):
                err = f'{type(item).__name__}: {item}'
                roi_failures.append({'asin': asin_key, 'error': err})
                emit_progress(f'失败 {asin_key}：{item}')
                print(f"警告: ROI 任务异常 ({asin_key}): {item}")
                continue
            if item:
                info_result.append(item)
                done_count += 1
                emit_progress(f'进度 {done_count}/{len(target_asins)}：{asin_key} ROI 已完成')
            else:
                if not any(f.get('asin') == asin_key for f in roi_failures):
                    roi_failures.append({'asin': asin_key, 'error': 'ROI-US-pack 生成失败'})

        if ban_pending:
            emit_progress('ROI 生成中检测到子账号被禁，正在切换批量账号并续算…')
            clear_seller_login_cache()
            if not await bulk_rotate_if_available(
                ban_pending,
                rotation_state=rotation_state,
                pending_task='roi',
            ):
                from bulk_account_pool import record_ban_pending_asins

                record_ban_pending_asins(ban_pending, task='roi')
                for a in ban_pending:
                    if not any(f.get('asin') == a for f in roi_failures):
                        roi_failures.append(
                            {'asin': a, 'error': '账号被禁且无法轮换，已记入续算队列'}
                        )
                merging_data = await async_merging_data(info_result, info_dict)
                if not isinstance(merging_data, dict):
                    merging_data = {}
                merging_data['__roi_failures__'] = roi_failures
                merging_data['__ban__'] = {
                    'pending': ban_pending,
                    'message': '账号被禁且无法轮换',
                }
                if info_result:
                    emit_progress(
                        f'无法换号：已保留成功 {len(info_result)} 个，'
                        f'待续算 {len(ban_pending)} 个'
                    )
                    return _finish(merging_data)
                raise SellerAccountBannedError(
                    '账号被禁且无法轮换',
                    partial_results=merging_data,
                    ban_pending=ban_pending,
                )
            await ensure_seller_login()
            # 换号后补拉缺失的卖家精灵数据与主图，避免续算仍用空广告字典
            fba_info_dict, asin_info_dict = await _refill_missing_sellersprite(
                ban_pending,
                fba_info_dict,
                asin_info_dict,
                max_concurrent=max_ss,
                max_rounds=3,
                rotation_state=rotation_state,
                pending_task='roi',
                marketplace=mp,
            )
            asin_info_dict, asin_to_image_path = await _prepare_asin_images(
                ban_pending,
                asin_info_dict,
                asin_to_image_path,
                max_concurrent=max_ss,
                marketplace=mp,
            )
            remaining_roi = ban_pending
            continue
        break

    print(info_result, 'wangxian2')

    merging_data = await async_merging_data(info_result, info_dict)
    if not isinstance(merging_data, dict):
        merging_data = {}
    merging_data['__roi_failures__'] = roi_failures
    merging_data['__image_meta__'] = _build_image_meta(
        target_asins, asin_info_dict, asin_to_image_path
    )
    ok_n = len(info_result)
    fail_n = len(roi_failures)
    emit_progress(f'全部完成：成功 {ok_n} 个，失败 {fail_n} 个')
    if roi_failures:
        for row in roi_failures[:15]:
            emit_progress(f"  失败 {row.get('asin')}: {row.get('error')}")
        if len(roi_failures) > 15:
            emit_progress(f'  … 另有 {len(roi_failures) - 15} 个失败 ASIN')
    print(merging_data, 'www')
    print("所有表创建完成！")
    return _finish(merging_data)


if __name__ == "__main__":
    asyncio.run(seller_wizard_main(6.88))
