from __future__ import annotations

import re
from io import BytesIO
from pathlib import Path

# Amazon ASIN 为 10 位大写字母与数字（常见以 B0 等开头）
ASIN_RE = re.compile(r'\b([A-Z0-9]{10})\b')


def extract_asins_from_text(text: str) -> list[str]:
    if not text or not text.strip():
        return []
    found = ASIN_RE.findall(text.upper())
    seen: set[str] = set()
    out: list[str] = []
    for a in found:
        if a not in seen:
            seen.add(a)
            out.append(a)
    return out


def extract_asins_from_file(path: Path) -> list[str]:
    ext = path.suffix.lower()
    if ext in ('.txt', '.csv'):
        text = path.read_text(encoding='utf-8', errors='ignore')
        return extract_asins_from_text(text)
    if ext in ('.xlsx', '.xls'):
        from openpyxl import load_workbook

        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        parts: list[str] = []
        for row in ws.iter_rows(values_only=True):
            for c in row:
                if c is not None:
                    parts.append(str(c))
        wb.close()
        return extract_asins_from_text(' '.join(parts))
    return []


def extract_asins_from_upload(ext: str, raw: bytes) -> list[str]:
    """从上传文件字节内容解析 ASIN，不落盘。"""
    ext = ext.lower()
    if ext in ('.txt', '.csv'):
        text = raw.decode('utf-8', errors='ignore')
        return extract_asins_from_text(text)
    if ext in ('.xlsx', '.xls'):
        from openpyxl import load_workbook

        bio = BytesIO(raw)
        wb = load_workbook(bio, read_only=True, data_only=True)
        ws = wb.active
        parts: list[str] = []
        for row in ws.iter_rows(values_only=True):
            for c in row:
                if c is not None:
                    parts.append(str(c))
        wb.close()
        return extract_asins_from_text(' '.join(parts))
    return []


def merge_asin_lists(*lists: list[str]) -> list[str]:
    seen: set[str] = set()
    out: list[str] = []
    for lst in lists:
        for a in lst:
            if a not in seen:
                seen.add(a)
                out.append(a)
    return out


def product_grade_from_monthly_profit1(v: float | None) -> str:
    """monthly_profit1 为美元量级时：A/B/C/D/E。"""
    print(v,'1qqq')
    if v is None:
        return 'E'
    try:
        x = float(v)
    except (TypeError, ValueError):
        return 'E'
    if x > 12000:
        return 'A'
    if x > 8000:
        return 'B'
    if x > 4000:
        return 'C'
    if x > 2000:
        return 'D'
    return 'E'


def count_asins_in_text(text: str) -> int:
    """与 extract_asins_from_text 一致，统计其中识别到的 ASIN 个数。"""
    return len(extract_asins_from_text(text))




