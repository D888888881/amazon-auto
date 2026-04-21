"""Excel 读取（样式/图片）与保留格式的保存。"""
import base64

from openpyxl import load_workbook
from openpyxl.utils import range_boundaries
def _cell_to_str(value):
    if value is None:
        return ''
    return str(value)


def _color_to_css(color):
    if color is None:
        return None
    try:
        if getattr(color, 'rgb', None):
            rgb = color.rgb
            if isinstance(rgb, str) and len(rgb) >= 6:
                if len(rgb) == 8:
                    return '#' + rgb[2:8]
                return '#' + rgb
        if getattr(color, 'indexed', None) is not None and color.indexed:
            return None
    except Exception:
        pass
    return None


def serialize_cell_style(cell) -> dict:
    """序列化为前端可用的样式字典。"""
    st: dict = {}
    try:
        f = cell.font
        if f:
            st['font_bold'] = bool(f.bold)
            st['font_italic'] = bool(f.italic)
            if f.sz:
                st['font_size'] = float(f.sz)
            c = _color_to_css(f.color)
            if c:
                st['color'] = c
    except Exception:
        pass
    try:
        fill = cell.fill
        if fill and fill.fill_type and fill.fill_type != 'none':
            c = _color_to_css(fill.fgColor) or _color_to_css(fill.bgColor)
            if c:
                st['bg'] = c
    except Exception:
        pass
    try:
        al = cell.alignment
        if al:
            if al.horizontal:
                st['h_align'] = al.horizontal
            if al.vertical:
                st['v_align'] = al.vertical
            if al.wrap_text:
                st['wrap'] = True
    except Exception:
        pass
    try:
        if cell.number_format and cell.number_format != 'General':
            st['num_fmt'] = cell.number_format
    except Exception:
        pass
    try:
        b = cell.border
        if b and b.left and b.left.style:
            st['border'] = True
    except Exception:
        pass
    return st


def style_dict_to_css(st: dict) -> str:
    parts = []
    if st.get('font_bold'):
        parts.append('font-weight:700')
    if st.get('font_italic'):
        parts.append('font-style:italic')
    if st.get('font_size'):
        parts.append('font-size:%spx' % st['font_size'])
    if st.get('color'):
        parts.append('color:%s' % st['color'])
    if st.get('bg'):
        parts.append('background-color:%s' % st['bg'])
    if st.get('h_align'):
        ha = st['h_align']
        if ha == 'center':
            parts.append('text-align:center')
        elif ha in ('right',):
            parts.append('text-align:right')
        else:
            parts.append('text-align:left')
    if st.get('wrap'):
        parts.append('white-space:pre-wrap;word-break:break-word')
    if st.get('border'):
        parts.append('border:1px solid #d0d0d0')
    return ';'.join(parts)


def _top_left_for_cell(ws, row: int, col: int) -> tuple[int, int]:
    """合并区域内统一用左上角单元格表示值与样式。"""
    merged = getattr(ws, "merged_cells", None)
    if merged is None:
        return row, col
    for merged_range in merged.ranges:
        try:
            min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
        except Exception:
            continue
        if min_row <= row <= max_row and min_col <= col <= max_col:
            return min_row, min_col
    return row, col


def read_active_sheet_rich(path) -> tuple:
    """
    返回 (payload_dict, None) 或 失败时 (None, error_str)
    payload: sheet_name, rows, images, merge_ranges, max_row, max_col
    rows: [[{v, style, css}, ...], ...]
    """
    # Excel 本体对部分 xlsx 有“容错/自动修复”能力，但 openpyxl 对 XML 更严格；
    # 因此这里做两次尝试，并在必要时降级为“仅值”读取，保证页面至少能打开与编辑。
    wb = None
    first_err = None
    try:
        wb = load_workbook(path, read_only=False, data_only=True, keep_links=True)
    except Exception as e:
        first_err = e
        try:
            # 更宽松：不保留外部链接，且用只读模式减少对复杂对象的解析
            wb = load_workbook(path, read_only=True, data_only=True, keep_links=False)
        except Exception as e2:
            return None, str(e2)
    try:
        ws = wb.active

        # ReadOnlyWorksheet 对 ws.cell / merged_cells 等支持不完整；
        # 在只读降级模式下改用 iter_rows，确保至少能把“值”读出来。
        if getattr(wb, "read_only", False):
            tmp_rows = []
            max_col = 1
            for r in ws.iter_rows(values_only=True):
                rr = [_cell_to_str(v) for v in (r or [])]
                max_col = max(max_col, len(rr))
                tmp_rows.append(rr)
            if not tmp_rows:
                tmp_rows = [['']]
            max_row = len(tmp_rows)
            # 统一列宽
            for rr in tmp_rows:
                while len(rr) < max_col:
                    rr.append('')
            rows = [
                [{'v': v, 'style': {}, 'css': ''} for v in rr]
                for rr in tmp_rows
            ]
        else:
            max_row = ws.max_row or 1
            max_col = ws.max_column or 1
            rows = []
            for r in range(1, max_row + 1):
                row = []
                for c in range(1, max_col + 1):
                    tr, tc = _top_left_for_cell(ws, r, c)
                    cell = ws.cell(tr, tc)
                    st = {}
                    css = ''
                    try:
                        st = serialize_cell_style(cell)
                        css = style_dict_to_css(st)
                    except Exception:
                        st = {}
                        css = ''
                    row.append({
                        'v': _cell_to_str(cell.value),
                        'style': st,
                        'css': css,
                    })
                rows.append(row)

        merge_ranges = []
        if not getattr(wb, "read_only", False):
            try:
                merge_ranges = [str(rng) for rng in ws.merged_cells.ranges]
            except Exception:
                merge_ranges = []

        images = []
        # 降级模式下不解析图片（图片解析最容易触发兼容性问题）
        if not getattr(wb, "read_only", False):
            for img in getattr(ws, '_images', []):
                try:
                    raw = img._data()
                except Exception:
                    continue
                b64 = base64.b64encode(raw).decode('ascii')
                fmt = (getattr(img, 'format', None) or 'png').lower()
                if fmt == 'jpeg':
                    mime = 'image/jpeg'
                elif fmt == 'gif':
                    mime = 'image/gif'
                else:
                    mime = 'image/png'
                row0, col0 = 0, 0
                anchor = getattr(img, 'anchor', None)
                if anchor is not None and hasattr(anchor, '_from'):
                    row0 = int(anchor._from.row)
                    col0 = int(anchor._from.col)
                elif anchor is not None and isinstance(anchor, str):
                    from openpyxl.utils.cell import coordinate_to_tuple
                    try:
                        row0, col0 = coordinate_to_tuple(anchor.upper())
                        row0 -= 1
                        col0 -= 1
                    except Exception:
                        pass
                width_px = None
                height_px = None
                if anchor is not None and hasattr(anchor, 'ext') and anchor.ext is not None:
                    try:
                        from openpyxl.utils.units import EMU_to_pixels

                        width_px = EMU_to_pixels(anchor.ext.width)
                        height_px = EMU_to_pixels(anchor.ext.height)
                    except Exception:
                        pass
                images.append({
                    'row': row0,
                    'col': col0,
                    'mime': mime,
                    'data': b64,
                    'width': width_px,
                    'height': height_px,
                })

        payload = {
            'sheet_name': ws.title,
            'rows': rows,
            'images': images,
            'merge_ranges': merge_ranges,
            'max_row': max_row,
            'max_col': max_col,
        }
        wb.close()
        return payload, None
    except Exception as e:
        try:
            wb.close()
        except Exception:
            pass
        return None, str(e)


def is_merged_subordinate(ws, row: int, col: int) -> bool:
    """(row,col) 为 1-based。若为合并区域内非左上角则返回 True。"""
    merged = getattr(ws, "merged_cells", None)
    if merged is None:
        return False
    for merged_range in merged.ranges:
        try:
            min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
        except Exception:
            continue
        if min_row <= row <= max_row and min_col <= col <= max_col:
            return (row, col) != (min_row, min_col)
    return False


def _extract_cell_val(raw):
    if isinstance(raw, dict):
        return raw.get('v', '')
    if raw is None:
        return ''
    return str(raw)


def _clear_rect(ws, r1: int, r2: int, c1: int, c2: int) -> None:
    """将矩形区域内每个合并块的左上角清空（用于 delete_rows/cols 失败时的降级）。"""
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            tl_r, tl_c = _top_left_for_cell(ws, r, c)
            if tl_r == r and tl_c == c:
                ws.cell(r, c).value = None


def save_sheet_values_preserving_format(path, rows: list):
    """
    在原文件上按网格写回文本值，并删去网格外的多余行列，尽量保留样式、图片。
    rows: 二维数组，元素为字符串或 {v: ...}
    """
    try:
        wb = load_workbook(path, read_only=False, data_only=False, keep_links=True)
    except Exception as e:
        return False, str(e)
    try:
        ws = wb.active
        old_max_row = ws.max_row or 1
        old_max_col = ws.max_column or 1

        if not rows:
            rows = [['']]

        norm = []
        for row in rows:
            if not isinstance(row, list):
                norm.append([row])
            else:
                norm.append(list(row))
        rows = norm

        new_max_row = len(rows)
        lengths = [len(r) for r in rows]
        new_max_col = max(lengths) if lengths else 1
        new_max_col = max(1, new_max_col)

        for r in rows:
            while len(r) < new_max_col:
                r.append('')
            del r[new_max_col:]

        for r_idx in range(1, new_max_row + 1):
            row = rows[r_idx - 1]
            for c_idx in range(1, new_max_col + 1):
                val = _extract_cell_val(row[c_idx - 1])
                tl_r, tl_c = _top_left_for_cell(ws, r_idx, c_idx)
                ws.cell(tl_r, tl_c).value = val

        if old_max_row > new_max_row:
            n = old_max_row - new_max_row
            try:
                ws.delete_rows(new_max_row + 1, n)
            except Exception:
                _clear_rect(ws, new_max_row + 1, old_max_row, 1, old_max_col)

        old_max_col = ws.max_column or new_max_col
        if old_max_col > new_max_col:
            n = old_max_col - new_max_col
            try:
                ws.delete_cols(new_max_col + 1, n)
            except Exception:
                mr = ws.max_row or new_max_row
                _clear_rect(ws, 1, mr, new_max_col + 1, old_max_col)

        wb.save(path)
        wb.close()
        return True, None
    except Exception as e:
        try:
            wb.close()
        except Exception:
            pass
        return False, str(e)
