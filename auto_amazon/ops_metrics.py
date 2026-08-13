"""从看板行运营难度 JSON 提取代表值（用于定时消息与预警）。"""
from __future__ import annotations

import json
import re
from pathlib import Path

from openpyxl import load_workbook

from .asin_access import normalize_asin
from .marketplace import (
    MARKETPLACE_US,
    normalize_marketplace,
    ops_ignore_label,
    ops_segment_count,
    resolve_review_interval_path,
)
from .media_paths import media_root
from .models import AsinDashboardRow


def _to_num(x) -> float | None:
    t = str(x).replace('%', '').replace('％', '').strip()
    if not t:
        return None
    try:
        return float(t)
    except (TypeError, ValueError):
        m = re.search(r'(\d+(?:\.\d+)?)', t)
        if m:
            return float(m.group(1))
        return None


def parse_ops_percent_values(
    raw: str,
    *,
    marketplace: str | None = None,
) -> list[float]:
    """从运营难度 JSON 中提取各区间百分比；忽略末档（US:200以上 / UK:150以上）。"""
    if not raw or not str(raw).strip():
        return []
    s = str(raw).strip()
    out: list[float] = []
    ignore = ops_ignore_label(marketplace)
    last_idx = ops_segment_count(marketplace) - 1

    if s.startswith('{'):
        try:
            obj = json.loads(s)
        except json.JSONDecodeError:
            obj = None
        if isinstance(obj, dict) and obj:
            kw = next(iter(obj.keys()))
            wrap = obj.get(kw)
            if isinstance(wrap, dict):
                ri = wrap.get('review_interval')
                if isinstance(ri, dict):
                    labels = list(ri.get('区间') or [])
                    ops = list(ri.get('运营难度') or [])
                    for i, x in enumerate(ops):
                        label = str(labels[i]).strip() if i < len(labels) else f'idx-{i}'
                        if ignore in label or i == last_idx:
                            continue
                        v = _to_num(x)
                        if v is not None:
                            out.append(v)
                    return out

    nums = re.findall(r'(\d+(?:\.\d+)?)\s*[%％]', s)
    for i, n in enumerate(nums):
        if i >= last_idx:
            break
        out.append(float(n))
    return out


def _ops_cell_json(keyword: str, review_inner: dict) -> str:
    payload = {keyword: {'review_interval': review_inner}}
    return json.dumps(payload, ensure_ascii=False)


def _read_review_interval_xlsx(path: Path, *, marketplace: str | None = None) -> dict | None:
    """从 review_interval_analysis*.xlsx 读取区间/运营难度数组。"""
    max_rows = ops_segment_count(marketplace)
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return None
    try:
        ws = wb.active
        labels: list[str] = []
        avg_sales: list = []
        ops: list[str] = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or row[0] is None:
                break
            label_s = str(row[0]).strip()
            if not label_s:
                break
            if '自然排名' in label_s:
                break
            if len(labels) >= max_rows:
                break
            labels.append(label_s)
            avg_sales.append(row[1] if len(row) > 1 else None)
            raw_ops = row[2] if len(row) > 2 else '0%'
            ops.append(str(raw_ops) if raw_ops is not None else '0%')
        if not ops:
            return None
        return {'区间': labels, '平均销量': avg_sales, '运营难度': ops}
    finally:
        wb.close()


def scan_ops_fields_from_media(
    asin: str,
    *,
    marketplace: str | None = None,
) -> tuple[str, str, str]:
    """扫描 media/file/<ASIN>/<关键词>/ 下对应站点的区间分析表，最多取前 3 个关键词。"""
    root = media_root() / normalize_asin(asin)
    if not root.is_dir():
        return '', '', ''

    mp = normalize_marketplace(marketplace) or MARKETPLACE_US
    entries: list[tuple[str, Path]] = []
    for kw_dir in sorted(root.iterdir(), key=lambda p: p.name.lower()):
        if not kw_dir.is_dir():
            continue
        path = resolve_review_interval_path(kw_dir, mp)
        if path is not None:
            entries.append((kw_dir.name, path))

    out: list[str] = []
    for keyword, path in entries[:3]:
        ri = _read_review_interval_xlsx(path, marketplace=mp)
        out.append(_ops_cell_json(keyword, ri) if ri else '')
    while len(out) < 3:
        out.append('')
    return out[0], out[1], out[2]


def refresh_row_ops_from_media(row: AsinDashboardRow) -> bool:
    """ROI 重算后从本地 Excel 同步运营难度 1/2/3 到看板行。"""
    mp = getattr(row, 'marketplace', None) or MARKETPLACE_US
    o1, o2, o3 = scan_ops_fields_from_media(row.asin, marketplace=mp)
    updates: dict[str, str] = {}
    if o1:
        updates['ops_difficulty_1'] = o1
    if o2:
        updates['ops_difficulty_2'] = o2
    if o3:
        updates['ops_difficulty_3'] = o3
    if not updates:
        return False
    AsinDashboardRow.objects.filter(pk=row.pk).update(**updates)
    return True


def ops_percent_from_media(asin: str, *, marketplace: str | None = None) -> float | None:
    """不依赖 DB，直接从 media 下 Excel 取运营难度代表值。"""
    mp = normalize_marketplace(marketplace) or MARKETPLACE_US
    max_val: float | None = None
    for field_val in scan_ops_fields_from_media(asin, marketplace=mp):
        for v in parse_ops_percent_values(field_val, marketplace=mp):
            if max_val is None or v > max_val:
                max_val = v
    return max_val


def row_ops_difficulty_percent(row: AsinDashboardRow, *, media_fallback: bool = False) -> float | None:
    """取三个运营难度字段中区间百分比的最大值作为代表运营难度。"""
    mp = getattr(row, 'marketplace', None) or MARKETPLACE_US
    max_val: float | None = None
    for field in ('ops_difficulty_1', 'ops_difficulty_2', 'ops_difficulty_3'):
        for v in parse_ops_percent_values(getattr(row, field, '') or '', marketplace=mp):
            if max_val is None or v > max_val:
                max_val = v
    if max_val is None and media_fallback:
        return ops_percent_from_media(normalize_asin(row.asin), marketplace=mp)
    return max_val


def row_metrics_snapshot(row: AsinDashboardRow, *, media_fallback: bool = False) -> dict[str, float | None]:
    return {
        'ad_roi': row.ad_removed_roi,
        'ad_difficulty': row.ranking_percent,
        'ops_difficulty': row_ops_difficulty_percent(row, media_fallback=media_fallback),
    }


# 与 scheduled_jobs.compute_alert_status 中「立即淘汰」阈值一致
ELIMINATE_AD_ROI_MAX = 80.0
ELIMINATE_OPS_MAX = 10.0


def should_skip_scheduled_ad_difficulty(
    row: AsinDashboardRow,
    *,
    media_fallback: bool = False,
) -> bool:
    """
    当「最新去广告投产比」或「最新运营难度」任一已达立即淘汰线时，跳过广告难度重算。
    """
    snap = row_metrics_snapshot(row, media_fallback=media_fallback)
    ad_roi = snap.get('ad_roi')
    ops = snap.get('ops_difficulty')
    roi_eliminate = False
    ops_eliminate = False
    if ad_roi is not None:
        try:
            roi_eliminate = float(ad_roi) <= ELIMINATE_AD_ROI_MAX
        except (TypeError, ValueError):
            pass
    if ops is not None:
        try:
            ops_eliminate = float(ops) <= ELIMINATE_OPS_MAX
        except (TypeError, ValueError):
            pass
    return roi_eliminate or ops_eliminate
