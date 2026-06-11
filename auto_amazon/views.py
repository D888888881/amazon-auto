import json
from datetime import datetime
import shutil
import uuid
import zipfile
from io import BytesIO
from pathlib import Path
import re
from collections import defaultdict

from django.conf import settings
from django.contrib import messages
from django.contrib.auth import login as auth_login
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.models import User
from django.core.cache import cache
from django.core.paginator import Paginator
from django.db import close_old_connections
from django.db import DatabaseError
from django.db import transaction
from django.db.models import Q
from django.http import FileResponse, Http404, HttpResponse, JsonResponse
from django.utils import timezone
from django.shortcuts import redirect, render
from django.urls import reverse
from django.views.decorators.http import require_GET, require_POST
from openpyxl import Workbook, load_workbook

from .asin_job_lock import AsinComputeLock
from .resilient_wizard import (
    AsinFailure,
    BatchRunResult,
    batch_result_from_wizard_output,
    ordered_unique_asins,
    run_ad_difficulty_asins_batch,
    run_ad_difficulty_asins_sequential,
    run_roi_asins_sequential,
    use_sequential_roi,
)
from .asin_wizard import run_seller_wizard
from .excel_io import (
    read_active_sheet_rich,
    save_sheet_values_preserving_format,
    search_xlsx_save_is_rewrite,
)
from .excel_search_restore import build_origin_row_from_search, search_row_to_api_record
from .forms import RegisterForm
from .media_paths import media_root, parent_rel, safe_media_path_global
from .dashboard_ops_filter import run_dashboard_ops_filter
from .asin_access import (
    dedupe_dashboard_rows,
    normalize_asin,
    resolve_dashboard_row_for_persist,
    user_assigned_asin_codes,
    user_can_access_excel_media_path,
    user_can_delete_dashboard_row,
    user_can_operate_dashboard_row,
    user_dashboard_rows_qs,
    user_imported_asin_codes,
)
from .excel_import_utils import ensure_dir_nodes_registered, extract_zip_to_media_root
from .media_import_staging import (
    append_chunk as import_append_chunk,
    cleanup_staging_dir,
    load_ready_staging,
    refresh_conflicts_in_meta,
)
from .models import (
    AsinCatalogItem,
    AsinDashboardRow,
    AsinDataUpdateStamp,
    AsinFolderAssignment,
    AsinRoiPackVerification,
    AsinUploadBatch,
    ImportedMediaPath,
    ScheduledTaskMessage,
    UserProfile,
)
from .asin_upload import batch_asin_lines, ingest_asin_upload
from .roi_us_pack_recalc import (
    extract_dashboard_metrics_from_roi_us_pack,
    is_roi_us_pack_filename,
    recalc_roi_us_pack_rows,
)
from .utils import product_grade_from_monthly_profit1
from .exchange_rate import fetch_usd_cny_rate
from .wizard_jobs import (
    WIZARD_JOB_TTL,
    clear_user_active_job as _clear_user_active_job,
    dismiss_active_job,
    get_active_job_for_user,
    job_status_payload,
    new_job_entry,
    refresh_job_entry,
    set_user_active_job as _set_user_active_job,
    wizard_job_key as _wizard_job_key,
)


def _default_exchange_rate_for_form() -> str:
    """页面汇率输入框默认值（实时获取，用户仍可修改）。"""
    try:
        rate = fetch_usd_cny_rate()
    except Exception:
        rate = float(getattr(settings, 'USD_CNY_RATE_FALLBACK', 7.2))
    text = f'{rate:.4f}'.rstrip('0').rstrip('.')
    return text or '7.2'


def _merge_cost_overrides_from_db(
    asins: list[str] | None, cost_overrides: dict | None
) -> dict:
    """图搜失败时可用看板已填写的采购价/头程作为 unit_purchase_override。"""
    merged: dict = {}
    asin_set = {normalize_asin(a) for a in (asins or []) if normalize_asin(a)}
    if asin_set:
        for r in AsinDashboardRow.objects.filter(asin__in=asin_set).only(
            'asin', 'unit_purchase', 'head_distance'
        ):
            a = normalize_asin(r.asin)
            if not a:
                continue
            one = merged.setdefault(a, {})
            if r.unit_purchase is not None:
                one.setdefault('unit_purchase', float(r.unit_purchase))
            if r.head_distance is not None:
                one.setdefault('head_distance', float(r.head_distance))
    for a, one in (cost_overrides or {}).items():
        key = normalize_asin(a)
        if not key or not isinstance(one, dict):
            continue
        merged.setdefault(key, {}).update(one)
    return merged


def _dashboard_per_page(request) -> int:
    raw = (request.GET.get('per_page') or '').strip().lower()
    custom = _get_int(request.GET.get('per_page_custom'))
    if raw == 'custom' and custom is not None:
        return max(1, min(custom, 500))
    try:
        n = int(raw)
    except (TypeError, ValueError):
        n = 10
    if n in (10, 20, 40, 100):
        return n
    return 10


def _compute_roi_per_page(request) -> int:
    raw = (request.GET.get('per_page') or '').strip().lower()
    custom = _get_int(request.GET.get('per_page_custom'))
    if raw == 'custom' and custom is not None:
        return max(1, min(custom, 500))
    try:
        n = int(raw)
    except (TypeError, ValueError):
        n = 20
    if n in (20, 50, 100):
        return n
    return 20


def _assignment_label_map(asins: set[str]) -> dict[str, str]:
    if not asins:
        return {}
    rows = (
        AsinFolderAssignment.objects.filter(asin__in=asins)
        .prefetch_related('assignees')
        .only('id', 'asin')
    )
    out: dict[str, str] = {}
    for a in rows:
        names = sorted({u.username for u in a.assignees.all()})
        out[normalize_asin(a.asin)] = '、'.join(names) if names else ''
    return out


def _touch_asin_updates(asins: list[str] | set[str]) -> None:
    uniq = sorted({normalize_asin(a) for a in asins if str(a or '').strip()})
    for a in uniq:
        AsinDataUpdateStamp.objects.update_or_create(asin=a, defaults={})


def _parse_dt_local(raw: str | None, end_of_day: bool = False):
    s = (raw or '').strip()
    if not s:
        return None
    for fmt in ('%Y-%m-%dT%H:%M', '%Y-%m-%d'):
        try:
            dt = datetime.strptime(s, fmt)
            if fmt == '%Y-%m-%d' and end_of_day:
                dt = dt.replace(hour=23, minute=59, second=59)
            if timezone.is_naive(dt):
                dt = timezone.make_aware(dt, timezone.get_current_timezone())
            return dt
        except ValueError:
            continue
    return None


def _attach_row_assignee_labels(
    rows,
    label_map: dict[str, str],
    viewer_id: int,
    viewer_assigned_asins: set[str],
    viewer_is_superuser: bool,
    roi_verified_asins: set[str] | None = None,
) -> None:
    verified = roi_verified_asins or set()
    for r in rows:
        key = normalize_asin(r.asin)
        r.assignee_display = label_map.get(key) or '—'
        can_operate = viewer_is_superuser or r.user_id == viewer_id or key in viewer_assigned_asins
        r.ops_readonly = not can_operate
        r.can_delete_row = viewer_is_superuser or r.user_id == viewer_id
        r.can_open_dir = viewer_is_superuser or (key in viewer_assigned_asins)
        r.roi_pack_verified = key in verified


def _require_superuser_for_excel_audit(request) -> bool:
    return bool(getattr(request.user, 'is_superuser', False))


def _first_word_from_filename(name: str) -> str:
    stem = Path(name).stem.strip()
    if not stem:
        return 'Excel'
    return stem.split()[0]


def _is_search_sheet_filename(name: str) -> bool:
    n = (name or '').lower()
    return n.startswith('search(') and n.endswith('.xlsx')


def _deleted_asin_set_from_data_origin(search_path: Path) -> set[str]:
    """
    对比同目录 *_data_origin.xlsx：
    返回在 Search 中存在、但在 data_origin 中不存在的 asin 集合（视为被清洗删除）。
    """
    try:
        parent = search_path.parent
        origins = sorted(parent.glob('*_data_origin.xlsx'))
        if not origins:
            return set()
        origin_path = origins[0]

        def _asin_set(xlsx: Path) -> set[str]:
            wb = load_workbook(xlsx, read_only=True, data_only=True)
            ws = wb.active
            rows = ws.iter_rows(values_only=True)
            header = next(rows, None)
            if not header:
                wb.close()
                return set()
            idx = -1
            for i, h in enumerate(header):
                if str(h or '').strip().lower() == 'asin':
                    idx = i
                    break
            if idx < 0:
                wb.close()
                return set()
            out: set[str] = set()
            for r in rows:
                if not r or idx >= len(r):
                    continue
                v = str(r[idx] or '').strip().upper()
                if v:
                    out.add(v)
            wb.close()
            return out

        s_set = _asin_set(search_path)
        o_set = _asin_set(origin_path)
        if not s_set:
            return set()
        return {x for x in s_set if x not in o_set}
    except Exception:
        return set()


def _highlight_deleted_rows_for_search(payload: dict, deleted_asins: set[str]) -> None:
    """将 Search 表中命中 deleted_asins 的整行高亮。"""
    if not deleted_asins:
        return
    rows = payload.get('rows')
    if not isinstance(rows, list) or len(rows) < 2:
        return
    head = rows[0] if isinstance(rows[0], list) else []
    asin_col = -1
    for i, c in enumerate(head):
        txt = ''
        if isinstance(c, dict):
            txt = str(c.get('v', '') or '').strip().lower()
        else:
            txt = str(c or '').strip().lower()
        if txt == 'asin':
            asin_col = i
            break
    if asin_col < 0:
        return
    for r in range(1, len(rows)):
        row = rows[r]
        if not isinstance(row, list) or asin_col >= len(row):
            continue
        c = row[asin_col]
        asin = (str(c.get('v', '') if isinstance(c, dict) else c or '')).strip().upper()
        if not asin or asin not in deleted_asins:
            continue
        for j, one in enumerate(row):
            if isinstance(one, dict):
                base_css = str(one.get('css', '') or '')
                if 'background-color' not in base_css:
                    one['css'] = (base_css + ';background-color:#fff3cd').strip(';')
            else:
                row[j] = {'v': str(one or ''), 'style': {}, 'css': 'background-color:#fff3cd'}


def _rows_to_plain(rows: list) -> list[list[str]]:
    out: list[list[str]] = []
    for r in rows or []:
        if not isinstance(r, list):
            r = [r]
        one: list[str] = []
        for c in r:
            if isinstance(c, dict):
                one.append(str(c.get('v', '') or ''))
            else:
                one.append('' if c is None else str(c))
        out.append(one)
    return out


def _header_index_map(header: list[str]) -> dict[str, int]:
    m: dict[str, int] = {}
    for i, h in enumerate(header):
        key = str(h or '').strip().lower()
        if key and key not in m:
            m[key] = i
    return m


def _map_row_to_target_header(src_header: list[str], src_row: list[str], tgt_header: list[str]) -> list[str]:
    src_map = _header_index_map(src_header)
    out: list[str] = []
    for th in tgt_header:
        key = str(th or '').strip().lower()
        idx = src_map.get(key)
        if idx is None or idx >= len(src_row):
            out.append('')
        else:
            out.append(src_row[idx])
    return out


def _find_data_origin_path(keyword_dir: Path) -> Path | None:
    cands = sorted(keyword_dir.glob('*_data_origin.xlsx'))
    return cands[0] if cands else None


def _asin_set_from_plain_rows(rows: list[list[str]]) -> set[str]:
    if not rows:
        return set()
    header = rows[0]
    idx = _header_index_map(header).get('asin', -1)
    if idx < 0:
        return set()
    out: set[str] = set()
    for r in rows[1:]:
        if idx < len(r):
            v = str(r[idx] or '').strip().upper()
            if v:
                out.add(v)
    return out


def register(request):
    if request.user.is_authenticated:
        return redirect('index')
    if request.method == 'POST':
        form = RegisterForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(
                request,
                '注册申请已提交，请等待超级管理员审核。审核通过后即可使用账号登录。',
            )
            return redirect('login')
    else:
        form = RegisterForm()
    return render(request, 'registration/register.html', {'form': form})


def login_view(request):
    if request.user.is_authenticated:
        return redirect(settings.LOGIN_REDIRECT_URL)
    next_url = request.GET.get('next') or ''
    if request.method == 'POST':
        username = (request.POST.get('username') or '').strip()
        password = request.POST.get('password') or ''
        next_url = (request.POST.get('next') or '').strip() or next_url
        try:
            user = User.objects.get(username=username)
        except User.DoesNotExist:
            messages.error(request, '用户名或密码不正确。')
            return render(request, 'registration/login.html', {'next': next_url})
        if not user.check_password(password):
            messages.error(request, '用户名或密码不正确。')
            return render(request, 'registration/login.html', {'next': next_url})
        if not user.is_active:
            try:
                st = user.profile.registration_status
            except UserProfile.DoesNotExist:
                st = UserProfile.Status.PENDING
            if st == UserProfile.Status.REJECTED:
                messages.error(request, '您的注册申请已被拒绝，无法登录。')
            else:
                messages.warning(request, '您的账号尚未通过管理员审核，请耐心等待。')
            return render(request, 'registration/login.html', {'next': next_url})
        auth_login(request, user, backend='django.contrib.auth.backends.ModelBackend')
        return redirect(next_url or settings.LOGIN_REDIRECT_URL)
    return render(request, 'registration/login.html', {'next': next_url})


@login_required
@user_passes_test(lambda u: u.is_superuser)
def pending_registrations(request):
    if request.method == 'POST':
        user_id = request.POST.get('user_id')
        action = request.POST.get('action')
        if not user_id or action not in ('approve', 'reject'):
            messages.error(request, '请求无效。')
            return redirect('pending_registrations')
        try:
            target = User.objects.get(pk=int(user_id))
        except (User.DoesNotExist, ValueError):
            messages.error(request, '用户不存在。')
            return redirect('pending_registrations')
        if target.is_superuser:
            messages.error(request, '不能审核超级管理员账号。')
            return redirect('pending_registrations')
        try:
            profile = target.profile
        except UserProfile.DoesNotExist:
            profile = UserProfile.objects.create(
                user=target,
                registration_status=UserProfile.Status.PENDING,
            )
        if profile.registration_status != UserProfile.Status.PENDING:
            messages.warning(request, f'用户「{target.username}」当前不是待审核状态。')
            return redirect('pending_registrations')
        now = timezone.now()
        if action == 'approve':
            with transaction.atomic():
                target.is_active = True
                target.save(update_fields=['is_active'])
                profile.registration_status = UserProfile.Status.APPROVED
                profile.reviewed_at = now
                profile.reviewed_by = request.user
                profile.save(
                    update_fields=['registration_status', 'reviewed_at', 'reviewed_by'],
                )
            messages.success(request, f'已通过用户「{target.username}」的注册申请。')
        else:
            with transaction.atomic():
                target.is_active = False
                target.save(update_fields=['is_active'])
                profile.registration_status = UserProfile.Status.REJECTED
                profile.reviewed_at = now
                profile.reviewed_by = request.user
                profile.save(
                    update_fields=['registration_status', 'reviewed_at', 'reviewed_by'],
                )
            messages.success(request, f'已拒绝用户「{target.username}」的注册申请。')
        return redirect('pending_registrations')

    pending = (
        UserProfile.objects.filter(registration_status=UserProfile.Status.PENDING)
        .select_related('user')
        .order_by('user__date_joined')
    )
    hist_base = (
        UserProfile.objects.filter(
            registration_status__in=(
                UserProfile.Status.APPROVED,
                UserProfile.Status.REJECTED,
            ),
            reviewed_at__isnull=False,
        )
        .select_related('user', 'reviewed_by')
        .order_by('-reviewed_at')
    )
    review_history_total = hist_base.count()
    review_history = hist_base[:100]
    return render(
        request,
        'auto_amazon/pending_registrations.html',
        {
            'pending': pending,
            'pending_count': pending.count(),
            'review_history': review_history,
            'review_history_total': review_history_total,
        },
    )


@login_required
@user_passes_test(lambda u: u.is_superuser)
def user_management(request):
    if request.method == 'POST':
        user_id = request.POST.get('user_id')
        action = request.POST.get('action')

        if not user_id or action not in ('disable', 'enable', 'delete'):
            messages.error(request, '请求无效。')
            return redirect('user_management')

        try:
            target = User.objects.get(pk=int(user_id))
        except (User.DoesNotExist, ValueError):
            messages.error(request, '用户不存在。')
            return redirect('user_management')

        # 安全保护：不能操作其它超级管理员、不能对自己操作。
        if target.is_superuser:
            messages.error(request, '不能操作超级管理员账号。')
            return redirect('user_management')
        if target == request.user:
            messages.error(request, '不能对自己执行该操作。')
            return redirect('user_management')

        if action == 'disable':
            if target.is_active:
                target.is_active = False
                target.save(update_fields=['is_active'])
                messages.success(request, f'已禁用用户「{target.username}」。')
            else:
                messages.warning(request, f'用户「{target.username}」当前已处于禁用状态。')
        elif action == 'enable':
            if not target.is_active:
                target.is_active = True
                target.save(update_fields=['is_active'])
                messages.success(request, f'已启用用户「{target.username}」。')
            else:
                messages.warning(request, f'用户「{target.username}」当前已处于启用状态。')
        else:  # delete
            target.delete()
            messages.success(request, '用户已删除。')

        return redirect('user_management')

    # GET 请求部分保持不变
    users_qs = User.objects.filter(is_superuser=False).order_by('-date_joined')
    users_count = users_qs.count()
    active_count = users_qs.filter(is_active=True).count()
    inactive_count = users_qs.filter(is_active=False).count()

    return render(
        request,
        'auto_amazon/user_manage.html',
        {
            'users': users_qs,
            'users_count': users_count,
            'active_count': active_count,
            'inactive_count': inactive_count,
        },
    )


DASHBOARD_SORT_FIELDS = {
    'asin': 'asin',
    'profit_margin': 'profit_margin',
    'ranking_percent': 'ranking_percent',
    'unit_purchase': 'unit_purchase',
    'monthly_results': 'monthly_results',
    'profit_per_order': 'profit_per_order',
    'monthly_sales_total': 'monthly_sales_total',
    'ad_removed_roi': 'ad_removed_roi',
    'head_actual_total': 'head_actual_total',
    'monthly_profit1': 'monthly_profit1',
    'product_grade': 'monthly_profit1',
    'created_at': 'created_at',
    'updated_at': 'updated_at',
    'follow_status': 'follow_status',
}


def _get_float(val: str | None):
    if val is None:
        return None
    s = str(val).strip()
    if s == '':
        return None
    try:
        return float(s)
    except (TypeError, ValueError):
        return None


def _get_int(val: str | None):
    if val is None:
        return None
    s = str(val).strip()
    if s == '':
        return None
    try:
        return int(s)
    except (TypeError, ValueError):
        return None


OPS_REVIEW_INTERVAL_OPTIONS = ('0-30', '31-50', '51-100', '101-200', '200以上')


def _is_200_plus_review_label(label: str) -> bool:
    t = str(label).strip()
    return '200以上' in t or t.replace(' ', '') in ('200+', '200＋')


def _review_labels_equivalent(selected: str, actual: str) -> bool:
    """看板筛选用的评价数量区间与 JSON 内「区间」标签是否等同。"""
    sel = str(selected).strip()
    act = str(actual).strip()
    if not sel or not act:
        return False
    if sel == act:
        return True
    if sel == '101-200' and act in ('101-150', '151-200'):
        return True
    if sel == '200以上' and _is_200_plus_review_label(act):
        return True
    return False


def _parse_ops_json(raw: str) -> list[tuple[str, float]]:
    """从运营难度 JSON 中提取(评价数量区间, 运营难度百分比)。"""
    if not raw or not str(raw).strip():
        return []
    s = str(raw).strip()
    out: list[tuple[str, float]] = []

    def _to_num(x):
        t = str(x).replace('%', '').replace('％', '').strip()
        if not t:
            return None
        try:
            return float(t)
        except (TypeError, ValueError):
            m = re.search(r'(\d+(?:\.\d+)?)', t)
            if m:
                return float(m.group(1))
            return None

    # 新格式：严格 JSON
    if s.startswith('{'):
        try:
            obj = json.loads(s)
        except json.JSONDecodeError:
            obj = None
        if isinstance(obj, dict) and obj:
            kw = next(iter(obj.keys()))
            wrap = obj.get(kw)
            if isinstance(wrap, dict):
                ri = wrap.get('review_interval')
                if isinstance(ri, dict):
                    labels = list(ri.get('区间') or [])
                    ops = list(ri.get('运营难度') or [])
                    if ops:
                        for i, x in enumerate(ops):
                            label = str(labels[i]).strip() if i < len(labels) else f'idx-{i}'
                            v = _to_num(x)
                            if v is not None:
                                out.append((label, v))
                        if out:
                            return out

    # 兼容旧格式文本：兜底提取百分比
    fallback_labels = ['0-30', '31-50', '51-100', '101-200', '200以上']
    nums = re.findall(r'(\d+(?:\.\d+)?)\s*[%％]', s)
    for i, n in enumerate(nums):
        label = fallback_labels[i] if i < len(fallback_labels) else f'idx-{i}'
        out.append((label, float(n)))
    return out


def _ops_match(
    raw: str,
    lo: float | None,
    hi: float | None,
    *,
    review_interval: str = 'all',
) -> bool:
    pairs = _parse_ops_json(raw)
    if not pairs:
        return False
    for label, v in pairs:
        if review_interval == 'all':
            if _is_200_plus_review_label(label):
                continue
        elif not _review_labels_equivalent(review_interval, label):
            continue
        if lo is not None and v < lo:
            continue
        if hi is not None and v > hi:
            continue
        return True
    return False


def _delete_asin_folder_if_allowed(user, asin: str) -> tuple[bool, str]:
    """
    删除 media/file/<ASIN> 目录。
    返回 (是否成功删除, 说明信息)。
    普通用户仅可删除本人导入过的 ASIN 目录；超管可删除任意 ASIN 目录。
    """
    asin = normalize_asin(asin)
    if not asin:
        return False, 'ASIN 无效'
    if not re.match(r'^B0[A-Z0-9]{8}$', asin):
        return False, f'ASIN 非法：{asin}'
    if not getattr(user, 'is_superuser', False):
        if not ImportedMediaPath.objects.filter(user=user, rel_path=asin).exists():
            return False, f'无权限删除 ASIN 文件夹：{asin}（仅可删除本人导入目录）'
    d = media_root() / asin
    if not d.exists():
        ImportedMediaPath.objects.filter(Q(rel_path=asin) | Q(rel_path__startswith=f'{asin}/')).delete()
        return True, f'{asin} 目录不存在，已清理导入记录'
    try:
        shutil.rmtree(d)
    except OSError as e:
        return False, f'删除 {asin} 目录失败：{e}'
    ImportedMediaPath.objects.filter(Q(rel_path=asin) | Q(rel_path__startswith=f'{asin}/')).delete()
    return True, f'已删除 ASIN 文件夹：{asin}'


@login_required
def index(request):
    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'delete_batch':
            ids = request.POST.getlist('row_ids')
            if ids:
                delete_with_folder = str(request.POST.get('delete_with_folder') or '').lower() in (
                    '1',
                    'true',
                    'on',
                    'yes',
                )
                rows = list(AsinDashboardRow.objects.filter(user=request.user, pk__in=ids).only('asin'))
                asins = sorted({normalize_asin(r.asin) for r in rows if r.asin})
                AsinDashboardRow.objects.filter(user=request.user, pk__in=ids).delete()
                messages.success(request, f'已删除 {len(ids)} 条记录。')
                if delete_with_folder and asins:
                    ok_cnt = 0
                    err_msgs: list[str] = []
                    for asin in asins:
                        ok, msg = _delete_asin_folder_if_allowed(request.user, asin)
                        if ok:
                            ok_cnt += 1
                        else:
                            err_msgs.append(msg)
                    if ok_cnt:
                        messages.success(request, f'已同步删除 {ok_cnt} 个 ASIN 文件夹。')
                    for m in err_msgs[:5]:
                        messages.warning(request, m)
        elif action == 'calc_roi_selected':
            ids = request.POST.getlist('row_ids')
            if not ids:
                messages.error(request, '请先勾选需要计算 ROI 的 ASIN。')
                return redirect('index')
            active = get_active_job_for_user(request.user.id)
            if active:
                messages.warning(
                    request,
                    f'已有 ROI 任务进行中（{active["status_label"]}）。已跳转到进度页。',
                )
                return redirect(f"{reverse('compute_roi')}?job_id={active['job_id']}")
            parity, err = _parse_compute_local_form(request)
            if err:
                messages.error(request, err)
                return redirect('index')
            accessible = user_dashboard_rows_qs(request.user)
            rows_sel = list(
                accessible.filter(pk__in=ids).only(
                    'pk', 'asin', 'unit_purchase', 'head_distance'
                )
            )
            rows_sel = [r for r in rows_sel if user_can_operate_dashboard_row(request.user, r)]
            asins = sorted({r.asin for r in rows_sel if r.asin})
            if not asins:
                messages.error(request, '未找到可计算的 ASIN。')
                return redirect('index')
            target_row_ids = {r.asin: r.pk for r in rows_sel}
            cost_overrides: dict = {}
            for r in rows_sel:
                up_in = _get_float(request.POST.get(f'unit_purchase_{r.pk}'))
                hd_in = _get_float(request.POST.get(f'head_distance_{r.pk}'))
                up_val = up_in if up_in is not None else r.unit_purchase
                hd_val = hd_in if hd_in is not None else r.head_distance
                updates = {}
                if up_in is not None:
                    updates['unit_purchase'] = up_in
                if hd_in is not None:
                    updates['head_distance'] = hd_in
                if up_val is not None and hd_val is not None:
                    updates['head_actual_total'] = round(up_val + hd_val, 2)
                if updates:
                    AsinDashboardRow.objects.filter(pk=r.pk).update(**updates)
                one = {}
                if up_val is not None:
                    one['unit_purchase'] = float(up_val)
                if hd_val is not None:
                    one['head_distance'] = float(hd_val)
                if one:
                    cost_overrides[r.asin] = one
            jid = str(uuid.uuid4())
            key = _wizard_job_key(jid)
            _set_user_active_job(request.user.id, jid)
            cache.set(
                key,
                new_job_entry(
                    request.user.id,
                    [
                        '看板 ROI 计算任务已创建…',
                        f'当前汇率：{parity}。',
                        f'本次勾选 ASIN：{len(asins)} 个。',
                    ],
                    task_type='wizard',
                    parity=parity,
                    asins=asins,
                    cost_overrides=cost_overrides,
                    target_row_ids=target_row_ids,
                ),
                WIZARD_JOB_TTL,
            )
            from .rq_enqueue import dispatch_wizard_job

            dispatch_wizard_job(
                jid,
                request.user.id,
                asins,
                parity,
                cost_overrides=cost_overrides,
                target_row_ids=target_row_ids,
            )
            messages.success(request, f'已开始计算 {len(asins)} 个 ASIN 的 ROI，已跳转到进度页。')
            return redirect(f"{reverse('compute_roi')}?job_id={jid}")
        elif action == 'calc_ranking_selected':
            ids = request.POST.getlist('row_ids')
            if not ids:
                messages.error(request, '请先勾选需要计算广告难度的 ASIN。')
                return redirect('index')
            active = get_active_job_for_user(request.user.id)
            if active:
                messages.warning(
                    request,
                    f'已有任务进行中（{active["status_label"]}）。已跳转到进度页。',
                )
                return redirect(f"{reverse('compute_roi')}?job_id={active['job_id']}")
            accessible = user_dashboard_rows_qs(request.user)
            rows_sel = list(accessible.filter(pk__in=ids).only('pk', 'asin'))
            rows_sel = [r for r in rows_sel if user_can_operate_dashboard_row(request.user, r)]
            asins = sorted({r.asin for r in rows_sel if r.asin})
            if not asins:
                messages.error(request, '未找到可计算的记录。')
                return redirect('index')
            target_row_pks = [r.pk for r in rows_sel]
            jid = str(uuid.uuid4())
            key = _wizard_job_key(jid)
            _set_user_active_job(request.user.id, jid)
            cache.set(
                key,
                new_job_entry(
                    request.user.id,
                    [
                        '广告难度计算任务已创建…',
                        f'本次勾选 ASIN：{len(asins)} 个。',
                    ],
                    task_type='ad_difficulty',
                    asins=asins,
                    target_row_pks=target_row_pks,
                ),
                WIZARD_JOB_TTL,
            )
            from .rq_enqueue import dispatch_ad_difficulty_job

            dispatch_ad_difficulty_job(
                jid, request.user.id, asins, target_row_pks=target_row_pks
            )
            messages.success(request, f'已开始计算 {len(asins)} 个 ASIN 的广告难度，已跳转到进度页。')
            return redirect(f"{reverse('compute_roi')}?job_id={jid}")
        elif action == 'delete_one':
            rid = request.POST.get('row_id')
            if rid:
                delete_with_folder = str(request.POST.get('delete_with_folder') or '').lower() in (
                    '1',
                    'true',
                    'on',
                    'yes',
                )
                row = AsinDashboardRow.objects.filter(user=request.user, pk=rid).only('asin').first()
                asin = normalize_asin(row.asin) if row and row.asin else ''
                AsinDashboardRow.objects.filter(user=request.user, pk=rid).delete()
                messages.success(request, '已删除该条记录。')
                if delete_with_folder and asin:
                    ok, msg = _delete_asin_folder_if_allowed(request.user, asin)
                    if ok:
                        messages.success(request, msg)
                    else:
                        messages.warning(request, msg)
        elif action == 'set_follow_status':
            rid = (request.POST.get('row_id') or '').strip()
            status = (request.POST.get('follow_status') or '').strip()
            if rid.isdigit() and status in (
                AsinDashboardRow.FollowStatus.NORMAL,
                AsinDashboardRow.FollowStatus.PRIORITY,
            ):
                row = user_dashboard_rows_qs(request.user).filter(pk=int(rid)).first()
                if row and user_can_operate_dashboard_row(request.user, row):
                    AsinDashboardRow.objects.filter(pk=row.pk).update(follow_status=status)
            nxt = (request.POST.get('next') or '').strip()
            if nxt.startswith('/') and not nxt.startswith('//'):
                return redirect(nxt)
            return redirect('index')
        return redirect('index')

    sort_key = request.GET.get('sort', 'updated_at')
    direction = request.GET.get('dir', 'desc')
    if direction not in ('asc', 'desc'):
        direction = 'desc'
    field = DASHBOARD_SORT_FIELDS.get(sort_key, 'updated_at')
    order_prefix = '-' if direction == 'desc' else ''
    order_by = f'{order_prefix}{field}'

    assigned_codes = user_assigned_asin_codes(request.user)
    assigned_set = {normalize_asin(a) for a in assigned_codes}
    viewer_is_superuser = bool(request.user.is_superuser)
    rows_qs = user_dashboard_rows_qs(request.user)
    updated_stamp_map = {
        normalize_asin(a): t
        for a, t in AsinDataUpdateStamp.objects.values_list('asin', 'updated_at')
    }
    asin_kw = (request.GET.get('asin_kw') or '').strip()
    if asin_kw:
        rows_qs = rows_qs.filter(asin__icontains=asin_kw)

    assignee_username = (request.GET.get('assignee_user') or '').strip()
    if assignee_username:
        u = User.objects.filter(username__iexact=assignee_username, is_active=True).first()
        if u:
            asins_for_u = AsinFolderAssignment.objects.filter(assignees=u).values_list('asin', flat=True)
            rows_qs = rows_qs.filter(asin__in=list(asins_for_u))

    # 更新时间筛选
    updated_from = _parse_dt_local(request.GET.get('updated_from'))
    updated_to = _parse_dt_local(request.GET.get('updated_to'), end_of_day=True)
    if updated_from or updated_to:
        filtered_asins = set()
        for a, dt in updated_stamp_map.items():
            if dt is None:
                continue
            if updated_from and dt < updated_from:
                continue
            if updated_to and dt > updated_to:
                continue
            filtered_asins.add(a)
        rows_qs = rows_qs.filter(asin__in=list(filtered_asins))

    # 数值区间筛选
    numeric_fields = [
        ('profit_margin_min', 'profit_margin__gte'),
        ('profit_margin_max', 'profit_margin__lte'),
        ('ad_removed_roi_min', 'ad_removed_roi__gte'),
        ('ad_removed_roi_max', 'ad_removed_roi__lte'),
        ('monthly_results_min', 'monthly_results__gte'),
        ('monthly_results_max', 'monthly_results__lte'),
        ('profit_per_order_min', 'profit_per_order__gte'),
        ('profit_per_order_max', 'profit_per_order__lte'),
        ('monthly_sales_total_min', 'monthly_sales_total__gte'),
        ('monthly_sales_total_max', 'monthly_sales_total__lte'),
        ('unit_purchase_min', 'unit_purchase__gte'),
        ('unit_purchase_max', 'unit_purchase__lte'),
        ('head_actual_total_min', 'head_actual_total__gte'),
        ('head_actual_total_max', 'head_actual_total__lte'),
        ('ranking_percent_min', 'ranking_percent__gte'),
        ('ranking_percent_max', 'ranking_percent__lte'),
    ]
    for key, lookup in numeric_fields:
        v = _get_float(request.GET.get(key))
        if v is not None:
            rows_qs = rows_qs.filter(**{lookup: v})

    # 产品等级筛选（A-E，多选）
    grades = [g.upper() for g in request.GET.getlist('product_grade') if str(g).strip()]
    allowed = [g for g in grades if g in ('A', 'B', 'C', 'D', 'E')]
    if allowed:
        rows_qs = rows_qs.filter(product_grade__in=allowed)

    roi_verified_f = (request.GET.get('roi_verified') or '').strip().lower()
    if roi_verified_f in ('yes', 'no'):
        verified_codes = list(AsinRoiPackVerification.objects.values_list('asin', flat=True))
        if roi_verified_f == 'yes':
            rows_qs = rows_qs.filter(asin__in=verified_codes)
        else:
            rows_qs = rows_qs.exclude(asin__in=verified_codes)

    follow_f = (request.GET.get('follow_filter') or '').strip()
    if follow_f == AsinDashboardRow.FollowStatus.NORMAL:
        rows_qs = rows_qs.filter(follow_status=AsinDashboardRow.FollowStatus.NORMAL)
    elif follow_f == AsinDashboardRow.FollowStatus.PRIORITY:
        rows_qs = rows_qs.filter(follow_status=AsinDashboardRow.FollowStatus.PRIORITY)

    if field == 'updated_at':
        rows = list(rows_qs)
        rows.sort(
            key=lambda r: updated_stamp_map.get(normalize_asin(r.asin)) or timezone.make_aware(
                datetime(1970, 1, 1), timezone.get_current_timezone()
            ),
            reverse=(direction == 'desc'),
        )
    else:
        rows_qs = rows_qs.order_by(order_by)
        rows = rows_qs

    # 运营难度筛选（列选择 + 百分比区间 + 评价数量区间）
    ops_slot = (request.GET.get('ops_slot') or 'any').strip()
    ops_range = (request.GET.get('ops_range') or 'all').strip()
    ops_review_interval = (request.GET.get('ops_review_interval') or 'all').strip()
    if ops_review_interval not in OPS_REVIEW_INTERVAL_OPTIONS:
        ops_review_interval = 'all'
    ops_min = None
    ops_max = None
    preset = {
        '0-5': (0.0, 5.0),
        '5-10': (5.0, 10.0),
        '10-20': (10.0, 20.0),
        '30-100': (30.0, 100.0),
    }
    if ops_range in preset:
        ops_min, ops_max = preset[ops_range]
    elif ops_range == 'custom':
        ops_min = _get_float(request.GET.get('ops_custom_min'))
        ops_max = _get_float(request.GET.get('ops_custom_max'))

    if field != 'updated_at':
        rows = rows_qs
    if ops_range != 'all':
        if ops_slot == '1':
            fields = ('ops_difficulty_1',)
        elif ops_slot == '2':
            fields = ('ops_difficulty_2',)
        elif ops_slot == '3':
            fields = ('ops_difficulty_3',)
        else:
            fields = ('ops_difficulty_1', 'ops_difficulty_2', 'ops_difficulty_3')
        filtered = []
        src_rows = rows if field == 'updated_at' else rows_qs
        for r in src_rows:
            hit = False
            for fn in fields:
                if _ops_match(
                    getattr(r, fn, ''),
                    ops_min,
                    ops_max,
                    review_interval=ops_review_interval,
                ):
                    hit = True
                    break
            if hit:
                filtered.append(r)
        rows = filtered

    if not isinstance(rows, list):
        rows = list(rows)
    rows = dedupe_dashboard_rows(rows, request.user.id, assigned_set)

    page_num = request.GET.get('page') or 1
    per_page = _dashboard_per_page(request)
    paginator = Paginator(rows, per_page)
    page_obj = paginator.get_page(page_num)
    rows = page_obj.object_list
    asins_on_page = {normalize_asin(r.asin) for r in rows}
    verified_norm = {
        normalize_asin(a)
        for a in AsinRoiPackVerification.objects.filter(asin__in=list(asins_on_page)).values_list(
            'asin', flat=True
        )
    }
    _attach_row_assignee_labels(
        rows,
        _assignment_label_map(asins_on_page),
        request.user.id,
        assigned_set,
        viewer_is_superuser,
        roi_verified_asins=verified_norm,
    )
    for r in rows:
        r.data_updated_at = updated_stamp_map.get(normalize_asin(r.asin))
    page_start = page_obj.number
    page_end = min(paginator.num_pages, page_start + 9)
    page_numbers = list(range(page_start, page_end + 1))

    kept = request.GET.copy()
    kept.pop('sort', None)
    kept.pop('dir', None)
    kept.pop('page', None)
    kept.pop('_rv', None)
    filter_qs = kept.urlencode()
    assignee_choices = list(
        User.objects.filter(is_active=True).order_by('username').values_list('username', flat=True)
    )
    return render(
        request,
        'auto_amazon/asin_dashboard.html',
        {
            'rows': rows,
            'page_obj': page_obj,
            'page_numbers': page_numbers,
            'sort': sort_key,
            'dir': direction,
            'filter_qs': filter_qs,
            'per_page': per_page,
            'assignee_choices': assignee_choices,
            'filters': {
                'ops_slot': ops_slot,
                'ops_range': ops_range,
                'ops_review_interval': ops_review_interval,
                'selected_grades': allowed,
                'roi_verified': roi_verified_f if roi_verified_f in ('yes', 'no') else '',
                'follow_filter': follow_f if follow_f in (
                    AsinDashboardRow.FollowStatus.NORMAL,
                    AsinDashboardRow.FollowStatus.PRIORITY,
                ) else '',
                'updated_from': (request.GET.get('updated_from') or '').strip(),
                'updated_to': (request.GET.get('updated_to') or '').strip(),
            },
            'default_exchange_rate': _default_exchange_rate_for_form(),
        },
    )


@login_required
@require_POST
def dashboard_ops_filter(request):
    """按运营难度列 JSON 筛选 data_origin 并写入同目录新文件。"""
    row_id = (request.POST.get('row_id') or '').strip()
    slot = (request.POST.get('slot') or '1').strip()
    field_map = {'1': 'ops_difficulty_1', '2': 'ops_difficulty_2', '3': 'ops_difficulty_3'}
    field = field_map.get(slot)
    if not field or not row_id.isdigit():
        messages.error(request, '参数无效。')
        return redirect('index')
    row = AsinDashboardRow.objects.filter(pk=int(row_id)).first()
    if not row or not user_can_operate_dashboard_row(request.user, row):
        messages.error(request, '无权操作该数据行。')
        return redirect('index')
    payload = getattr(row, field, '') or ''
    ok, msg = run_dashboard_ops_filter(row.asin, payload)
    if ok:
        messages.success(request, msg)
    else:
        messages.error(request, msg)
    return redirect('index')



def _batch_failures_to_json(failures: list[AsinFailure]) -> list[dict]:
    return [
        {'asin': f.asin, 'error': f.error, 'attempts': f.attempts}
        for f in failures
    ]


def _apply_batch_result_to_job(
    key: str,
    *,
    user_id: int,
    batch: BatchRunResult,
    rows_written: int,
    ttl: int = 900,
) -> None:
    """根据批量执行结果写入 job 状态（done / done_with_errors / error）。"""
    ent = cache.get(key) or {}
    ent['success_count'] = batch.success_count
    ent['fail_count'] = batch.fail_count
    ent['failed_asins'] = batch.failed_asins
    ent['failures'] = _batch_failures_to_json(batch.failures)
    ent['rows_written'] = rows_written
    ent['redirect'] = reverse('index')
    ent['user_id'] = user_id

    prog = list(ent.get('progress') or [])
    summary = batch.summary_text()
    if summary and (not prog or prog[-1] != summary):
        prog.append(summary)
    ent['progress'] = prog[-160:]

    if batch.success_count == 0 and batch.fail_count > 0:
        ent['status'] = 'error'
        ent['error'] = batch.summary_text()
        detail_lines = [f'{f.asin}: {f.error}' for f in batch.failures[:50]]
        if len(batch.failures) > 50:
            detail_lines.append(f'… 另有 {len(batch.failures) - 50} 个失败 ASIN')
        ent['error_detail'] = '\n'.join(detail_lines)
        cache.set(key, ent, 3600)
        return

    if batch.fail_count > 0:
        ent['status'] = 'done_with_errors'
        ent['error'] = batch.summary_text()
        cache.set(key, ent, ttl)
        return

    ent['status'] = 'done'
    ent.pop('error', None)
    ent.pop('error_detail', None)
    cache.set(key, ent, ttl)


def _json_sanitize(obj):
    """写入 DB 前把 numpy 等转为可 JSON 序列化的原生类型。"""
    try:
        import numpy as np

        if isinstance(obj, (np.floating, np.float32, np.float64)):
            return float(obj)
        if isinstance(obj, (np.integer, np.int32, np.int64)):
            return int(obj)
    except ImportError:
        pass
    if isinstance(obj, dict):
        return {k: _json_sanitize(v) for k, v in obj.items()}
    if isinstance(obj, (list, tuple)):
        return [_json_sanitize(x) for x in obj]
    return obj


def _ops_cell_json(keyword: str, review_inner: dict) -> str:
    """
    一个关键词对应一个运营难度字段，结构为：
    { "关键词文本": { "review_interval": { 区间[], 平均销量[], 运营难度[] } } }
    整体作为一块数据，供前端图形化展示，不拆行展示区间。
    """
    inner = _json_sanitize(review_inner)
    payload = {keyword: {'review_interval': inner}}
    return json.dumps(payload, ensure_ascii=False)


def _ops_difficulty_three_fields(data: dict, asin: str) -> tuple[str, str, str]:
    """
    运营难度1/2/3 = 前三个关键词，每个字段存该关键词下完整 review_interval（见 _ops_cell_json）。
    """
    ri_outer = data.get('review_interval')
    if not isinstance(ri_outer, dict):
        return '', '', ''

    asin_node = ri_outer.get(asin)
    if not isinstance(asin_node, dict):
        if len(ri_outer) == 1:
            asin_node = next(iter(ri_outer.values()))
        else:
            return '', '', ''
    if not isinstance(asin_node, dict):
        return '', '', ''

    out: list[str] = []
    for kw in asin_node.keys():
        inner = asin_node.get(kw)
        if not isinstance(inner, dict):
            continue
        block = inner.get('review_interval')
        if not isinstance(block, dict):
            continue
        out.append(_ops_cell_json(kw, block))
        if len(out) >= 3:
            break

    while len(out) < 3:
        out.append('')
    return out[0], out[1], out[2]


def _sync_dashboard_from_roi_us_pack(user: User, asin: str, rows: list) -> None:
    """Excel ROI-US-pack 保存/重算后，将去广告毛利率、去广告投产比等同步到看板行。"""
    metrics = extract_dashboard_metrics_from_roi_us_pack(rows)
    if not metrics:
        return
    row = resolve_dashboard_row_for_persist(user, asin)
    if row is None:
        return
    updates = {k: v for k, v in metrics.items() if v is not None}
    if not updates:
        return
    AsinDashboardRow.objects.filter(pk=row.pk).update(**updates)
    AsinDashboardRow.objects.filter(user_id=user.id, asin=normalize_asin(asin)).exclude(
        pk=row.pk
    ).delete()


def _persist_wizard_results(
    user_id: int,
    result: dict,
    parity: float,
    target_row_ids: dict[str, int] | None = None,
) -> int:
    """按指定看板行或 (user, asin) 覆盖写入：已存在则 update，不新增重复行。"""

    def _to_float(v):
        if v is None:
            return None
        try:
            return float(v)
        except (TypeError, ValueError):
            return None

    n = 0
    touched_asins: set[str] = set()
    row_id_map = target_row_ids or {}
    user = User.objects.filter(pk=user_id).first()
    for asin, data in result.items():
        if str(asin).startswith('__'):
            continue
        if not isinstance(data, dict):
            continue
        mr = _to_float(data.get('monthly_results'))
        pm = _to_float(data.get('profit_margin'))
        up = _to_float(data.get('unit_purchase'))
        mp1 = _to_float(data.get('monthly_profit1'))
        ppo = _to_float(data.get('profit_per_order'))
        rp = _to_float(data.get('ranking_percent'))
        hd = _to_float(data.get('head_distance'))
        ac = _to_float(data.get('actual_cost'))
        ad_roi = _to_float(data.get('ad_removed_roi'))
        if ad_roi is not None:
            ad_roi = ad_roi * 100

        existing = resolve_dashboard_row_for_persist(user, asin, row_id_map) if user else None
        if existing is None:
            existing = AsinDashboardRow.objects.filter(user_id=user_id, asin=asin).first()
        ranking_to_store = rp
        if existing is not None and existing.ranking_percent is not None:
            try:
                if float(existing.ranking_percent) != 0.0:
                    ranking_to_store = existing.ranking_percent
            except (TypeError, ValueError):
                pass

        # mst = None
        # if mr is not None and ppo is not None:
        #     mst = round(mr * ppo, 2)
        hat = None
        if hd is not None and up is not None:
            hat = round(hd + up, 2)

        o1, o2, o3 = _ops_difficulty_three_fields(data, asin)
        if existing is not None:
            if not (o1 or '').strip() and (existing.ops_difficulty_1 or '').strip():
                o1 = existing.ops_difficulty_1
            if not (o2 or '').strip() and (existing.ops_difficulty_2 or '').strip():
                o2 = existing.ops_difficulty_2
            if not (o3 or '').strip() and (existing.ops_difficulty_3 or '').strip():
                o3 = existing.ops_difficulty_3

        defaults = {
            'profit_margin': pm,
            'ranking_percent': ranking_to_store,
            'ops_difficulty_1': o1,
            'ops_difficulty_2': o2,
            'ops_difficulty_3': o3,
            'unit_purchase': up,
            'monthly_results': mr,
            'profit_per_order': ppo,
            'monthly_profit1': mp1,
            'monthly_sales_total': mp1,
            'head_distance': hd,
            'actual_cost': ac,
            'head_actual_total': hat,
            'ad_removed_roi': ad_roi,
            'product_grade': product_grade_from_monthly_profit1(mp1),
            'sales_trend_json': '',
            'exchange_rate': parity,
        }
        if existing is not None:
            AsinDashboardRow.objects.filter(pk=existing.pk).update(**defaults)
            AsinDashboardRow.objects.filter(user_id=user_id, asin=asin).exclude(pk=existing.pk).delete()
        else:
            AsinDashboardRow.objects.create(user_id=user_id, asin=asin, **defaults)
        n += 1
        touched_asins.add(asin)
    _touch_asin_updates(touched_asins)
    return n


def _run_wizard_job(
        job_id: str,
        user_id: int,
        asins: list[str] | None,
        parity: float,
        cost_overrides: dict | None = None,
        target_row_ids: dict[str, int] | None = None,
) -> None:
    key = _wizard_job_key(job_id)
    close_old_connections()
    lock = AsinComputeLock(asins, f'user:{user_id}:{job_id}')
    blocked = lock.acquire()
    if blocked:
        ent = cache.get(key) or {}
        ent['status'] = 'error'
        ent['error'] = f'ASIN {blocked} 正在被其他任务计算，请稍后重试。'
        ent['user_id'] = user_id
        cache.set(key, ent, 3600)
        _clear_user_active_job(user_id, job_id)
        close_old_connections()
        return
    try:

        def on_line(line: str) -> None:
            if line.startswith('PROGRESS:'):
                line = line[9:].strip() or line
            ent = cache.get(key) or {}
            prog = ent.get('progress', [])
            short = line[:500] if len(line) > 500 else line
            prog.append(short)
            if len(prog) > 160:
                prog = prog[-160:]
            ent['progress'] = prog
            ent['status'] = 'running'
            ent['user_id'] = user_id
            cache.set(key, ent, WIZARD_JOB_TTL)

        ent = cache.get(key) or {}
        if ent.get('status') == 'queued':
            ent['status'] = 'running'
            prog = ent.get('progress', [])
            prog.append('Worker 已开始执行 ROI 计算…')
            ent['progress'] = prog
            cache.set(key, ent, WIZARD_JOB_TTL)

        def on_progress(msg: str) -> None:
            on_line(msg)

        asin_list = ordered_unique_asins(asins)
        cost_overrides = _merge_cost_overrides_from_db(asin_list or asins, cost_overrides)
        written = 0

        if asin_list:
            if use_sequential_roi():
                def _persist_one(asin: str, part: dict) -> None:
                    nonlocal written
                    row_ids = None
                    if target_row_ids and asin in target_row_ids:
                        row_ids = {asin: target_row_ids[asin]}
                    written += _persist_wizard_results(
                        user_id, part, parity, target_row_ids=row_ids
                    )

                def _on_asin_failed(failure: AsinFailure) -> None:
                    ent = cache.get(key) or {}
                    ent['success_count'] = len(ent.get('succeeded_asins', []))
                    failed = list(ent.get('failed_asins', []))
                    if failure.asin not in failed:
                        failed.append(failure.asin)
                    ent['failed_asins'] = failed
                    ent['fail_count'] = len(failed)
                    failures = list(ent.get('failures', []))
                    failures.append(
                        {
                            'asin': failure.asin,
                            'error': failure.error,
                            'attempts': failure.attempts,
                        }
                    )
                    ent['failures'] = failures[-200:]
                    cache.set(key, ent, WIZARD_JOB_TTL)

                batch = run_roi_asins_sequential(
                    asin_list,
                    parity,
                    cost_overrides=cost_overrides,
                    on_stderr_line=on_line,
                    on_progress=on_progress,
                    on_asin_done=_persist_one,
                    on_asin_failed=_on_asin_failed,
                )
            else:
                on_progress(
                    f'批量计算 ROI：共 {len(asin_list)} 个 ASIN'
                    f'（一次拉取卖家精灵数据，请勿与逐个模式混淆）…'
                )
                merged = run_seller_wizard(
                    asin_list,
                    parity,
                    cost_overrides=cost_overrides,
                    on_stderr_line=on_line,
                )
                failure_details = None
                if isinstance(merged, dict):
                    failure_details = merged.pop('__roi_failures__', None)
                written = _persist_wizard_results(
                    user_id,
                    merged,
                    parity,
                    target_row_ids=target_row_ids,
                )
                batch = batch_result_from_wizard_output(
                    asin_list, merged, failure_details
                )
                on_progress(batch.summary_text())
                for failure in batch.failures[:30]:
                    on_progress(f'  {failure.asin}: {failure.error}')
                if len(batch.failures) > 30:
                    on_progress(f'  … 另有 {len(batch.failures) - 30} 个失败 ASIN')
            _apply_batch_result_to_job(
                key,
                user_id=user_id,
                batch=batch,
                rows_written=written,
            )
        else:
            result = run_seller_wizard(
                asins, parity, cost_overrides=cost_overrides, on_stderr_line=on_line
            )
            written = _persist_wizard_results(
                user_id, result, parity, target_row_ids=target_row_ids
            )
            ent = cache.get(key) or {}
            ent['status'] = 'done'
            ent['rows_written'] = written
            ent['success_count'] = written
            ent['fail_count'] = 0
            ent['failed_asins'] = []
            ent['failures'] = []
            ent['redirect'] = reverse('index')
            ent['user_id'] = user_id
            cache.set(key, ent, 900)
    except Exception as e:
        import traceback

        tb = traceback.format_exc()
        ent = cache.get(key) or {}
        ent['status'] = 'error'
        ent['error'] = f'{type(e).__name__}: {e}'
        ent['error_detail'] = tb[-6000:] if len(tb) > 6000 else tb
        ent['user_id'] = user_id
        cache.set(key, ent, 3600)
    finally:
        lock.release_all()
        _clear_user_active_job(user_id, job_id)
        close_old_connections()


def _run_ad_difficulty_job(
    job_id: str,
    user_id: int,
    asins: list[str],
    target_row_pks: list[int] | None = None,
) -> None:
    key = _wizard_job_key(job_id)
    close_old_connections()
    lock = AsinComputeLock(asins, f'user:{user_id}:{job_id}')
    blocked = lock.acquire()
    if blocked:
        ent = cache.get(key) or {}
        ent['status'] = 'error'
        ent['error'] = f'ASIN {blocked} 正在被其他任务计算，请稍后重试。'
        ent['user_id'] = user_id
        cache.set(key, ent, 3600)
        _clear_user_active_job(user_id, job_id)
        close_old_connections()
        return
    try:
        ent = cache.get(key) or {}
        if ent.get('status') == 'queued':
            ent['status'] = 'running'
            prog = ent.get('progress', [])
            prog.append('Worker 已开始执行广告难度计算…')
            ent['progress'] = prog
            cache.set(key, ent, WIZARD_JOB_TTL)

        def on_line(line: str) -> None:
            if line.startswith('PROGRESS:'):
                line = line[9:].strip() or line
            ent = cache.get(key) or {}
            prog = ent.get('progress', [])
            short = line[:500] if len(line) > 500 else line
            prog.append(short)
            if len(prog) > 160:
                prog = prog[-160:]
            ent['progress'] = prog
            ent['status'] = 'running'
            ent['user_id'] = user_id
            cache.set(key, ent, WIZARD_JOB_TTL)

        def on_progress(msg: str) -> None:
            on_line(msg)

        if target_row_pks:
            rows_sel = list(
                AsinDashboardRow.objects.filter(pk__in=target_row_pks).only('pk', 'asin')
            )
        else:
            rows_sel = list(
                AsinDashboardRow.objects.filter(user_id=user_id, asin__in=asins).only('pk', 'asin')
            )
        asin_to_pk = {normalize_asin(r.asin): r.pk for r in rows_sel if normalize_asin(r.asin)}
        asin_list = ordered_unique_asins(asins)
        asin_list = [a for a in asin_list if a in asin_to_pk]

        updated = 0

        def _persist_ad(asin: str, part: dict) -> None:
            nonlocal updated
            pk = asin_to_pk.get(asin)
            if not pk:
                return
            payload = part.get(asin) or {}
            rp = payload.get('ranking_percent')
            try:
                rp_num = float(rp)
            except (TypeError, ValueError):
                rp_num = 0.0
            AsinDashboardRow.objects.filter(pk=pk).update(ranking_percent=rp_num)
            updated += 1

        def _on_ad_failed(failure: AsinFailure) -> None:
            ent = cache.get(key) or {}
            failed = list(ent.get('failed_asins', []))
            if failure.asin not in failed:
                failed.append(failure.asin)
            ent['failed_asins'] = failed
            ent['fail_count'] = len(failed)
            failures = list(ent.get('failures', []))
            failures.append(
                {
                    'asin': failure.asin,
                    'error': failure.error,
                    'attempts': failure.attempts,
                }
            )
            ent['failures'] = failures[-200:]
            cache.set(key, ent, WIZARD_JOB_TTL)

        batch = run_ad_difficulty_asins_batch(
            asin_list,
            on_stderr_line=on_line,
            on_progress=on_progress,
            on_asin_done=_persist_ad,
            on_asin_failed=_on_ad_failed,
        )
        _touch_asin_updates(batch.succeeded)

        _apply_batch_result_to_job(
            key,
            user_id=user_id,
            batch=batch,
            rows_written=updated,
        )
    except Exception as e:
        import traceback

        tb = traceback.format_exc()
        ent = cache.get(key) or {}
        ent['status'] = 'error'
        ent['error'] = f'{type(e).__name__}: {e}'
        ent['error_detail'] = tb[-6000:] if len(tb) > 6000 else tb
        ent['user_id'] = user_id
        cache.set(key, ent, 3600)
    finally:
        lock.release_all()
        _clear_user_active_job(user_id, job_id)
        close_old_connections()


def _parse_compute_local_form(request):
    """计算本地数据：仅要求汇率，ASIN 由本地逻辑决定。"""
    parity_raw = (request.POST.get('exchange_rate') or '').strip()

    try:
        parity = float(parity_raw)
    except ValueError:
        return None, '请输入有效汇率（数字）。'
    if parity <= 0:
        return None, '汇率必须大于 0。'
    return parity, None


def _discover_local_asins(force_recompute: bool) -> tuple[list[str], list[str]]:
    """
    扫描 media/file 下 ASIN 目录。
    - force_recompute=False: 跳过目录根下已有 *ROI-US-pack*.xlsx 的 ASIN
    - force_recompute=True: 全部纳入
    返回 (待计算 asins, 被跳过 asins)。
    """
    root = media_root()
    pending: list[str] = []
    skipped: list[str] = []
    for p in root.iterdir():
        if not p.is_dir():
            continue
        asin = p.name.strip().upper()
        if not re.match(r'^B0[A-Z0-9]{8}$', asin):
            continue
        has_roi = any(
            c.is_file() and 'ROI-US-pack' in c.name and c.suffix.lower() in ('.xlsx', '.xlsm')
            for c in p.iterdir()
        )
        if has_roi and not force_recompute:
            skipped.append(asin)
            continue
        pending.append(asin)
    return sorted(pending), sorted(skipped)


def _compute_allowed_asins_for_user(user) -> set[str] | None:
    """
    计算当前用户可参与 ROI 计算的 ASIN 集合。
    - 超管：None（不限制）
    - 普通用户：本人导入 + 被分配
    """
    if getattr(user, 'is_superuser', False):
        return None
    imported = user_imported_asin_codes(user)
    assigned = {normalize_asin(a) for a in user_assigned_asin_codes(user)}
    return {a for a in (imported | assigned) if a}


@login_required
@require_POST
def upload_start(request):
    active = get_active_job_for_user(request.user.id)
    if active:
        return JsonResponse(
            {
                'ok': False,
                'error': (
                    f'已有任务进行中（{active["status_label"]}），'
                    f'请查看进度或解除占用后再发起新任务。'
                ),
                'job_id': active['job_id'],
                'active_job': active,
            },
            status=409,
        )
    parity, err = _parse_compute_local_form(request)
    if err:
        return JsonResponse({'ok': False, 'error': err}, status=400)
    selected = [normalize_asin(x) for x in request.POST.getlist('asins') if str(x).strip()]
    if not selected:
        return JsonResponse({'ok': False, 'error': '请先勾选至少一个待计算 ASIN。'}, status=400)
    include_recomputed = str(request.POST.get('recompute_existing') or '').lower() in ('1', 'true', 'on', 'yes')
    pending, skipped = _discover_local_asins(force_recompute=include_recomputed)
    allowed_asins = _compute_allowed_asins_for_user(request.user)
    if allowed_asins is not None:
        pending = [a for a in pending if a in allowed_asins]
        skipped = [a for a in skipped if a in allowed_asins]
    allowed_set = set(pending)
    asins = sorted({a for a in selected if a in allowed_set})
    if not asins:
        return JsonResponse(
            {
                'ok': False,
                'error': '勾选的 ASIN 不在当前可计算列表中，请刷新页面后重试。',
            },
            status=400,
        )
    job_id = uuid.uuid4()
    jid = str(job_id)
    key = _wizard_job_key(jid)
    _set_user_active_job(request.user.id, jid)
    cost_overrides = _merge_cost_overrides_from_db(asins, None)
    prog = [
        '本地计算任务已创建…',
        f'当前汇率：{parity}。',
        f'本次待计算 ASIN：{len(asins)} 个。',
    ]
    if skipped and not include_recomputed:
        prog.append(f'当前未计算列表之外（已存在 ROI-US-pack）：{len(skipped)} 个。')
    cache.set(
        key,
        new_job_entry(
            request.user.id,
            prog,
            task_type='wizard',
            parity=parity,
            asins=asins,
            cost_overrides=cost_overrides,
        ),
        WIZARD_JOB_TTL,
    )
    from .rq_enqueue import dispatch_wizard_job

    dispatch_wizard_job(jid, request.user.id, asins, parity, cost_overrides=cost_overrides)
    return JsonResponse({'ok': True, 'job_id': jid, 'selected_count': len(asins), 'skipped_count': len(skipped)})


@login_required
def upload_job_status(request, job_id):
    ent = refresh_job_entry(str(job_id))
    if not ent:
        return JsonResponse({'ok': False, 'error': '任务不存在或已过期'}, status=404)
    if ent.get('user_id') != request.user.id:
        return JsonResponse({'ok': False, 'error': '无权访问该任务'}, status=403)
    payload = job_status_payload(str(job_id), ent)
    payload['ok'] = True
    return JsonResponse(payload)


@login_required
@require_GET
def active_wizard_job(request):
    """当前用户占用中的 ROI/广告难度任务（用于页面提示条）。"""
    active = get_active_job_for_user(request.user.id)
    if not active:
        return JsonResponse({'ok': True, 'has_active': False, 'active_job': None})
    return JsonResponse({'ok': True, 'has_active': True, 'active_job': active})


@login_required
@require_POST
def dismiss_active_wizard_job(request):
    """解除任务占用（Worker 未启动导致长期 queued 时可手动释放）。"""
    job_id = (request.POST.get('job_id') or '').strip() or None
    ok, msg = dismiss_active_job(request.user.id, job_id)
    if not ok:
        return JsonResponse({'ok': False, 'error': msg}, status=403)
    return JsonResponse({'ok': True, 'message': msg})


@login_required
def upload_page(request):
    return redirect('compute_roi')


@login_required
def compute_roi_page(request):
    include_recomputed = str(request.GET.get('recompute_existing') or '').lower() in ('1', 'true', 'on', 'yes')
    pending_asins, skipped_asins = _discover_local_asins(force_recompute=False)
    allowed_asins = _compute_allowed_asins_for_user(request.user)
    if allowed_asins is not None:
        pending_asins = [a for a in pending_asins if a in allowed_asins]
        skipped_asins = [a for a in skipped_asins if a in allowed_asins]
    all_asins = sorted(set(pending_asins + skipped_asins)) if include_recomputed else pending_asins
    asin_kw = (request.GET.get('asin_kw') or '').strip().upper()
    if asin_kw:
        all_asins = [a for a in all_asins if asin_kw in a]
    per_page = _compute_roi_per_page(request)
    paginator = Paginator(all_asins, per_page)
    page_obj = paginator.get_page(request.GET.get('page') or 1)
    kept = request.GET.copy()
    kept.pop('page', None)
    filter_qs = kept.urlencode()
    active_job = get_active_job_for_user(request.user.id)
    return render(
        request,
        'auto_amazon/compute_roi.html',
        {
            'rows': list(page_obj.object_list),
            'page_obj': page_obj,
            'per_page': per_page,
            'filter_qs': filter_qs,
            'asin_kw': asin_kw,
            'total_pending': len(pending_asins),
            'total_all': len(all_asins),
            'include_recomputed': include_recomputed,
            'default_exchange_rate': _default_exchange_rate_for_form(),
            'active_job': active_job,
        },
    )


@login_required
def schedule_messages_page(request):
    """定时任务推送消息：普通用户仅看发给自己的消息。"""
    qs = ScheduledTaskMessage.objects.select_related('recipient', 'dashboard_row')
    if not request.user.is_superuser:
        qs = qs.filter(recipient=request.user)

    alert_filter = (request.GET.get('alert') or '').strip().lower()
    if alert_filter == 'alert':
        qs = qs.filter(alert_status=ScheduledTaskMessage.AlertStatus.ALERT)
    elif alert_filter == 'eliminate':
        qs = qs.filter(alert_status=ScheduledTaskMessage.AlertStatus.ELIMINATE)
    elif alert_filter in ('normal', 'no'):
        qs = qs.filter(alert_status=ScheduledTaskMessage.AlertStatus.NORMAL)
    elif alert_filter == 'yes':
        qs = qs.filter(
            alert_status__in=(
                ScheduledTaskMessage.AlertStatus.ALERT,
                ScheduledTaskMessage.AlertStatus.ELIMINATE,
            )
        )

    asin_q = (request.GET.get('asin') or '').strip().upper()
    if asin_q:
        qs = qs.filter(asin__icontains=asin_q)

    date_from_raw = (request.GET.get('date_from') or '').strip()
    date_to_raw = (request.GET.get('date_to') or '').strip()
    date_from = None
    date_to = None
    if date_from_raw:
        try:
            date_from = datetime.strptime(date_from_raw, '%Y-%m-%d').date()
            qs = qs.filter(sent_at__date__gte=date_from)
        except ValueError:
            date_from_raw = ''
    if date_to_raw:
        try:
            date_to = datetime.strptime(date_to_raw, '%Y-%m-%d').date()
            qs = qs.filter(sent_at__date__lte=date_to)
        except ValueError:
            date_to_raw = ''

    qs = qs.order_by('-sent_at', '-id')

    per_page = 30
    paginator = Paginator(qs, per_page)
    page_obj = paginator.get_page(request.GET.get('page') or 1)

    page_start = max(1, page_obj.number - 4)
    page_end = min(paginator.num_pages, page_obj.number + 5)
    page_numbers = list(range(page_start, page_end + 1)) if paginator.num_pages else []

    kept = request.GET.copy()
    kept.pop('page', None)
    filter_qs = kept.urlencode()

    return render(
        request,
        'auto_amazon/schedule_messages.html',
        {
            'page_obj': page_obj,
            'page_numbers': page_numbers,
            'filter_qs': filter_qs,
            'filters': {
                'alert': alert_filter,
                'asin': asin_q,
                'date_from': date_from_raw,
                'date_to': date_to_raw,
            },
        },
    )


@login_required
@user_passes_test(lambda u: u.is_superuser)
def credentials_config_page(request):
    """SIF / 卖家精灵凭证配置（超级管理员）。"""
    from .credentials_config import (
        read_credentials_page_context,
        read_seller_password,
        refresh_sif_token,
        write_ao_lo_to_n,
        write_seller_child_ids,
        write_seller_password,
        write_seller_username,
        write_sif_authorization,
    )

    if request.method == 'POST':
        action = (request.POST.get('action') or 'save').strip()
        if action == 'refresh':
            ok, msg = refresh_sif_token()
            if ok:
                preview = msg[:48] + '…' if len(msg) > 48 else msg
                messages.success(request, f'SIF Token 刷新成功：{preview}')
            else:
                messages.error(request, f'刷新失败：{msg}')
        else:
            child_raw = (request.POST.get('child_ids') or '').strip()
            child_ids = [x.strip() for x in child_raw.replace('\n', ',').split(',') if x.strip()]
            if not child_ids:
                messages.error(request, '子账号 ID 不能为空。')
                return redirect('credentials_config')

            username = (request.POST.get('seller_username') or '').strip()
            if not username:
                messages.error(request, '子账号用户名不能为空。')
                return redirect('credentials_config')

            password_in = (request.POST.get('seller_password') or '').strip()
            ao_lo_to_n = (request.POST.get('ao_lo_to_n') or '').strip()
            auth = (request.POST.get('authorization') or '').strip()

            write_seller_child_ids(child_ids)
            write_seller_username(username)
            if password_in:
                write_seller_password(password_in)
            elif not read_seller_password():
                messages.error(request, '首次配置须填写子账号密码。')
                return redirect('credentials_config')
            write_ao_lo_to_n(ao_lo_to_n)
            write_sif_authorization(auth)

            messages.success(request, '凭证已保存。')
            if not auth:
                messages.warning(request, 'SIF authorization 为空，CPC 将使用默认值直至填写。')
            if not ao_lo_to_n:
                messages.warning(request, 'ao_lo_to_n 为空，卖家精灵登录可能失败，请尽快填写。')
        return redirect('credentials_config')

    return render(request, 'auto_amazon/credentials_config.html', read_credentials_page_context())


def sif_config_page(request):
    """旧 URL 兼容。"""
    return credentials_config_page(request)


@login_required
def asin_upload_page(request):
    """上传 ASIN 页面：所有登录用户可见。"""
    if request.method == 'POST':
        up = request.FILES.get('asin_file')
        if up is None:
            messages.error(request, '请选择要上传的文件。')
            return redirect('asin_upload')
        raw = up.read()
        if not raw:
            messages.error(request, '文件为空。')
            return redirect('asin_upload')
        batch, err = ingest_asin_upload(request.user, getattr(up, 'name', '') or '', raw)
        if err:
            messages.error(request, err)
            return redirect('asin_upload')
        if batch.new_count == 0:
            messages.warning(
                request,
                f'文件共识别 {batch.total_in_file} 个 ASIN，均已存在于库中，未新增。',
            )
        else:
            msg = (
                f'上传成功：文件内 {batch.total_in_file} 个 ASIN，'
                f'新增 {batch.new_count} 个'
            )
            if batch.skipped_count:
                msg += f'，跳过重复 {batch.skipped_count} 个'
            msg += '。'
            messages.success(request, msg)
        return redirect('asin_upload')

    qs = AsinUploadBatch.objects.select_related('user', 'downloaded_by').all()

    uploader = (request.GET.get('uploader') or '').strip()
    if uploader:
        qs = qs.filter(user__username__iexact=uploader)

    dl_status = (request.GET.get('download_status') or '').strip().lower()
    if dl_status == 'yes':
        qs = qs.filter(is_downloaded=True)
    elif dl_status == 'no':
        qs = qs.filter(is_downloaded=False)

    paginator = Paginator(qs, 20)
    page_obj = paginator.get_page(request.GET.get('page') or 1)

    uploader_choices = list(
        User.objects.filter(asin_upload_batches__isnull=False)
        .distinct()
        .order_by('username')
        .values_list('username', flat=True)
    )
    total_catalog = AsinCatalogItem.objects.count()

    kept = request.GET.copy()
    kept.pop('page', None)
    filter_qs = kept.urlencode()

    return render(
        request,
        'auto_amazon/asin_upload.html',
        {
            'page_obj': page_obj,
            'batches': page_obj.object_list,
            'uploader_choices': uploader_choices,
            'filter_qs': filter_qs,
            'filters': {
                'uploader': uploader,
                'download_status': dl_status,
            },
            'total_catalog': total_catalog,
            'total_batches': AsinUploadBatch.objects.count(),
        },
    )


@login_required
def asin_upload_export(request, batch_id: int):
    """导出批次 ASIN 为 txt，并标记为已下载。"""
    is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
    batch = AsinUploadBatch.objects.filter(pk=batch_id).first()
    if batch is None:
        if is_ajax:
            return JsonResponse({'ok': False, 'error': '上传记录不存在'}, status=404)
        raise Http404('上传记录不存在')
    lines = batch_asin_lines(batch)
    if not lines:
        err = '该批次没有可导出的 ASIN（可能全部为重复未入库）。'
        if is_ajax:
            return JsonResponse({'ok': False, 'error': err}, status=400)
        messages.warning(request, err)
        return redirect('asin_upload')

    if not batch.is_downloaded:
        batch.is_downloaded = True
        batch.downloaded_at = timezone.now()
        batch.downloaded_by = request.user
        batch.save(update_fields=['is_downloaded', 'downloaded_at', 'downloaded_by'])

    content = '\n'.join(lines) + '\n'
    fn = f'asin_batch_{batch_id}_{batch.created_at:%Y%m%d_%H%M%S}.txt'
    resp = HttpResponse(content, content_type='text/plain; charset=utf-8')
    resp['Content-Disposition'] = f'attachment; filename="{fn}"'
    if is_ajax:
        resp['X-Download-Username'] = request.user.username
    return resp


@login_required
@user_passes_test(lambda u: bool(getattr(u, 'is_superuser', False)))
def fetch_data_page(request):
    return redirect('asin_upload')


@login_required
def excel_page(request):
    assignable_users = list(
        User.objects.filter(is_active=True).order_by('username').values('id', 'username')
    )
    uploader_user_ids = ImportedMediaPath.objects.values_list('user_id', flat=True).distinct()
    uploader_users = list(
        User.objects.filter(id__in=uploader_user_ids, is_active=True)
        .order_by('username')
        .values('id', 'username')
    )
    return render(
        request,
        'auto_amazon/excel.html',
        {
            'assignable_users': assignable_users,
            'assignable_users_json': json.dumps(assignable_users, ensure_ascii=False),
            'uploader_users_json': json.dumps(uploader_users, ensure_ascii=False),
            'is_superuser': bool(request.user.is_superuser),
        },
    )


@login_required
def excel_editor_page(request):
    if request.GET.get('local') == '1':
        return render(
            request,
            'auto_amazon/excel_editor.html',
            {'rel_path': '', 'local_mode': True, 'page_title': 'Excel', 'read_only_excel': False},
        )
    rel = (request.GET.get('path') or '').strip().replace('\\', '/')
    excel_home = 'excel' if _require_superuser_for_excel_audit(request) else 'index'
    if not rel:
        messages.error(request, '请从文件库选择要打开的文件。')
        return redirect(excel_home)
    path = safe_media_path_global(rel)
    if path is None or not path.is_file():
        messages.error(request, '文件不存在或路径无效。')
        return redirect(excel_home)
    if path.suffix.lower() not in ('.xlsx', '.xlsm'):
        messages.error(request, '仅支持 .xlsx / .xlsm 文件。')
        return redirect(excel_home)
    if not _require_superuser_for_excel_audit(request) and not user_can_access_excel_media_path(
        request.user, rel
    ):
        messages.error(request, '无权打开该文件（需由管理员将对应 ASIN 文件夹分配给您）。')
        return redirect('index')
    return render(
        request,
        'auto_amazon/excel_editor.html',
        {
            'rel_path': rel,
            'local_mode': False,
            'page_title': _first_word_from_filename(path.name),
            'read_only_excel': False,
        },
    )


@login_required
@require_GET
def excel_browse(request):
    rel = request.GET.get('path', '').strip().replace('\\', '/')
    is_super = _require_superuser_for_excel_audit(request)
    if not is_super and rel:
        if not user_can_access_excel_media_path(request.user, rel):
            return JsonResponse({'ok': False, 'error': '无权限访问该目录'}, status=403)
    target = safe_media_path_global(rel)
    if target is None:
        return JsonResponse({'ok': False, 'error': '路径非法'}, status=400)
    if not target.exists():
        return JsonResponse({'ok': False, 'error': '路径不存在'}, status=404)
    if not target.is_dir():
        return JsonResponse({'ok': False, 'error': '不是文件夹'}, status=400)
    try:
        dashboard_asins = {
            normalize_asin(a) for a in AsinDashboardRow.objects.values_list('asin', flat=True)
        }
        roi_verified_asins = {
            normalize_asin(a) for a in AsinRoiPackVerification.objects.values_list('asin', flat=True)
        }
        updated_map = {
            normalize_asin(a): t
            for a, t in AsinDataUpdateStamp.objects.values_list('asin', 'updated_at')
        }
        assign_map: dict[str, list[str]] = {}
        for obj in AsinFolderAssignment.objects.prefetch_related('assignees').all():
            assign_map[normalize_asin(obj.asin)] = sorted(
                {u.username for u in obj.assignees.all()}
            )
        assigned_set = set(user_assigned_asin_codes(request.user)) if not is_super else set()
        import_roots = user_imported_asin_codes(request.user) if not is_super else set()
        owned_paths = set()
        if not is_super:
            oq = ImportedMediaPath.objects.filter(user=request.user)
            if rel:
                owned_paths = set(
                    oq.filter(Q(rel_path=rel) | Q(rel_path__startswith=f'{rel}/')).values_list(
                        'rel_path', flat=True
                    )
                )
            else:
                owned_paths = set(oq.values_list('rel_path', flat=True))
        uploaders_by_asin: dict[str, set[str]] = defaultdict(set)
        for rpath, uname in ImportedMediaPath.objects.values_list('rel_path', 'user__username'):
            rp = str(rpath).replace('\\', '/').strip('/')
            if not rp:
                continue
            seg = rp.split('/')[0].strip().upper()
            if re.match(r'^B0[A-Z0-9]{8}$', seg):
                uploaders_by_asin[seg].add(uname)
    except DatabaseError as e:
        return JsonResponse(
            {
                'ok': False,
                'error': (
                    '数据库未就绪：请在服务器项目目录执行 python manage.py migrate 后再试。'
                    f' 详情：{e}'
                ),
            },
            status=503,
        )
    uploaded_by_filter = (request.GET.get('uploaded_by') or '').strip()
    items = []
    try:
        entries = sorted(target.iterdir(), key=lambda x: (not x.is_dir(), x.name.lower()))
    except OSError as e:
        return JsonResponse({'ok': False, 'error': str(e)}, status=400)
    path_batch: list[str] = []
    for p in entries:
        if p.name.startswith('.'):
            continue
        cr = f'{rel}/{p.name}' if rel else p.name
        cr = cr.replace('\\', '/').strip('/')
        path_batch.append(cr)
    uploader_by_rel_file: dict[str, str] = {}
    if path_batch:
        uploader_by_rel_file = dict(
            ImportedMediaPath.objects.filter(rel_path__in=path_batch).values_list(
                'rel_path', 'user__username'
            )
        )
    for p in entries:
        if p.name.startswith('.'):
            continue
        child_rel = f'{rel}/{p.name}' if rel else p.name
        child_rel = child_rel.replace('\\', '/').strip('/')
        can_delete_item = True if is_super else (child_rel in owned_paths)
        if p.is_dir():
            nm = p.name.strip().upper()
            is_asin_dir = re.match(r'^B0[A-Z0-9]{8}$', nm) is not None
            if not is_super:
                if rel == '':
                    if not is_asin_dir:
                        continue
                    if nm not in assigned_set and nm not in import_roots:
                        continue
                elif not user_can_access_excel_media_path(request.user, child_rel):
                    continue
            uploaders_list = sorted(uploaders_by_asin.get(nm, [])) if is_asin_dir else []
            if rel == '' and uploaded_by_filter and is_asin_dir:
                if uploaded_by_filter not in uploaders_by_asin.get(nm, set()):
                    continue
            assignees = assign_map.get(nm, []) if is_asin_dir else []
            updated_at = updated_map.get(nm)
            updated_label = timezone.localtime(updated_at).strftime('%Y-%m-%d %H:%M:%S') if updated_at else ''
            if is_asin_dir:
                dir_uploaders = uploaders_list
            else:
                ud = uploader_by_rel_file.get(child_rel)
                dir_uploaders = [ud] if ud else []
            items.append(
                {
                    'name': p.name,
                    'type': 'dir',
                    'calculated': bool(is_asin_dir and nm in dashboard_asins),
                    'is_asin_dir': bool(is_asin_dir),
                    'assigned': bool(assignees),
                    'assignees': assignees,
                    'roi_verified': bool(is_asin_dir and nm in roi_verified_asins),
                    'updated_at': updated_label,
                    'updated_at_ts': int(updated_at.timestamp() * 1000) if updated_at else 0,
                    'imported_by_me': bool(is_asin_dir and nm in import_roots),
                    'uploaders': dir_uploaders,
                    'can_delete': can_delete_item,
                }
            )
        else:
            if not is_super and not user_can_access_excel_media_path(request.user, child_rel):
                continue
            try:
                st = p.stat()
                size = st.st_size
            except OSError:
                size = 0
            fu_file = uploader_by_rel_file.get(child_rel)
            items.append({
                'name': p.name,
                'type': 'file',
                'ext': p.suffix.lower() or '',
                'size': size,
                'calculated': None,
                'is_asin_dir': False,
                'assigned': False,
                'assignees': [],
                'updated_at': '',
                'updated_at_ts': 0,
                'imported_by_me': bool(child_rel in owned_paths),
                'uploaders': [fu_file] if fu_file else [],
                'can_delete': can_delete_item,
            })
    if rel == '':
        items.sort(
            key=lambda it: (
                it.get('type') != 'dir',
                not bool(it.get('is_asin_dir')),
                -(it.get('updated_at_ts') or 0),
                str(it.get('name') or '').lower(),
            )
        )
    parent = parent_rel(rel) if rel else None
    return JsonResponse({
        'ok': True,
        'path': rel,
        'parent': parent,
        'items': items,
    })


def _build_media_zip_for_paths(request, paths: list[str]) -> tuple[BytesIO | None, int]:
    """将文件或文件夹（递归）打包为 ZIP；返回 (buffer, 文件数)。"""
    root = media_root()
    is_super = _require_superuser_for_excel_audit(request)
    pending_writes: list[tuple[Path, str]] = []
    seen_arcs: set[str] = set()

    for rel in paths:
        rel = str(rel).strip().replace('\\', '/')
        if not rel:
            continue
        if not is_super and not user_can_access_excel_media_path(request.user, rel):
            continue
        path = safe_media_path_global(rel)
        if path is None:
            continue
        try:
            path.relative_to(root)
        except ValueError:
            continue
        if path.is_file():
            arc = str(path.relative_to(root)).replace('\\', '/')
            if arc not in seen_arcs:
                seen_arcs.add(arc)
                pending_writes.append((path, arc))
        elif path.is_dir():
            for fp in path.rglob('*'):
                if not fp.is_file():
                    continue
                try:
                    sub_rel = str(fp.relative_to(root)).replace('\\', '/')
                except ValueError:
                    continue
                if not is_super and not user_can_access_excel_media_path(request.user, sub_rel):
                    continue
                if sub_rel not in seen_arcs:
                    seen_arcs.add(sub_rel)
                    pending_writes.append((fp, sub_rel))

    if not pending_writes:
        return None, 0

    buf = BytesIO()
    with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zf:
        for file_path, arc in pending_writes:
            zf.write(file_path, arcname=arc)
    buf.seek(0)
    return buf, len(pending_writes)


@login_required
@require_GET
def excel_download(request):
    rel = (request.GET.get('path') or '').strip().replace('\\', '/')
    if not rel:
        raise Http404
    if not _require_superuser_for_excel_audit(request) and not user_can_access_excel_media_path(
        request.user, rel
    ):
        raise Http404
    path = safe_media_path_global(rel)
    if path is None:
        raise Http404
    if path.is_dir():
        buf, added = _build_media_zip_for_paths(request, [rel])
        if not buf or added == 0:
            raise Http404
        return FileResponse(
            buf,
            as_attachment=True,
            filename=f'{path.name}.zip',
            content_type='application/zip',
        )
    if not path.is_file():
        raise Http404
    try:
        fh = open(path, 'rb')
    except OSError:
        raise Http404
    return FileResponse(fh, as_attachment=True, filename=path.name)


@login_required
@require_POST
def excel_delete(request):
    is_super = _require_superuser_for_excel_audit(request)
    try:
        payload = json.loads(request.body.decode('utf-8'))
    except json.JSONDecodeError:
        return JsonResponse({'ok': False, 'error': '无效的 JSON'}, status=400)
    paths = payload.get('paths')
    if not isinstance(paths, list):
        one = payload.get('path')
        paths = [one] if one else []
    paths = [str(p).strip().replace('\\', '/') for p in paths if p is not None and str(p).strip()]
    if not paths:
        return JsonResponse({'ok': False, 'error': '缺少 path(s)'}, status=400)
    root = media_root()
    for rel in paths:
        if not rel:
            return JsonResponse({'ok': False, 'error': '不能删除根目录'}, status=400)
        path = safe_media_path_global(rel)
        if path is None:
            return JsonResponse({'ok': False, 'error': f'非法路径: {rel}'}, status=400)
        try:
            path.relative_to(root)
        except ValueError:
            return JsonResponse({'ok': False, 'error': f'非法路径: {rel}'}, status=400)
        if path == root:
            return JsonResponse({'ok': False, 'error': '不能删除根目录'}, status=400)
        if not is_super:
            if not user_can_access_excel_media_path(request.user, rel):
                return JsonResponse({'ok': False, 'error': '无权限删除该路径'}, status=403)
            if not ImportedMediaPath.objects.filter(user=request.user, rel_path=rel).exists():
                return JsonResponse(
                    {'ok': False, 'error': '只能删除本人通过「导入」上传的文件或文件夹（管理员分配的资料无删除权限）'},
                    status=403,
                )
            if path.is_dir():
                for fp in path.rglob('*'):
                    if not fp.is_file():
                        continue
                    sub_rel = str(fp.relative_to(root)).replace('\\', '/')
                    if not ImportedMediaPath.objects.filter(user=request.user, rel_path=sub_rel).exists():
                        return JsonResponse(
                            {
                                'ok': False,
                                'error': f'文件夹内存在非本人导入的文件，无法整夹删除：{sub_rel}',
                            },
                            status=403,
                        )
        try:
            if is_super:
                ImportedMediaPath.objects.filter(Q(rel_path=rel) | Q(rel_path__startswith=f'{rel}/')).delete()
            else:
                ImportedMediaPath.objects.filter(user=request.user).filter(
                    Q(rel_path=rel) | Q(rel_path__startswith=f'{rel}/')
                ).delete()
        except DatabaseError as e:
            return JsonResponse(
                {
                    'ok': False,
                    'error': (
                        '数据库未就绪或表缺失：请在项目目录执行 python manage.py migrate 后再试。'
                        f' 详情：{e}'
                    ),
                },
                status=503,
            )
        try:
            if path.is_dir():
                shutil.rmtree(path)
            elif path.is_file():
                path.unlink()
        except OSError as e:
            return JsonResponse({'ok': False, 'error': str(e)}, status=400)
    return JsonResponse({'ok': True})


@login_required
@require_POST
def excel_import_chunk(request):
    """分片上传 ZIP（用于大文件，避免单次请求内存过大）。"""
    try:
        chunk_index = int(request.POST.get('chunk_index', '-1'))
        total_chunks = int(request.POST.get('total_chunks', '0'))
    except ValueError:
        return JsonResponse({'ok': False, 'error': '分片参数无效'}, status=400)
    staging_id = (request.POST.get('staging_id') or '').strip()
    blob = request.FILES.get('file')
    if not blob:
        return JsonResponse({'ok': False, 'error': '缺少分片数据'}, status=400)
    data = b''.join(blob.chunks())
    try:
        sid, meta = import_append_chunk(request.user.id, staging_id or None, chunk_index, total_chunks, data)
    except PermissionError as e:
        return JsonResponse({'ok': False, 'error': str(e)}, status=403)
    except ValueError as e:
        return JsonResponse({'ok': False, 'error': str(e)}, status=400)
    out: dict = {
        'ok': True,
        'staging_id': sid,
        'next_chunk': meta.get('next_chunk'),
        'total_chunks': meta.get('total_chunks'),
    }
    if meta.get('status') == 'ready':
        out['ready'] = True
        out['conflicts'] = meta.get('conflicts') or []
        out['file_count'] = meta.get('file_count', 0)
    return JsonResponse(out)


@login_required
@require_POST
def excel_import_commit(request):
    """确认解压：与现有文件冲突时返回 409，前端可选择覆盖或仅跳过已存在项。"""
    try:
        payload = json.loads(request.body.decode('utf-8'))
    except json.JSONDecodeError:
        return JsonResponse({'ok': False, 'error': '无效的 JSON'}, status=400)
    staging_id = (payload.get('staging_id') or '').strip()
    overwrite = bool(payload.get('overwrite'))
    skip_existing = bool(payload.get('skip_existing'))
    if not staging_id:
        return JsonResponse({'ok': False, 'error': '缺少 staging_id'}, status=400)
    if overwrite and skip_existing:
        return JsonResponse({'ok': False, 'error': '不能同时选择覆盖与跳过已存在文件'}, status=400)
    try:
        sdir, meta = load_ready_staging(request.user.id, staging_id)
    except PermissionError as e:
        return JsonResponse({'ok': False, 'error': str(e)}, status=403)
    except (ValueError, FileNotFoundError) as e:
        return JsonResponse({'ok': False, 'error': str(e)}, status=400)
    mr = media_root()
    conflicts = refresh_conflicts_in_meta(mr, sdir, meta)
    if conflicts and not overwrite and not skip_existing:
        return JsonResponse(
            {
                'ok': False,
                'need_choice': True,
                'conflicts': conflicts,
                'staging_id': staging_id,
            },
            status=409,
        )
    zip_path = sdir / 'upload.zip'
    skip_if_exists = skip_existing and not overwrite
    try:
        written, skipped_existing = extract_zip_to_media_root(zip_path, mr, skip_if_exists=skip_if_exists)
    except Exception as e:
        return JsonResponse({'ok': False, 'error': f'解压失败：{e}'}, status=400)
    try:
        for rp in ensure_dir_nodes_registered(written):
            ImportedMediaPath.objects.update_or_create(rel_path=rp, defaults={'user': request.user})
        asins = {normalize_asin(x.split('/')[0]) for x in written if x}
        _touch_asin_updates(asins)
    except DatabaseError as e:
        cleanup_staging_dir(sdir)
        return JsonResponse(
            {
                'ok': False,
                'error': (
                    'ZIP 已解压到 media/file，但写入导入记录失败；请执行 python manage.py migrate 后刷新。'
                    f' 详情：{e}'
                ),
                'written': len(written),
            },
            status=503,
        )
    cleanup_staging_dir(sdir)
    return JsonResponse(
        {
            'ok': True,
            'written': len(written),
            'skipped_existing': len(skipped_existing),
        }
    )


@login_required
@require_POST
def excel_batch_download(request):
    try:
        payload = json.loads(request.body.decode('utf-8'))
    except json.JSONDecodeError:
        return JsonResponse({'ok': False, 'error': '无效的 JSON'}, status=400)
    paths = payload.get('paths')
    if not isinstance(paths, list) or not paths:
        return JsonResponse({'ok': False, 'error': '缺少 paths'}, status=400)
    buf, added = _build_media_zip_for_paths(request, paths)
    if not buf or added == 0:
        return JsonResponse({'ok': False, 'error': '没有可打包的文件'}, status=400)
    return FileResponse(
        buf,
        as_attachment=True,
        filename='download.zip',
        content_type='application/zip',
    )


@login_required
@require_POST
def excel_load_media(request):
    try:
        payload = json.loads(request.body.decode('utf-8'))
    except json.JSONDecodeError:
        return JsonResponse({'ok': False, 'error': '无效的 JSON'}, status=400)
    rel = (payload.get('path') or '').strip().replace('\\', '/')
    if not _require_superuser_for_excel_audit(request) and not user_can_access_excel_media_path(
        request.user, rel
    ):
        return JsonResponse({'ok': False, 'error': '无权限读取该文件'}, status=403)
    path = safe_media_path_global(rel)
    if path is None or not path.is_file():
        return JsonResponse({'ok': False, 'error': '文件不存在或路径非法'}, status=400)
    if path.suffix.lower() not in ('.xlsx', '.xlsm'):
        return JsonResponse({'ok': False, 'error': '仅支持 .xlsx / .xlsm'}, status=400)
    payload, err = read_active_sheet_rich(path)
    if err:
        # openpyxl 常见报错：xlsx 内部 XML 非法 / 工作表 XML 解析失败
        msg = str(err)
        hint = ''
        low = msg.lower()
        if 'invalid xml' in low or 'could not read worksheets' in low or 'unable to read workbook' in low:
            hint = (
                '（通常是 Excel 文件内部 XML 损坏或由非标准工具生成导致）\n'
                '建议：1) 用 Excel/WPS 打开该文件，选择“打开并修复/修复”，然后“另存为”新的 .xlsx；\n'
                '2) 若无法打开，重新导出/重新生成该文件；\n'
                '3) 也可尝试先下载到本地用 Excel 修复后再上传。'
            )
        return JsonResponse({'ok': False, 'error': f'无法读取 Excel：{msg}{hint}'}, status=400)
    if _is_search_sheet_filename(path.name):
        deleted_asins = _deleted_asin_set_from_data_origin(path)
        _highlight_deleted_rows_for_search(payload, deleted_asins)
    if is_roi_us_pack_filename(path.name):
        parent = path.parent.name.strip().upper()
        if re.match(r'^B0[A-Z0-9]{8}$', parent):
            a = normalize_asin(parent)
            payload['roi_pack_verified'] = AsinRoiPackVerification.objects.filter(asin=a).exists()
    payload['ok'] = True
    payload['path'] = rel
    return JsonResponse(payload)


@login_required
@require_POST
def excel_save_media(request):
    try:
        payload = json.loads(request.body.decode('utf-8'))
    except json.JSONDecodeError:
        return JsonResponse({'ok': False, 'error': '无效的 JSON'}, status=400)
    rel = (payload.get('path') or '').strip().replace('\\', '/')
    if not _require_superuser_for_excel_audit(request) and not user_can_access_excel_media_path(
        request.user, rel
    ):
        return JsonResponse({'ok': False, 'error': '无权限保存该文件'}, status=403)
    rows = payload.get('rows')
    if not isinstance(rows, list):
        return JsonResponse({'ok': False, 'error': '缺少 rows'}, status=400)
    path = safe_media_path_global(rel)
    if path is None or not path.is_file():
        return JsonResponse({'ok': False, 'error': '文件不存在或路径非法'}, status=400)
    if path.suffix.lower() not in ('.xlsx', '.xlsm'):
        return JsonResponse({'ok': False, 'error': '仅支持 .xlsx / .xlsm'}, status=400)
    ok, err = save_sheet_values_preserving_format(path, rows)
    if not ok:
        return JsonResponse({'ok': False, 'error': err or '保存失败'}, status=400)
    parent_asin = normalize_asin(path.parent.name)
    if re.match(r'^B0[A-Z0-9]{8}$', parent_asin):
        _touch_asin_updates({parent_asin})
        if is_roi_us_pack_filename(path.name):
            _sync_dashboard_from_roi_us_pack(request.user, parent_asin, rows)
    return JsonResponse({
        'ok': True,
        'path': rel,
        'rewritten': search_xlsx_save_is_rewrite(path),
    })


@login_required
@require_POST
def excel_import_data_origin(request):
    """将上传 Excel 的数据行追加导入到当前 *_data_origin.xlsx。"""
    rel = (request.POST.get('path') or '').strip().replace('\\', '/')
    if not rel:
        return JsonResponse({'ok': False, 'error': '缺少 path'}, status=400)
    if not _require_superuser_for_excel_audit(request) and not user_can_access_excel_media_path(
        request.user, rel
    ):
        return JsonResponse({'ok': False, 'error': '无权限导入到该文件'}, status=403)
    path = safe_media_path_global(rel)
    if path is None or not path.is_file():
        return JsonResponse({'ok': False, 'error': '目标文件不存在或路径非法'}, status=400)
    if not path.name.lower().endswith('_data_origin.xlsx'):
        return JsonResponse({'ok': False, 'error': '仅支持导入到 *_data_origin.xlsx'}, status=400)

    up = request.FILES.get('file')
    if up is None:
        return JsonResponse({'ok': False, 'error': '缺少上传文件'}, status=400)
    up_name = (getattr(up, 'name', '') or '').lower()
    if not (up_name.endswith('.xlsx') or up_name.endswith('.xlsm')):
        return JsonResponse({'ok': False, 'error': '仅支持上传 .xlsx / .xlsm'}, status=400)
    skip_header = str(request.POST.get('skip_header', '1')).strip() != '0'
    try:
        import_wb = load_workbook(up, read_only=True, data_only=True)
        import_ws = import_wb.active
        imported_rows = []
        max_cols = 0
        for i, row in enumerate(import_ws.iter_rows(values_only=True)):
            if skip_header and i == 0:
                continue
            vals = ['' if v is None else str(v) for v in (row or ())]
            while vals and vals[-1] == '':
                vals.pop()
            if not vals:
                continue
            imported_rows.append(vals)
            if len(vals) > max_cols:
                max_cols = len(vals)
        import_wb.close()
    except Exception as e:
        return JsonResponse({'ok': False, 'error': f'读取上传文件失败：{e}'}, status=400)

    if not imported_rows:
        return JsonResponse({'ok': False, 'error': '上传文件没有可导入的数据行'}, status=400)

    current_payload, err = read_active_sheet_rich(path)
    if err:
        return JsonResponse({'ok': False, 'error': f'读取目标文件失败：{err}'}, status=400)
    current_rows = current_payload.get('rows') or []
    target_cols = 0
    for r in current_rows:
        if isinstance(r, list) and len(r) > target_cols:
            target_cols = len(r)
    final_cols = max(target_cols, max_cols, 1)

    normalized_current = []
    for r in current_rows:
        if not isinstance(r, list):
            r = [r]
        one = []
        for c in r:
            if isinstance(c, dict):
                one.append(c.get('v', ''))
            else:
                one.append(c if c is not None else '')
        while len(one) < final_cols:
            one.append('')
        normalized_current.append(one[:final_cols])

    normalized_import = []
    for r in imported_rows:
        one = list(r)
        while len(one) < final_cols:
            one.append('')
        normalized_import.append(one[:final_cols])

    merged_rows = normalized_current + normalized_import
    ok, save_err = save_sheet_values_preserving_format(path, merged_rows)
    if not ok:
        return JsonResponse({'ok': False, 'error': save_err or '导入保存失败'}, status=400)

    parent_asin = normalize_asin(path.parent.name)
    if re.match(r'^B0[A-Z0-9]{8}$', parent_asin):
        _touch_asin_updates({parent_asin})
    return JsonResponse({'ok': True, 'path': rel, 'imported_rows': len(normalized_import)})


@login_required
@require_POST
def excel_restore_search_row(request):
    """将 Search 表中被标记删除（黄底）的一行恢复写入同目录 *_data_origin.xlsx。"""
    try:
        payload = json.loads(request.body.decode('utf-8'))
    except json.JSONDecodeError:
        return JsonResponse({'ok': False, 'error': '无效的 JSON'}, status=400)
    rel = (payload.get('path') or '').strip().replace('\\', '/')
    if not rel:
        return JsonResponse({'ok': False, 'error': '缺少 path'}, status=400)
    if not _require_superuser_for_excel_audit(request) and not user_can_access_excel_media_path(
        request.user, rel
    ):
        return JsonResponse({'ok': False, 'error': '无权限操作该文件'}, status=403)
    search_path = safe_media_path_global(rel)
    if search_path is None or not search_path.is_file():
        return JsonResponse({'ok': False, 'error': '文件不存在或路径非法'}, status=400)
    if not _is_search_sheet_filename(search_path.name):
        return JsonResponse({'ok': False, 'error': '仅支持 Search(...) 表恢复数据'}, status=400)

    try:
        row_index = int(payload.get('row_index'))
    except (TypeError, ValueError):
        return JsonResponse({'ok': False, 'error': '缺少有效的 row_index'}, status=400)
    if row_index < 1:
        return JsonResponse({'ok': False, 'error': '请选择数据行（不能选表头）'}, status=400)

    origin_path = _find_data_origin_path(search_path.parent)
    if origin_path is None or not origin_path.is_file():
        return JsonResponse({'ok': False, 'error': '同目录下未找到 *_data_origin.xlsx'}, status=400)

    search_payload, err = read_active_sheet_rich(search_path)
    if err:
        return JsonResponse({'ok': False, 'error': f'读取 Search 表失败：{err}'}, status=400)
    search_plain = _rows_to_plain(search_payload.get('rows') or [])
    if row_index >= len(search_plain):
        return JsonResponse({'ok': False, 'error': '行号超出范围'}, status=400)
    if len(search_plain) < 2:
        return JsonResponse({'ok': False, 'error': 'Search 表无数据行'}, status=400)

    search_header = search_plain[0]
    search_row = search_plain[row_index]

    api_record = search_row_to_api_record(search_header, search_row)
    asin = str(api_record.get('asin') or '').strip().upper()
    if not asin:
        return JsonResponse({'ok': False, 'error': '该行缺少 ASIN，无法恢复'}, status=400)

    origin_payload, err = read_active_sheet_rich(origin_path)
    if err:
        return JsonResponse({'ok': False, 'error': f'读取 data_origin 失败：{err}'}, status=400)
    origin_plain = _rows_to_plain(origin_payload.get('rows') or [])
    if not origin_plain:
        return JsonResponse({'ok': False, 'error': 'data_origin 表为空'}, status=400)

    origin_header = origin_plain[0]
    mapped = build_origin_row_from_search(search_header, search_row, origin_header)
    if not any(str(x or '').strip() for x in mapped):
        return JsonResponse({'ok': False, 'error': '该行无有效数据'}, status=400)

    existing = _asin_set_from_plain_rows(origin_plain)
    if asin in existing:
        return JsonResponse({'ok': False, 'error': f'ASIN {asin} 已在 data_origin 中，无需重复恢复'}, status=400)

    deleted_asins = _deleted_asin_set_from_data_origin(search_path)
    if asin not in deleted_asins:
        return JsonResponse({'ok': False, 'error': '该行并非「已删除」数据（黄底行才可恢复）'}, status=400)

    final_cols = max(len(r) for r in origin_plain)
    normalized = []
    for r in origin_plain:
        one = list(r)
        while len(one) < final_cols:
            one.append('')
        normalized.append(one[:final_cols])
    new_row = list(mapped)
    while len(new_row) < final_cols:
        new_row.append('')
    normalized.append(new_row[:final_cols])

    ok, save_err = save_sheet_values_preserving_format(origin_path, normalized)
    if not ok:
        return JsonResponse({'ok': False, 'error': save_err or '写入 data_origin 失败'}, status=400)

    parent_asin = normalize_asin(search_path.parent.parent.name)
    if re.match(r'^B0[A-Z0-9]{8}$', parent_asin):
        _touch_asin_updates({parent_asin})
    return JsonResponse({
        'ok': True,
        'asin': asin,
        'data_origin_path': str(origin_path.relative_to(media_root())).replace('\\', '/'),
    })


@login_required
@require_POST
def excel_recalc_roi_media(request):
    """仅对 ROI-US-pack.xlsx：按表内数据重算指标并写回文件（保留格式与图片）。"""
    try:
        payload = json.loads(request.body.decode('utf-8'))
    except json.JSONDecodeError:
        return JsonResponse({'ok': False, 'error': '无效的 JSON'}, status=400)
    rel = (payload.get('path') or '').strip().replace('\\', '/')
    if not _require_superuser_for_excel_audit(request) and not user_can_access_excel_media_path(
        request.user, rel
    ):
        return JsonResponse({'ok': False, 'error': '无权限重算该文件'}, status=403)
    path = safe_media_path_global(rel)
    if path is None or not path.is_file():
        return JsonResponse({'ok': False, 'error': '文件不存在或路径非法'}, status=400)
    if path.suffix.lower() not in ('.xlsx', '.xlsm'):
        return JsonResponse({'ok': False, 'error': '仅支持 .xlsx / .xlsm'}, status=400)
    if not is_roi_us_pack_filename(path.name):
        return JsonResponse({'ok': False, 'error': '仅支持文件名包含 ROI-US-pack 的 ROI 表'}, status=400)

    sheet_payload, err = read_active_sheet_rich(path)
    if err:
        return JsonResponse({'ok': False, 'error': f'无法读取 Excel：{err}'}, status=400)
    rows = sheet_payload.get('rows') or []
    ok, msg = recalc_roi_us_pack_rows(rows)
    if not ok:
        return JsonResponse({'ok': False, 'error': msg or '重算失败'}, status=400)
    save_ok, save_err = save_sheet_values_preserving_format(path, rows)
    if not save_ok:
        return JsonResponse({'ok': False, 'error': save_err or '保存失败'}, status=400)
    parent_asin = normalize_asin(path.parent.name)
    if re.match(r'^B0[A-Z0-9]{8}$', parent_asin):
        _touch_asin_updates({parent_asin})
        _sync_dashboard_from_roi_us_pack(request.user, parent_asin, rows)
    return JsonResponse({'ok': True, 'path': rel})


@login_required
@require_POST
def excel_confirm_roi_verify(request):
    """在 ROI-US-pack 编辑器中确认校验：将对应 ASIN 在看板与数据审核中标记为「是」。"""
    try:
        payload = json.loads(request.body.decode('utf-8'))
    except json.JSONDecodeError:
        return JsonResponse({'ok': False, 'error': '无效的 JSON'}, status=400)
    rel = (payload.get('path') or '').strip().replace('\\', '/')
    if not rel:
        return JsonResponse({'ok': False, 'error': '缺少 path'}, status=400)
    if not _require_superuser_for_excel_audit(request) and not user_can_access_excel_media_path(
        request.user, rel
    ):
        return JsonResponse({'ok': False, 'error': '无权限'}, status=403)
    path = safe_media_path_global(rel)
    if path is None or not path.is_file():
        return JsonResponse({'ok': False, 'error': '文件不存在或路径非法'}, status=400)
    if not is_roi_us_pack_filename(path.name):
        return JsonResponse({'ok': False, 'error': '仅支持 ROI-US-pack 表确认校验'}, status=400)
    parent = path.parent.name.strip().upper()
    if not re.match(r'^B0[A-Z0-9]{8}$', parent):
        return JsonResponse({'ok': False, 'error': '文件须位于 ASIN 文件夹下'}, status=400)
    asin = normalize_asin(parent)
    AsinRoiPackVerification.objects.update_or_create(
        asin=asin,
        defaults={'verified_by': request.user},
    )
    return JsonResponse({'ok': True, 'asin': asin})


@login_required
@require_POST
def excel_assign_folders(request):
    """超级管理员：将 ASIN 文件夹分配给若干用户（覆盖该 ASIN 下的分配名单）。"""
    if not _require_superuser_for_excel_audit(request):
        return JsonResponse({'ok': False, 'error': '无权限'}, status=403)
    try:
        body = json.loads(request.body.decode('utf-8'))
    except json.JSONDecodeError:
        return JsonResponse({'ok': False, 'error': '无效的 JSON'}, status=400)
    asins_raw = body.get('asins') or []
    user_ids = body.get('user_ids') or []
    if not isinstance(asins_raw, list) or not asins_raw:
        return JsonResponse({'ok': False, 'error': '缺少 asins'}, status=400)
    if not isinstance(user_ids, list):
        user_ids = []
    uid_set = {int(x) for x in user_ids if str(x).isdigit()}
    users = list(User.objects.filter(pk__in=uid_set, is_active=True))
    updated = 0
    for raw in asins_raw:
        a = normalize_asin(str(raw))
        if not re.match(r'^B0[A-Z0-9]{8}$', a):
            continue
        obj, _created = AsinFolderAssignment.objects.get_or_create(asin=a)
        obj.assignees.set(users)
        obj.assigned_by = request.user
        obj.save()
        updated += 1
    return JsonResponse({'ok': True, 'updated': updated})


@login_required
@require_POST
def dashboard_export_excel(request):
    """导出勾选看板行为 Excel（本人数据 + 已分配共享行）。"""
    ids = [int(x) for x in request.POST.getlist('row_ids') if str(x).isdigit()]
    if not ids:
        messages.error(request, '请先勾选要导出的记录。')
        return redirect('index')
    assigned = user_assigned_asin_codes(request.user)
    rows = list(
        user_dashboard_rows_qs(request.user)
        .filter(pk__in=ids)
        .select_related('user')
        .order_by('asin')
    )
    if not rows:
        messages.error(request, '没有可导出的数据。')
        return redirect('index')
    label_map = _assignment_label_map({normalize_asin(r.asin) for r in rows})
    headers = [
        'ASIN',
        '数据归属用户',
        '分配人',
        '去广告毛利率%',
        '去广告投产比%',
        '体量',
        '利润额￥',
        '(体量*利润额)$',
        '采购价￥',
        '(采购价+头程)￥',
        '广告难度%',
        '产品等级',
    ]
    wb = Workbook()
    ws = wb.active
    ws.append(headers)
    for r in rows:
        ak = normalize_asin(r.asin)
        assign_txt = label_map.get(ak) or ''
        ws.append(
            [
                r.asin,
                r.user.username if r.user_id else '',
                assign_txt,
                round(r.profit_margin, 2) if r.profit_margin is not None else '',
                round(r.ad_removed_roi, 2) if r.ad_removed_roi is not None else '',
                int(r.monthly_results) if r.monthly_results is not None else '',
                round(r.profit_per_order, 2) if r.profit_per_order is not None else '',
                round(r.monthly_sales_total, 2) if r.monthly_sales_total is not None else '',
                round(r.unit_purchase, 2) if r.unit_purchase is not None else '',
                round(r.head_actual_total, 2) if r.head_actual_total is not None else '',
                round(r.ranking_percent, 3) if r.ranking_percent is not None else '',
                r.product_grade or '',
            ]
        )
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    fn = f'asin_dashboard_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    return FileResponse(
        buf,
        as_attachment=True,
        filename=fn,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


@login_required
@require_POST
def excel_load(request):
    f = request.FILES.get('file')
    if not f:
        return JsonResponse({'ok': False, 'error': '未选择文件'}, status=400)
    if not f.name.lower().endswith(('.xlsx', '.xlsm')):
        return JsonResponse({'ok': False, 'error': '请上传 .xlsx 或 .xlsm 文件'}, status=400)

    try:
        raw = f.read()
        wb = load_workbook(BytesIO(raw), read_only=True, data_only=True)
    except Exception as e:
        return JsonResponse({'ok': False, 'error': f'无法读取 Excel：{e}'}, status=400)

    ws = wb.active
    sheet_title = ws.title
    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append([_cell_to_str(c) for c in row])
    wb.close()

    if not rows:
        rows = [['']]

    return JsonResponse({'ok': True, 'sheet_name': sheet_title, 'rows': rows})


def _cell_to_str(value):
    if value is None:
        return ''
    return str(value)


@login_required
@require_POST
def excel_save(request):
    try:
        payload = json.loads(request.body.decode('utf-8'))
    except json.JSONDecodeError:
        return JsonResponse({'ok': False, 'error': '无效的 JSON'}, status=400)

    rows = payload.get('rows')
    if not isinstance(rows, list):
        return JsonResponse({'ok': False, 'error': '缺少 rows 数组'}, status=400)

    wb = Workbook()
    ws = wb.active
    for row in rows:
        if not isinstance(row, list):
            row = [row]
        ws.append([_cell_to_str(c) for c in row])

    out = Path(settings.MEDIA_ROOT) / 'excel_exports'
    out.mkdir(parents=True, exist_ok=True)
    filename = f'export_{uuid.uuid4().hex}.xlsx'
    path = out / filename
    wb.save(path)

    return FileResponse(
        open(path, 'rb'),
        as_attachment=True,
        filename='edited.xlsx',
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
